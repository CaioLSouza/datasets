# -*- coding: utf-8 -*-
"""
=======================================================================
 CONGELADOR DO HISTORICO  (roda UMA vez, na migracao)
=======================================================================

 Numero que ja foi publicado nao muda. Este script le a aba Base da PA
 Principal e grava os valores COMO ESTAO PUBLICADOS hoje, onda a onda.

 A partir dai o pipeline passa a funcionar assim:

   ondas ate `ultima_onda_publicada`  -> valor CONGELADO (o publicado)
   ondas novas                        -> valor CALCULADO do bruto

 Os merges de alias continuam valendo, e nao mexem em numero nenhum:
 como "Riscos geopoliticos" (ate 202603) e "Riscos geopoliticos/ Guerra"
 (a partir de 202604) vivem em ondas disjuntas, juntar as duas na mesma
 linha so emenda a serie - os valores de cada mes seguem identicos ao
 que foi publicado.

 SAIDA
   config/valores_publicados.csv   onda;chave;pct
       O congelado. E este arquivo que o pipeline le para nao
       recalcular o que ja foi publicado. Versione junto com o
       registro de perguntas.

   _saida/chaves_base.csv          linha;chave;texto;reconhecido_como
       O mapa de como cada linha da Base antiga foi interpretada.
       Nao e usado pelo pipeline -- serve para voce auditar o
       congelamento e para achar os blocos mortos da planilha velha.

 USO
   python congelar_historico.py "\\\\xpdocs\\...\\PA Principal.xlsx"
=======================================================================
"""

import csv
import datetime
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl

from pipeline import RAIZ, Registro, norm

IGNORAR = {"total", "media", "média", "resposta media", "resposta média", ""}


def _n_respostas(v):
    """Le a linha 2 da Base (numero de respondentes da onda).

    O formato mudou tres vezes ao longo dos anos:
      ate 2021   texto  -> '590 respostas'
      2022       vazio  -> None
      2023+      numero -> 203
    """
    if isinstance(v, (int, float)) and v:
        return float(v)
    if isinstance(v, str):
        m = re.search(r"\d[\d.,]*", v)
        if m:
            return float(m.group().replace(".", "").replace(",", "."))
    return None


def em_notebook():
    """Num notebook o sys.argv traz os argumentos do kernel do Jupyter
    (-f kernel-....json), nao o que voce digitou. Detectar isso evita um
    erro obscuro la na frente, quando o openpyxl tentar abrir o .json."""
    return any("ipykernel" in str(a) or str(a).endswith(".json")
               for a in sys.argv)


def _resolver_caminho(caminho, exemplo):
    if caminho is None:
        if em_notebook():
            sys.exit(
                "Rodando num notebook: o sys.argv aqui traz os argumentos do\n"
                "kernel do Jupyter, não o caminho que você quer.\n\n"
                "Passe o caminho direto na chamada:\n\n"
                f"    {exemplo}\n\n"
                "Repare no r antes das aspas — sem ele o Python interpreta as\n"
                "barras invertidas como código de escape."
            )
        if len(sys.argv) < 2:
            sys.exit(__doc__)
        caminho = sys.argv[1]
    pa = Path(caminho)
    if pa.suffix.lower() not in (".xlsx", ".xlsm"):
        sys.exit(f"Esperava um .xlsx e recebi:\n   {pa}")
    if not pa.exists():
        sys.exit(f"Não achei o arquivo:\n   {pa}\n\n"
                 "Confira se a rede está acessível.")
    return pa


def main(caminho=None):
    """caminho: passe explicitamente se estiver num notebook."""
    pa = _resolver_caminho(
        caminho,
        'main(r"\\\\xpdocs\\Research\\Equities\\Estrategia\\Reports'
        '\\Pesquisa assessores\\PA Principal.xlsx")')
    reg = Registro(RAIZ / "config" / "perguntas.yaml")

    wb = openpyxl.load_workbook(pa, data_only=True)
    ws = wb["Base"]

    # ------------------------------------------------------------------
    #  Quais colunas sao ondas
    # ------------------------------------------------------------------
    #  A LINHA 4 (a data) e a chave, nao a linha 1.
    #
    #  O codigo AAAAMM da linha 1 so existe a partir de jul/2023 -- foi
    #  acrescentado quando a Raw Data comecou. As 40 colunas anteriores
    #  (fev/2020 a jun/2023) nao tem codigo nenhum, so a data. Ler pela
    #  linha 1 perderia 3 anos e meio de serie publicada.
    #
    #  A linha 3 nao serve: e um rotulo digitado a mao e tem erro (a
    #  coluna de mai/2021 esta escrita como "Edicao Mar/21").
    # ------------------------------------------------------------------
    achadas = {}        # coluna -> (onda, dia, rotulo)
    for c in range(4, 200):
        v4, v1 = ws.cell(4, c).value, ws.cell(1, c).value
        if isinstance(v4, datetime.datetime):
            onda, dia = v4.year * 100 + v4.month, v4.day
        elif isinstance(v1, (int, float)) and 201500 < v1 < 210000:
            onda, dia = int(v1), 1
        else:
            continue
        achadas[c] = (onda, dia, str(ws.cell(3, c).value or "")[:40])

    # ------------------------------------------------------------------
    #  Duas colunas no mesmo mes = duas pesquisas, nao um erro
    # ------------------------------------------------------------------
    #  Fev/2020 tem a edicao regular (dia 01) e uma "Edicao Coronavirus"
    #  (dia 15), fora do ciclo. A chave de onda e AAAAMM e nao comporta
    #  as duas, entao fica a REGULAR -- a que faz parte da serie mensal.
    #  A extraordinaria e descartada, com aviso: melhor perder um ponto
    #  fora de ciclo do que publicar ele no lugar do mes.
    # ------------------------------------------------------------------
    por_onda = {}
    for c, (onda, dia, rot) in achadas.items():
        if onda not in por_onda or dia < achadas[por_onda[onda]][1]:
            por_onda[onda] = c

    descartadas = [(c, achadas[c]) for c in achadas
                   if por_onda[achadas[c][0]] != c]
    if descartadas:
        print("AVISO: mais de uma coluna no mesmo mês. Fiquei com a do "
              "dia mais cedo (a edição regular).")
        print("Estas ficaram DE FORA do histórico congelado:")
        for c, (onda, dia, rot) in sorted(descartadas, key=lambda x: x[1]):
            print(f"   col {c}  {onda} dia {dia:>2}  {rot}")
        print()

    ondas = {c: achadas[c][0] for c in por_onda.values()}
    bases = {c: _n_respostas(ws.cell(2, c).value) for c in ondas}

    linhas_mapa, congelado, rankings = [], [], []
    atual, atual_safra, buffer = None, None, []
    stats = {"bloco": 0, "opcao": 0, "orfa": 0, "contagem": 0}

    def fechar_bloco():
        """Grava o bloco acumulado, decidindo % x contagem por COLUNA.

        A Base guarda a escala 0-10 em contagem num bloco e em % noutro.
        A decisao nao pode ser por celula (contagem de 1 pessoa seria
        confundida com 100%) nem pela soma (em multipla escolha os
        percentuais somam bem mais que 1).

        A regra que vale para os dois casos: percentual nunca passa de 1.
        Se o MAIOR valor da coluna dentro do bloco estoura 1, o bloco
        esta em contagem.
        """
        if not buffer:
            return
        for col, onda in ondas.items():
            vals = [(ch, mt, ws.cell(r, col).value) for r, ch, mt in buffer]
            nums = [v for _, _, v in vals if isinstance(v, (int, float))]
            if not nums:
                continue
            # Em jun/2023 a Base guardou o RANKING dos setores (1 a 14),
            # nao o percentual. Ranking dividido pelo nº de respondentes
            # viraria percentual inventado, entao e melhor nao congelar
            # nada: nao existe conversao honesta de posicao para %.
            #
            # A assinatura e inconfundivel -- os valores sao exatamente
            # a permutacao 1..N. Contagem real nunca cai nisso: com 178
            # respondentes e 14 setores os valores seriam 90, 86, 77...
            if (len(nums) >= 4
                    and all(float(v).is_integer() for v in nums)
                    and sorted(int(v) for v in nums) == list(range(1, len(nums) + 1))):
                rankings.append((onda, buffer[0][1].split("|")[0], len(nums)))
                continue

            divisor = 1.0
            if max(nums) > 1.05:                # bloco em contagem
                divisor = bases.get(col) or 0
                if not divisor:
                    continue
                stats["contagem"] += 1
            era_pct = divisor == 1.0
            for ch, mt, v in vals:
                if isinstance(v, (int, float)):
                    congelado.append((onda, ch, v / divisor, era_pct, mt))

    for r in range(1, ws.max_row + 1):
        c = ws.cell(r, 3).value
        txt = str(c).strip() if c not in (None, "") else ""
        if norm(txt) in IGNORAR:
            linhas_mapa.append((r, "", txt[:60], "—"))
            continue
        p, safra = reg.identificar(txt)
        if p:
            fechar_bloco()
            buffer = []
            atual, atual_safra = p, safra
            stats["bloco"] += 1
            marca = f" safra {safra}" if safra else ""
            linhas_mapa.append((r, "", txt[:60], f"cabeçalho de {p.id}{marca}"))
            continue
        if atual is None:
            linhas_mapa.append((r, "", txt[:60], "fora de bloco conhecido"))
            continue
        res = atual.resolver(txt)
        if not res or res[0][3] == "desconhecida":
            fechar_bloco()
            buffer = []
            atual = None
            stats["orfa"] += 1
            linhas_mapa.append((r, "", txt[:60], "não reconhecida — bloco encerrado"))
            continue

        oid, rot_pt, rot_en, _ = res[0]
        chave = (f"{atual.id}|{atual_safra}|{oid}" if atual_safra
                 else f"{atual.id}|{oid}")
        stats["opcao"] += 1
        linhas_mapa.append((r, chave, txt[:60], atual.id))

        # O rotulo viaja junto com o valor. Sem isso, alternativa que so
        # existiu antes de jul/2023 (as safras velhas do Ibovespa, por
        # exemplo) nao teria de quem herdar o texto na hora de montar o
        # agregado, e sumiria do historico.
        meta = (atual.id, atual.familia, atual_safra or "", oid,
                rot_pt, rot_en, atual.ordem_opcao.get(oid, 500))

        # Chave repetida dentro do mesmo bloco = comecou um sub-bloco.
        # A Base empilha, sem cabecalho no meio, a versao em CONTAGEM e a
        # versao em % da escala 0-10 (o "Resposta media" entre as duas
        # nao fecha nada). Sem este corte, os dois viram um bloco so e a
        # deteccao de escala se perde.
        if any(ch == chave for _, ch, _ in buffer):
            fechar_bloco()
            buffer = []

        buffer.append((r, chave, meta))

    fechar_bloco()

    # --------- grava ---------
    (RAIZ / "config").mkdir(exist_ok=True)
    (RAIZ / "_saida").mkdir(exist_ok=True)

    # Uma mesma chave pode vir de DUAS linhas da Base: o rotulo antigo e
    # o novo da alternativa renomeada (ex.: "Instabilidade politica" e
    # "Instabilidade politica/ Eleicoes"). Como as janelas sao disjuntas,
    # em cada onda so uma das duas linhas tem valor de verdade e a outra
    # esta zerada. Fico com a que tem valor.
    #  Regra de desempate, nesta ordem:
    #   1. valor nao-zerado ganha do zerado  (rotulo renomeado: em cada
    #      onda so uma das duas linhas tem valor)
    #   2. bloco que ja estava em % ganha do bloco em contagem  (a escala
    #      0-10 aparece nos dois formatos; o publicado como % e o que
    #      saiu no report)
    #  Sobra conflito de verdade so quando duas linhas em % discordam —
    #  ai provavelmente ha alias errado no perguntas.yaml.
    dest_val = RAIZ / "config" / "valores_publicados.csv"
    melhor, conflitos = {}, []
    for onda, chave, v, era_pct, mt in congelado:
        k = (onda, chave)
        if k not in melhor:
            melhor[k] = (v, era_pct, mt)
            continue
        atual_v, atual_pct, _ = melhor[k]
        if atual_v in (0, None) and v not in (0, None):
            melhor[k] = (v, era_pct, mt)
        elif v in (0, None):
            pass
        elif era_pct and not atual_pct:
            melhor[k] = (v, era_pct, mt)
        elif era_pct == atual_pct and abs(atual_v - v) > 0.0015:
            # Duas linhas em % discordando. Na Base, correcao entra como
            # linha NOVA logo abaixo da errada (foi o que aconteceu com
            # "Melhora na recuperacao economica global": a de cima tem um
            # ';' colado no rotulo e subconta). Fico com a de baixo e
            # aviso, para voce conferir e, se quiser, editar o CSV na mao.
            conflitos.append((onda, chave, atual_v, v))
            melhor[k] = (v, era_pct, mt)

    with open(dest_val, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(["onda", "chave", "pct", "q_id", "familia", "safra",
                    "opcao_id", "opcao_pt", "opcao_en", "ordem_opcao"])
        for (onda, chave), (v, _, mt) in sorted(melhor.items()):
            w.writerow([onda, chave, repr(v), *mt])
    vistos = melhor

    dest_map = RAIZ / "_saida" / "chaves_base.csv"
    with open(dest_map, "w", encoding="utf-8-sig", newline="") as fh:
        fh.write("linha;chave;texto_na_coluna_C;reconhecido_como\n")
        for r, ch, txt, como in linhas_mapa:
            fh.write(f"{r};{ch};{txt.replace(';', ',')};{como}\n")

    ondas_u = sorted({o for o, _ in melhor})
    print(f"Blocos reconhecidos ....... {stats['bloco']}")
    print(f"Linhas de alternativa ..... {stats['opcao']}")
    print(f"Linhas fora do registro ... {stats['orfa']}  (perguntas do mês antigas)")
    print(f"\nValores congelados ........ {len(vistos)}")
    print(f"Ondas cobertas ............ {len(ondas_u)}  ({ondas_u[0]}..{ondas_u[-1]})")
    if rankings:
        print(f"\nATENÇÃO: {len(rankings)} bloco(s) guardam RANKING, não "
              f"percentual — não dá para congelar.")
        print("Esses meses ficam sem esta pergunta no histórico. O dado "
              "publicado era\numa ordem de preferência (1º, 2º, 3º...), "
              "e não existe conversão honesta\npara percentual.")
        for onda, qid, n in sorted(set(rankings)):
            print(f"   {onda}  {qid:<22} ranking de 1 a {n}")

    if conflitos:
        print(f"\nATENÇÃO: {len(conflitos)} chave(s) com duas linhas na Base "
              f"publicando valores diferentes.")
        print("Fiquei com a linha de baixo (na Base, correção entra abaixo da")
        print("linha errada). Confira; se preferir a outra, edite o valor à mão")
        print("em config/valores_publicados.csv — é um CSV simples.")
        print(f"\n   {'onda':>7}  {'chave':<40} {'de cima':>9} {'DE BAIXO':>9}")
        for onda, chave, a, b in conflitos[:10]:
            print(f"   {onda:>7}  {chave:<40} {a:>9.4f} {b:>9.4f}")
    print(f"\nGerado: {dest_val}")
    print(f"Gerado: {dest_map}")
    print(f"\nAgora ponha em config/config.yaml:")
    print(f"    ultima_onda_publicada: {ondas_u[-1]}")
    print("A partir da onda seguinte, o pipeline calcula do bruto.")


if __name__ == "__main__":
    main()
