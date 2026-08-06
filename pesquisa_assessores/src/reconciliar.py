# -*- coding: utf-8 -*-
"""
=======================================================================
 RECONCILIADOR  (rode depois da migracao, e sempre que tiver duvida)
=======================================================================

 Compara, celula a celula, o agregado do pipeline contra os valores que
 a aba Base da PA Principal publica hoje. Serve para responder a
 pergunta que sempre aparece: "o numero mudou, mas mudou por que?"

 Classifica cada diferenca em tres baldes:

   BATEM EXATO           o pipeline reproduz o publicado
   DIVERGEM POR MERGE    de proposito: a Base mostra rotulo antigo e
                         novo como duas linhas, o pipeline soma numa
                         serie so (aliases / divide_em / texto livre)
   DIVERGEM SEM EXPLICAR  <- e aqui que voce olha

 USO
   python reconciliar.py "\\\\xpdocs\\...\\PA Principal.xlsx"

 O bloco 1 cobre as 76 ondas (fev/2020 em diante): mesmo sem dado bruto
 por tras, o valor congelado tem que reproduzir o publicado.

 O bloco 2 so cobre de jul/2023 em diante, que e de onde existe Raw
 Data. Antes disso nao ha o que recalcular.
=======================================================================
"""

import datetime
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
import yaml

from congelar_historico import _resolver_caminho  # cai no config.yaml sozinho
from pipeline import (RAIZ, Registro, agregar, carregar_congelados,
                      carregar_store, norm)

IGNORAR = {"total", "media", "média", "resposta media", "resposta média", ""}


def main(caminho=None):
    """caminho: passe explicitamente se estiver num notebook."""
    pa = _resolver_caminho(
        caminho,
        'main(r"\\\\xpdocs\\Research\\Equities\\Estrategia\\Reports'
        '\\Pesquisa assessores\\PA Principal.xlsx")')

    with open(RAIZ / "config" / "config.yaml", encoding="utf-8") as fh:
        cfg = yaml.safe_load(fh)
    reg = Registro(RAIZ / "config" / "perguntas.yaml")
    regs = carregar_store(RAIZ / cfg["caminhos"]["store"])
    if not regs:
        sys.exit("Store vazio. Rode o pipeline (ou --bootstrap) antes.")

    congelados, metas = carregar_congelados(
        RAIZ / "config" / "valores_publicados.csv")
    corte = cfg["parametros"].get("ultima_onda_publicada")

    # DUAS visoes, que respondem duas perguntas diferentes:
    #
    #  saida  = o que o pipeline realmente entrega (historico congelado).
    #           Tem que bater 100% com o publicado. Se nao bater, e bug.
    #  calc   = o que o recalculo do bruto diria. Serve so para auditoria:
    #           e onde aparecem os erros do processo antigo.
    agg = agregar(regs, "onda", congelados, corte, metas, reg)
    saida = {(a["onda"], a["q_id"], a["opcao_id"]): a["pct"] for a in agg}
    pct = {(a["onda"], a["q_id"], a["opcao_id"]): a["pct_calculado"]
           for a in agg if a.get("pct_calculado") is not None}
    cnt = {(a["onda"], a["q_id"], a["opcao_id"]): a["n"] for a in agg}

    wb = openpyxl.load_workbook(pa, data_only=True)
    ws = wb["Base"]

    # A data da linha 4 e a chave da coluna, nao o codigo da linha 1 --
    # esse so existe de jul/2023 em diante. Mesma regra do congelador.
    cols, vistas = {}, {}
    for c in range(4, 130):
        v4, v1 = ws.cell(4, c).value, ws.cell(1, c).value
        if isinstance(v4, datetime.datetime):
            onda, dia = v4.year * 100 + v4.month, v4.day
        elif isinstance(v1, (int, float)) and 201500 < v1 < 210000:
            onda, dia = int(v1), 1
        else:
            continue
        if onda not in vistas or dia < vistas[onda][1]:
            vistas[onda] = (c, dia)
    cols = {c: o for o, (c, _) in vistas.items()}

    atual, comp_idx, sem_match = None, {}, set()
    for r in range(1, ws.max_row + 1):
        c = ws.cell(r, 3).value
        txt = str(c).strip() if c not in (None, "") else ""
        if norm(txt) in IGNORAR:
            continue
        p, _ = reg.identificar(txt)
        if p:
            atual = p
            continue
        if atual is None:
            continue
        res = atual.resolver(txt)
        if not res or res[0][3] == "desconhecida":
            sem_match.add((atual.id, txt[:55]))
            atual = None
            continue
        oid = res[0][0]
        for col, onda in cols.items():
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                # Duas linhas da Base podem cair na mesma chave: o rotulo
                # antigo e o novo de uma alternativa renomeada. Em cada
                # onda so uma tem valor; a outra fica zerada. Fico com a
                # que tem valor, senao a comparacao acusaria falso erro.
                k = (onda, atual.id, oid)
                if k not in comp_idx or (comp_idx[k][0] == 0 and v != 0):
                    comp_idx[k] = (float(v), txt)

    comp = [(o, q, oid, v, t) for (o, q, oid), (v, t) in comp_idx.items()]

    # opcoes que o registro mescla de proposito
    mescladas = set()
    for p in reg.perguntas:
        for oid, n in Counter(p._idx.values()).items():
            if n > 2:                       # mais que pt+en => tem alias
                mescladas.add((p.id, oid))
        mescladas |= {(p.id, o) for o in set(p.grupos_livres.values())}
        mescladas |= {(p.id, o) for ids in p.divide_em.values() for o in ids}

    # ---------------------------------------------------------------
    # 1) O TESTE QUE VALE: a saida do pipeline == o publicado?
    # ---------------------------------------------------------------
    bate = fura = 0
    furos = []
    for onda, q, oid, bv, rot in comp:
        sv, cv = saida.get((onda, q, oid)), cnt.get((onda, q, oid))
        if sv is None:
            continue
        if abs(bv - sv) <= 0.0015 or (cv is not None and abs(bv - cv) <= 0.5):
            bate += 1
        else:
            fura += 1
            furos.append((abs(bv - sv), onda, q, rot, bv, sv))

    tot1 = max(bate + fura, 1)
    ondas_all = sorted(cols.values())
    print(f"Ondas na Base ..... {len(ondas_all)}  "
          f"({ondas_all[0]}..{ondas_all[-1]})")
    print(f"  com Raw Data .... {len([o for o in ondas_all if o >= 202307])}")
    print(f"Pares comparados .. {len(comp)}")
    print(f"Congelado ate ..... {corte}\n")
    print("=" * 66)
    print(" 1) SAIDA DO PIPELINE  x  PUBLICADO   (tem que dar 100%)")
    print("=" * 66)
    print(f"  IDENTICOS ... {bate:>5}  ({bate/tot1:6.2%})")
    print(f"  DIFERENTES .. {fura:>5}  ({fura/tot1:6.2%})")
    if furos:
        print("\n  ATENCAO: o historico saiu diferente do publicado.")
        furos.sort(reverse=True)
        for d, onda, q, rot, bv, sv in furos[:15]:
            print(f"   {onda} {q:<22}{rot[:32]:<32} "
                  f"Base={bv:7.4f}  saida={sv:7.4f}  dif={d:.4f}")
    else:
        print("\n  OK - nenhum numero publicado foi alterado.")

    # ---------------------------------------------------------------
    # 2) AUDITORIA: o que o recalculo do bruto diria
    # ---------------------------------------------------------------
    ok = merge = ruim = falta = 0
    piores, porq = [], Counter()
    for onda, q, oid, bv, rot in comp:
        if onda < 202307:          # sem Raw Data, nao ha o que recalcular
            continue
        pv, cv = pct.get((onda, q, oid)), cnt.get((onda, q, oid))
        if pv is None:
            if bv > 0.0005:
                falta += 1
            continue
        if abs(bv - pv) <= 0.0015 or (cv is not None and abs(bv - cv) <= 0.5):
            ok += 1
        elif (q, oid) in mescladas:
            merge += 1
        else:
            ruim += 1
            porq[q] += 1
            piores.append((abs(bv - pv), onda, q, rot, bv, pv))

    tot = max(ok + merge + ruim, 1)
    print("\n" + "=" * 66)
    print(" 2) AUDITORIA: o que o recalculo do bruto diria")
    print("    (nao vai pro grafico enquanto a onda estiver congelada)")
    print("=" * 66)
    print(f"  Bateria com o publicado ..... {ok:>5}  ({ok/tot:6.2%})")
    print(f"  Difere por merge de alias ... {merge:>5}  ({merge/tot:6.2%})")
    print(f"  Difere sem explicacao ....... {ruim:>5}  ({ruim/tot:6.2%})")
    print(f"  Publicado sem recalculo ..... {falta:>5}")

    if piores:
        print(f"\n  Por pergunta: {dict(porq)}")
        print("\n  Maiores diferencas:")
        piores.sort(reverse=True)
        for d, onda, q, rot, bv, mv in piores[:15]:
            print(f"   {onda} {q:<22}{rot[:32]:<32} "
                  f"Base={bv:7.4f}  recalc={mv:7.4f}  dif={d:.4f}")
        print("\n  A maioria destas e o COUNTIF da Base subcontando quando")
        print("  falta ';' no fim da celula. O recalculo esta certo, mas o")
        print("  publicado fica como esta - so as ondas novas usam o calculo.")

    if sem_match:
        print(f"\n  Blocos da Base fora do registro ({len(sem_match)}) — "
              f"perguntas do mes antigas, normal:")
        for q, t in sorted(sem_match)[:12]:
            print(f"   [{q}] {t}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
