# -*- coding: utf-8 -*-
"""
=======================================================================
 PIPELINE DA PESQUISA DE ASSESSORES XP
=======================================================================

 Le o export do Forms, normaliza contra o registro canonico de
 perguntas (config/perguntas.yaml) e gera duas planilhas:

   PA Base Historica.xlsx   todo o historico, formato longo + agregado
   PA Base Mes Atual.xlsx   so o mes do report, em enderecos fixos

 Este script nao abre planilha de report nenhuma. Quem consome as duas
 acima, via Power Query, e a PA Report.xlsx -- montada uma vez so (ver
 powerquery/INSTRUCOES.md) e dai em diante apenas atualizada.

 A separacao e de proposito: os 22 links OLE dos PPTs apontam para a
 PA Report.xlsx, que nunca e reescrita. Se o pipeline regerasse ela
 todo mes, os links quebrariam.

 -------------------------------------------------------------------
 USO
 -------------------------------------------------------------------
   python pipeline.py
       processa o .xlsx mais recente da pasta input_forms

   python pipeline.py --arquivo "C:\\...\\pesquisa.xlsx" --onda 202608
       processa um arquivo especifico

   python pipeline.py --conferir
       so valida e mostra o diagnostico, nao escreve nada

   python pipeline.py --bootstrap "PA Principal.xlsx"
       carga inicial: importa as 37 ondas da aba Raw Data existente
       (as 39 ondas anteriores, que so existem na aba Base, entram
        pelo congelamento -- ver congelar_historico.py)

 Codigo de saida: 0 = ok, 1 = erro (alternativa nova nao declarada,
 pergunta nao reconhecida, etc). Use no agendador para ser avisado.
=======================================================================
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import yaml

try:
    import openpyxl
except ImportError:
    sys.exit("Falta o openpyxl. Rode:  pip install openpyxl pyyaml")


RAIZ = Path(__file__).resolve().parent.parent


# =====================================================================
#  Normalizacao de texto
# =====================================================================

def norm(s) -> str:
    """Minusculo, sem acento, sem espaco repetido.

    Tudo que e comparacao de texto no pipeline passa por aqui. E o que
    faz 'Riscos geopoliticos' casar com 'Riscos geopolíticos' e
    'os maiores risco' casar com 'o maiores riscos' (via regex).
    """
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s.lower()


_NUM = re.compile(r"\d[\d.,]*")


def _num(txt: str) -> float:
    """Converte numero em formato brasileiro. '13,75' -> 13.75, '1.500' -> 1500."""
    t = txt.strip()
    if "," in t:                      # virgula = decimal, ponto = milhar
        t = t.replace(".", "").replace(",", ".")
    elif re.fullmatch(r"\d{1,3}(\.\d{3})+", t):   # so ponto, padrao de milhar
        t = t.replace(".", "")
    return float(t)


_FAIXA_INICIO = re.compile(
    r"^([<>=≤≥]|entre|acima|abaixo|ate|mais de|menos de|superior|inferior|"
    r"menor|de \d|\d)")


def parse_faixa(rotulo: str):
    """'Entre 190 e 200 mil pontos' -> (190000.0, 200000.0)

    Reconhece: Entre X e Y / X a Y / Acima de X / Abaixo de X, com ou
    sem 'mil', com ou sem '%', com ou sem a preposicao 'de'.

    E este parser que faz a re-grade de faixas passar sem quebrar nada:
    quando as faixas do Ibovespa de 10 mil pontos viraram faixas de 20
    mil em abr/2026, nenhuma linha de configuracao precisou mudar.

    Devolve None para texto que nao TEM cara de faixa. Sem essa guarda,
    a frase inteira de uma pergunta ("Por favor, digite abaixo qual o
    nivel ... ao final de 2021.") viraria a faixa 2021-2021, e o valor
    da linha -- que e a media digitada, 141.964 pontos -- entraria no
    historico como se fosse um percentual.
    """
    t = norm(rotulo)
    if len(t) > 60 or not _FAIXA_INICIO.match(t):
        return None
    nums = [_num(m.group()) for m in _NUM.finditer(t)]
    if not nums:
        return None
    if "mil" in t:
        nums = [n * 1000 if n < 1000 else n for n in nums]

    if re.match(r"^(acima|mais de|superior)", t):
        return (nums[0], math.inf)
    if re.match(r"^(abaixo|ate|menos de|inferior|menor)", t):
        return (-math.inf, nums[0])
    if len(nums) >= 2:
        return (min(nums[0], nums[1]), max(nums[0], nums[1]))
    return (nums[0], nums[0])


def rotulo_faixa_en(rotulo_pt: str, faixa) -> str:
    """Traducao automatica dos rotulos de faixa (Ibovespa / Selic / %)."""
    lo, hi = faixa
    pct = "%" in rotulo_pt
    def fmt(v):
        if pct:
            return f"{v:g}%"
        return f"{v/1000:g}k"
    if hi == math.inf:
        return f"Above {fmt(lo)}"
    if lo == -math.inf:
        return f"Below {fmt(hi)}"
    return f"{fmt(lo)} to {fmt(hi)}"


# =====================================================================
#  Registro canonico
# =====================================================================

class Opcao:
    __slots__ = ("id", "pt", "en", "desde", "ate")

    def __init__(self, d):
        self.id = d["id"]
        self.pt = d.get("pt", d["id"])
        self.en = d.get("en", self.pt)
        self.desde = d.get("desde_onda")
        self.ate = d.get("ate_onda")


class Pergunta:
    def __init__(self, d, familia):
        self.id = d["id"]
        self.familia = familia            # 'recorrente' | 'safra' | 'mes'
        self.ordem = d.get("ordem", 999)
        self.tipo = d.get("tipo", "unica")
        self.pt = d.get("pt", d["id"])
        self.en = d.get("en", self.pt)
        self.parser_faixa = d.get("parser_faixa", False) or self.tipo == "faixa"
        self.substitui_no_mes = d.get("substitui_no_mes", False)
        self.safra_padrao = d.get("safra_padrao")
        self.destaque_capa = d.get("destaque_capa", False)
        self.slot = d.get("slot")

        self.match = [re.compile(p) for p in d.get("match", [])]

        self.opcoes = [Opcao(o) for o in d.get("opcoes", [])]
        self.ordem_opcao = {o.id: i for i, o in enumerate(self.opcoes)}

        # indice de lookup: texto normalizado -> id da opcao
        self._idx = {}
        for o in self.opcoes:
            self._idx[norm(o.pt)] = o.id
            self._idx[norm(o.en)] = o.id
        for o_d in d.get("opcoes", []):
            for al in o_d.get("aliases", []) or []:
                self._idx[norm(al)] = o_d["id"]

        # valor sujo que na verdade sao 2+ opcoes
        self.divide_em = {norm(k): v for k, v in (d.get("divide_em") or {}).items()}

        # write-ins agrupados (nao viram opcao, so limpam o log)
        self.grupos_livres = {}
        for gid, textos in (d.get("agrupa_texto_livre") or {}).items():
            for t in textos:
                self.grupos_livres[norm(t)] = gid

    # -----------------------------------------------------------------
    def casa(self, cabecalho: str):
        """Retorna (True, safra) se este cabecalho e desta pergunta."""
        h = norm(cabecalho)
        for rx in self.match:
            m = rx.search(h)
            if m:
                safra = None
                if m.groupdict().get("safra"):
                    safra = m.group("safra")
                elif self.familia == "safra":
                    safra = self.safra_padrao
                return True, safra
        return False, None

    # -----------------------------------------------------------------
    def resolver(self, valor: str):
        """valor bruto -> lista de (opcao_id, rotulo_pt, rotulo_en, status)

        status: 'ok' | 'faixa' | 'livre' | 'desconhecida'
        """
        v = str(valor).strip().rstrip(";").strip()
        if not v:
            return []
        n = norm(v)

        if n in self.divide_em:
            saida = []
            for oid in self.divide_em[n]:
                o = next((x for x in self.opcoes if x.id == oid), None)
                saida.append((oid, o.pt if o else oid, o.en if o else oid, "ok"))
            return saida

        if n in self._idx:
            oid = self._idx[n]
            o = next(x for x in self.opcoes if x.id == oid)
            return [(oid, o.pt, o.en, "ok")]

        if self.parser_faixa:
            f = parse_faixa(v)
            if f:
                lo, hi = f
                oid = "f_%s_%s" % (
                    "inf" if lo == -math.inf else f"{lo:g}",
                    "inf" if hi == math.inf else f"{hi:g}",
                )
                return [(oid, v, rotulo_faixa_en(v, f), "faixa")]

        if self.tipo == "escala":
            try:
                float(n.replace(",", "."))
                return [(v, v, v, "ok")]
            except ValueError:
                pass

        if n in self.grupos_livres:
            return [(self.grupos_livres[n], v, v, "livre")]

        # Pergunta do mes nao tem registro por definicao: as alternativas
        # sao o que vierem. Cada valor distinto vira sua propria opcao.
        if self.familia == "mes":
            return [(re.sub(r"[^a-z0-9]+", "_", n)[:60] or "vazio", v, v, "ok")]

        return [("_outros", v, v, "desconhecida")]


class Registro:
    def __init__(self, caminho: Path):
        with open(caminho, encoding="utf-8") as fh:
            d = yaml.safe_load(fh)
        self.bruto = d
        self.regras = d.get("regras", {})
        self.limite = self.regras.get("limite_alerta_desconhecida", 0.05)
        self.minimo = self.regras.get("minimo_respostas_desconhecida", 5)
        self.meta = {k: [norm(x) for x in v]
                     for k, v in (d.get("colunas_meta") or {}).items()}

        self.perguntas = []
        for chave, fam in (("perguntas", "recorrente"),
                           ("perguntas_safra", "safra"),
                           ("perguntas_mes", "mes")):
            for pd_ in (d.get(chave) or []):
                self.perguntas.append(Pergunta(pd_, fam))
        self.por_id = {p.id: p for p in self.perguntas}

    def identificar(self, cabecalho: str):
        """cabecalho do Forms -> (Pergunta, safra) ou (None, None)."""
        for p in self.perguntas:
            ok, safra = p.casa(cabecalho)
            if ok:
                return p, safra
        return None, None

    def eh_meta(self, cabecalho: str):
        h = norm(cabecalho)
        for papel, nomes in self.meta.items():
            if h in nomes:
                return papel
        return None


# =====================================================================
#  Leitura do export do Forms
# =====================================================================

def ler_planilha(caminho: Path, aba: str | None = None):
    """Retorna (cabecalhos, linhas) da primeira aba (ou da aba pedida)."""
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    ws = wb[aba] if aba else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    cab = [c if c is not None else "" for c in next(it)]
    linhas = []
    for r in it:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        linhas.append(list(r) + [None] * (len(cab) - len(r)))
    wb.close()
    return cab, linhas


def deduzir_onda(cab, linhas, registro: Registro):
    """Descobre AAAAMM pela coluna 'Survey' ou pela data de conclusao."""
    for i, h in enumerate(cab):
        if registro.eh_meta(h) == "onda":
            vals = [r[i] for r in linhas if r[i] is not None]
            if vals:
                return int(str(Counter(vals).most_common(1)[0][0])[:6])
    for i, h in enumerate(cab):
        if registro.eh_meta(h) == "fim":
            meses = []
            for r in linhas:
                v = r[i]
                if isinstance(v, datetime):
                    meses.append(v.year * 100 + v.month)
                elif isinstance(v, str) and len(v) >= 7:
                    try:
                        meses.append(int(v[:4]) * 100 + int(v[5:7]))
                    except ValueError:
                        pass
            if meses:
                return Counter(meses).most_common(1)[0][0]
    return None


# =====================================================================
#  Normalizacao para formato longo
# =====================================================================

class Diagnostico:
    def __init__(self):
        self.novas_perguntas = []      # (cabecalho, n_respostas)
        self.desconhecidas = defaultdict(Counter)   # q_id -> Counter(valor)
        self.livres = defaultdict(Counter)
        self.faixas_novas = defaultdict(set)
        self.perguntas_ausentes = []
        self.erros = []
        self.avisos = []


def normalizar(cab, linhas, registro: Registro, onda: int, origem: str):
    """Export do Forms -> lista de dicts (grao: respondente x pergunta x opcao)."""
    diag = Diagnostico()
    n_total = len(linhas)

    col_id = None
    for i, h in enumerate(cab):
        if registro.eh_meta(h) == "id":
            col_id = i
            break

    # mapeia cada coluna
    mapa = {}          # indice -> (Pergunta, safra)
    novas = []         # colunas sem match = pergunta do mes
    for i, h in enumerate(cab):
        if not str(h).strip():
            continue
        if registro.eh_meta(h):
            continue
        p, safra = registro.identificar(h)
        if p:
            mapa[i] = (p, safra, h)
        else:
            preenchidas = sum(1 for r in linhas
                              if r[i] is not None and str(r[i]).strip() != "")
            if preenchidas:
                novas.append((i, h, preenchidas))

    # perguntas do mes -> slots, na ordem em que aparecem no Forms
    for slot, (i, h, cnt) in enumerate(novas, start=1):
        diag.novas_perguntas.append((h, cnt))
        qid = "q_mes_%d" % slot
        pm = Pergunta({"id": qid, "pt": str(h).strip(), "tipo": "multipla",
                       "ordem": 900 + slot, "slot": slot}, "mes")
        mapa[i] = (pm, None, h)

    # recorrentes esperadas que nao vieram
    for p in registro.perguntas:
        if p.familia == "recorrente" and p.id not in {m[0].id for m in mapa.values()}:
            diag.perguntas_ausentes.append(p.id)

    registros = []
    respondentes = defaultdict(set)     # q_id -> set(resp_id)

    for nl, r in enumerate(linhas):
        rid = r[col_id] if col_id is not None and r[col_id] is not None else nl + 1
        rid = f"{onda}-{rid}"
        for i, (p, safra, cabec) in mapa.items():
            v = r[i]
            if v is None or str(v).strip() == "":
                continue
            respondentes[p.id].add(rid)

            partes = [x for x in str(v).split(";") if x.strip()] \
                if p.tipo in ("multipla",) or ";" in str(v) else [str(v)]

            for parte in partes:
                for oid, rot_pt, rot_en, status in p.resolver(parte):
                    if status == "desconhecida":
                        diag.desconhecidas[p.id][parte.strip()] += 1
                    elif status == "livre":
                        diag.livres[p.id][parte.strip()] += 1
                    elif status == "faixa":
                        diag.faixas_novas[p.id].add(rot_pt)

                    val_num = None
                    if p.tipo == "escala":
                        try:
                            val_num = float(str(parte).replace(",", "."))
                        except ValueError:
                            pass
                    elif p.parser_faixa:
                        f = parse_faixa(parte)
                        if f:
                            lo, hi = f
                            if lo == -math.inf:
                                val_num = hi
                            elif hi == math.inf:
                                val_num = lo
                            else:
                                val_num = (lo + hi) / 2

                    registros.append({
                        "onda": onda,
                        "resp_id": rid,
                        "q_id": p.id,
                        "familia": p.familia,
                        "safra": safra or "",
                        "pergunta_pt": p.pt,
                        "pergunta_en": p.en,
                        "opcao_id": oid,
                        "opcao_pt": rot_pt,
                        "opcao_en": rot_en,
                        "ordem_pergunta": p.ordem,
                        "ordem_opcao": p.ordem_opcao.get(oid, 500),
                        "valor_num": val_num if val_num is not None else "",
                        "valor_bruto": str(parte).strip(),
                        "origem": origem,
                    })

    # avaliacao das desconhecidas contra o limite
    for qid, cont in diag.desconhecidas.items():
        base = max(len(respondentes[qid]), 1)
        for valor, n in cont.items():
            share = n / base
            msg = (f"[{qid}] alternativa nao declarada: {valor!r} "
                   f"({n} resp., {share:.1%} da pergunta)")
            if share >= registro.limite and n >= registro.minimo:
                diag.erros.append(msg)
            else:
                diag.avisos.append(msg)

    return registros, diag, n_total, respondentes


# =====================================================================
#  Store (fonte da verdade, append-only)
# =====================================================================

CAMPOS = ["onda", "resp_id", "q_id", "familia", "safra", "pergunta_pt",
          "pergunta_en", "opcao_id", "opcao_pt", "opcao_en",
          "ordem_pergunta", "ordem_opcao", "valor_num", "valor_bruto", "origem"]


def carregar_store(caminho: Path):
    if not caminho.exists():
        return []
    with open(caminho, encoding="utf-8", newline="") as fh:
        return list(csv.DictReader(fh))


def salvar_store(caminho: Path, regs):
    caminho.parent.mkdir(parents=True, exist_ok=True)
    with open(caminho, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=CAMPOS)
        w.writeheader()
        w.writerows(regs)


def mesclar(store, novos, onda):
    """Substitui a onda inteira. Rodar duas vezes o mesmo mes e seguro."""
    mantidos = [r for r in store if int(r["onda"]) != int(onda)]
    return mantidos + novos


# =====================================================================
#  Agregacao
# =====================================================================

def carregar_congelados(caminho: Path):
    """Valores COMO FORAM PUBLICADOS, gerados por congelar_historico.py.

    -> (valores, metas)
       valores  {(onda, chave): pct}
       metas    {chave: {q_id, familia, safra, opcao_id, opcao_pt, ...}}

    A metadata viaja junto porque a Base cobre 76 ondas (fev/2020 em
    diante) e a Raw Data so 37 (jul/2023 em diante). Alternativa que so
    existiu nesse intervalo -- as safras velhas do Ibovespa, por exemplo
    -- nao aparece em onda nenhuma com dado bruto, entao nao teria de
    quem herdar o rotulo.
    """
    if not caminho.exists():
        return {}, {}
    valores, metas = {}, {}
    with open(caminho, encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh, delimiter=";"):
            try:
                valores[(int(r["onda"]), r["chave"])] = float(r["pct"])
            except (ValueError, KeyError):
                continue
            if r["chave"] not in metas and r.get("q_id"):
                metas[r["chave"]] = r
    return valores, metas


def agregar(regs, denominador="respondentes", congelados=None, ate_onda=None,
            metas=None, registro=None):
    """-> lista de dicts (onda, q_id, opcao_id, n, base, pct)

    Numero publicado nao muda: para as ondas ate `ate_onda`, o `pct` que
    sai e o CONGELADO (o que foi ao ar). O recalculo do bruto continua
    disponivel em `pct_calculado`, para auditoria - so nao manda no
    grafico.
    """
    congelados = congelados or {}
    sel = defaultdict(set)     # (onda,q_id,opcao_id) -> set(resp)
    resp_q = defaultdict(set)  # (onda,q_id) -> set(resp)
    resp_onda = defaultdict(set)
    info = {}
    for r in regs:
        o, q, oid, rid = int(r["onda"]), r["q_id"], r["opcao_id"], r["resp_id"]
        sel[(o, q, oid)].add(rid)
        resp_q[(o, q)].add(rid)
        resp_onda[o].add(rid)
        info[(q, oid)] = r

    saida = []
    chaves_calc = set()
    for (o, q, oid), s in sel.items():
        base = len(resp_q[(o, q)]) if denominador == "respondentes" else len(resp_onda[o])
        r = info[(q, oid)]
        # A safra entra na chave: o Ibovespa tem um bloco de faixas por
        # ano-alvo, todos com os mesmos rotulos ("Entre 120 e 130 mil
        # pontos" existe em 2023, 2024, 2025 e 2026). Sem a safra na
        # chave, as quatro safras colidiriam numa serie so.
        chave = f"{q}|{r['safra']}|{oid}" if r["safra"] else f"{q}|{oid}"
        calc = len(s) / base if base else 0.0
        congelado = congelados.get((o, chave))
        usa_congelado = (ate_onda is not None and o <= ate_onda
                         and congelado is not None)
        chaves_calc.add((o, chave))
        saida.append({
            "onda": o,
            "q_id": q,
            "familia": r["familia"],
            "safra": r["safra"],
            "pergunta_pt": r["pergunta_pt"],
            "pergunta_en": r["pergunta_en"],
            "opcao_id": oid,
            "opcao_pt": r["opcao_pt"],
            "opcao_en": r["opcao_en"],
            "ordem_pergunta": int(r["ordem_pergunta"]),
            "ordem_opcao": int(r["ordem_opcao"]),
            "chave": chave,
            "n": len(s),
            "base": base,
            "pct": congelado if usa_congelado else calc,
            "pct_calculado": calc,
            "fonte": "publicado" if usa_congelado else "calculado",
        })

    # Valor publicado que o recalculo nao reproduz entra assim mesmo,
    # senao a serie historica perderia um ponto que ja foi ao ar. Sao
    # dois casos, e a maioria e o segundo:
    #
    #   a) a alternativa existe em outra onda com dado bruto
    #      -> herda dali os rotulos
    #   b) a alternativa so existiu antes de jul/2023, onde nao ha Raw
    #      Data nenhuma (39 das 76 ondas)
    #      -> os rotulos vem do proprio congelamento
    metas = metas or {}
    if ate_onda is not None:
        rot = {}
        for a in saida:
            rot.setdefault(a["chave"], a)
        for (o, chave), v in congelados.items():
            if o > ate_onda or (o, chave) in chaves_calc:
                continue
            m = rot.get(chave)
            if m:
                saida.append({**m, "onda": o, "n": None, "base": None,
                              "pct": v, "pct_calculado": None,
                              "fonte": "publicado"})
                continue
            md = metas.get(chave)
            if not md:
                continue
            p = registro.por_id.get(md["q_id"]) if registro else None
            saida.append({
                "onda": o,
                "q_id": md["q_id"],
                "familia": md.get("familia", "recorrente"),
                "safra": md.get("safra", ""),
                "pergunta_pt": p.pt if p else md["q_id"],
                "pergunta_en": p.en if p else md["q_id"],
                "opcao_id": md["opcao_id"],
                "opcao_pt": md["opcao_pt"],
                "opcao_en": md.get("opcao_en") or md["opcao_pt"],
                "ordem_pergunta": p.ordem if p else 999,
                "ordem_opcao": int(md.get("ordem_opcao") or 500),
                "chave": chave,
                "n": None, "base": None,
                "pct": v, "pct_calculado": None,
                "fonte": "publicado",
            })
    saida.sort(key=lambda x: (x["onda"], x["ordem_pergunta"], x["q_id"],
                              x["ordem_opcao"], -x["pct"]))
    return saida


def checar_paineis(agg, onda, registro, lp):
    """Alternativa que nao cabe no painel viraria grafico truncado.

    E o pior tipo de erro: o grafico sai bonito, com o rodape faltando,
    e vai para o report publicado sem ninguem notar. Melhor parar a
    rodada e obrigar a decisao.
    """
    por_q = defaultdict(list)
    for a in agg:
        if a["onda"] == onda and a["familia"] != "mes":
            por_q[a["q_id"]].append(a)

    problemas = []
    for p in registro.perguntas:
        if p.familia not in ("recorrente", "safra"):
            continue
        opcoes = por_q.get(p.id, [])
        if len(opcoes) > lp:
            fora = sorted(opcoes, key=lambda x: (x["ordem_opcao"], -x["pct"]))[lp:]
            problemas.append(
                f"[{p.id}] {len(opcoes)} alternativas, o painel reserva {lp}. "
                f"Ficariam de fora: "
                + ", ".join(d["opcao_pt"][:40] for d in fora))
    return problemas


def medias(regs):
    """Media das perguntas de escala e faixa (Ibovespa, Selic, sentimento)."""
    acc = defaultdict(list)
    for r in regs:
        if r["valor_num"] not in ("", None):
            acc[(int(r["onda"]), r["q_id"], r["safra"])].append(float(r["valor_num"]))
    return [{"onda": o, "q_id": q, "safra": s, "n": len(v),
             "media": sum(v) / len(v)}
            for (o, q, s), v in sorted(acc.items())]


# =====================================================================
#  Escrita das planilhas
# =====================================================================

def _aba(wb, nome, cabec, linhas, largura=None):
    ws = wb.create_sheet(nome)
    ws.append(cabec)
    for c in ws[1]:
        c.font = openpyxl.styles.Font(bold=True)
    ws.freeze_panes = "A2"
    for l in linhas:
        ws.append(l)
    for i, w in enumerate(largura or [], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    return ws


def escrever_historica(caminho: Path, regs, agg, med, registro: Registro):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # `pct` e o que vai pro grafico: publicado no historico, calculado nas
    # ondas novas. `pct_calculado` fica ao lado so para auditoria.
    _aba(wb, "agregado",
         ["onda", "q_id", "familia", "safra", "pergunta_pt", "pergunta_en",
          "opcao_id", "opcao_pt", "opcao_en", "chave", "ordem_pergunta",
          "ordem_opcao", "n", "base", "pct", "pct_calculado", "fonte"],
         [[a["onda"], a["q_id"], a["familia"], a["safra"], a["pergunta_pt"],
           a["pergunta_en"], a["opcao_id"], a["opcao_pt"], a["opcao_en"],
           a["chave"], a["ordem_pergunta"], a["ordem_opcao"],
           a["n"], a["base"], a["pct"], a.get("pct_calculado"),
           a.get("fonte", "calculado")] for a in agg],
         [9, 20, 12, 8, 50, 50, 18, 45, 45, 34, 8, 8, 7, 7, 9, 13, 11])

    # ---- matriz: chave x onda. E deste retangulo que a aba Base le. ----
    ondas = sorted({a["onda"] for a in agg})
    chaves, meta_chave = [], {}
    for a in agg:
        if a["chave"] not in meta_chave:
            chaves.append(a["chave"])
            meta_chave[a["chave"]] = a
    idx = {(a["chave"], a["onda"]): a["pct"] for a in agg}
    _aba(wb, "matriz",
         ["chave", "q_id", "opcao_id", "opcao_pt", "opcao_en"] + [str(o) for o in ondas],
         [[k, meta_chave[k]["q_id"], meta_chave[k]["opcao_id"],
           meta_chave[k]["opcao_pt"], meta_chave[k]["opcao_en"]]
          + [idx.get((k, o), None) for o in ondas] for k in chaves],
         [34, 20, 18, 45, 45])

    _aba(wb, "medias", ["onda", "q_id", "safra", "n", "media"],
         [[m["onda"], m["q_id"], m["safra"], m["n"], m["media"]] for m in med],
         [9, 20, 8, 7, 14])

    # O grao respondente-a-respondente (~124 mil linhas) fica FORA do
    # xlsx, num CSV ao lado. Dentro do xlsx ele responderia por 97% do
    # peso do arquivo e deixaria a planilha lenta de abrir no \\xpdocs.
    # Em CSV o Power Query ainda le mais rapido do que leria em aba.
    # Grao respondente a respondente, dentro do proprio xlsx.
    # Os rotulos longos de pergunta ficam de fora: eles se repetem em
    # todas as ~124 mil linhas e triplicariam o arquivo. Estao em
    # `dicionario` e `agregado`, ligados por q_id / opcao_id.
    enxuto = ["onda", "resp_id", "q_id", "familia", "safra", "opcao_id",
              "opcao_pt", "valor_num", "valor_bruto"]
    _aba(wb, "respostas", enxuto,
         [[r.get(c) for c in enxuto] for r in regs],
         [9, 16, 20, 12, 8, 18, 45, 10, 45])

    # ---- raw_norm: a Raw Data que a Base ja sabe ler, so que limpa ----
    #
    # Mesmo formato largo de sempre (uma coluna por pergunta, multipla
    # escolha separada por ';'), mas:
    #   . cabecalho canonico, que nunca deriva de um mes pro outro
    #   . rotulo antigo ja convertido no atual (aliases aplicados)
    #   . SEMPRE com ';' no fim -> mata o bug do COUNTIF "*opcao;*",
    #     que hoje perde a ultima opcao marcada por cada respondente
    #   . sem espaco a mais e sem espaco nao-separavel (\xa0)
    #
    # Serve para a ETAPA 1 da migracao: as formulas da Base continuam
    # exatamente como estao, so passam a apontar para esta tabela.
    perg_ordem, vistas = [], set()
    for a in sorted(agg, key=lambda x: (x["ordem_pergunta"], x["q_id"])):
        if a["q_id"] not in vistas and a["familia"] != "mes":
            vistas.add(a["q_id"])
            perg_ordem.append((a["q_id"], a["pergunta_pt"]))
    celulas = defaultdict(list)
    for r in regs:
        celulas[(int(r["onda"]), r["resp_id"], r["q_id"])].append(r["opcao_pt"])
    resp_todos = sorted({(int(r["onda"]), r["resp_id"]) for r in regs})
    _aba(wb, "raw_norm",
         ["Survey", "Id"] + [pt.replace("{safra}", "(safra corrente)")
                             for _, pt in perg_ordem],
         [[onda, rid] + [";".join(celulas.get((onda, rid, q), [])) + ";"
                         if celulas.get((onda, rid, q)) else None
                         for q, _ in perg_ordem]
          for onda, rid in resp_todos],
         [9, 16] + [34] * len(perg_ordem))

    # ---- dicionario: o que o registro conhece hoje ----
    dic = []
    for p in registro.perguntas:
        for o in p.opcoes:
            dic.append([p.id, p.tipo, p.pt, p.en, o.id, o.pt, o.en,
                        o.desde or "", o.ate or ""])
    _aba(wb, "dicionario",
         ["q_id", "tipo", "pergunta_pt", "pergunta_en", "opcao_id",
          "opcao_pt", "opcao_en", "desde_onda", "ate_onda"],
         dic, [20, 10, 50, 50, 18, 45, 45, 11, 11])

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)


def escrever_mes(caminho: Path, agg, med, onda: int, cfg, registro: Registro,
                 n_respostas: int, origem: str):
    par = cfg["parametros"]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ondas = sorted({a["onda"] for a in agg})
    pos = ondas.index(onda)
    anterior = ondas[pos - 1] if pos > 0 else None
    janela = ondas[max(0, pos - par["janela_serie"] + 1): pos + 1]

    pct_ant = {a["chave"]: a["pct"] for a in agg if a["onda"] == anterior}
    atual = [a for a in agg if a["onda"] == onda]

    # ---------------- meta ----------------
    ws = wb.create_sheet("meta")
    dt = datetime(int(str(onda)[:4]), int(str(onda)[4:]), 1)
    meses = ["janeiro", "fevereiro", "março", "abril", "maio", "junho", "julho",
             "agosto", "setembro", "outubro", "novembro", "dezembro"]
    for linha in [
        ("onda", onda),
        ("data", dt),
        ("titulo_pt", f"Edição de {meses[dt.month-1].capitalize()} {dt.year}"),
        ("titulo_en", f"{dt.strftime('%B')} {dt.year} edition"),
        ("respostas", n_respostas),
        ("onda_anterior", anterior or ""),
        ("gerado_em", datetime.now().replace(microsecond=0)),
        ("arquivo_origem", origem),
        ("ibovespa", (cfg.get("ibovespa_fechamento") or {}).get(onda, "")),
    ]:
        ws.append(list(linha))
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 46

    # ---------------- report: blocos do mes ----------------
    linhas = []
    for a in sorted(atual, key=lambda x: (x["ordem_pergunta"], x["ordem_opcao"], -x["pct"])):
        if a["familia"] == "mes":
            continue
        ant = pct_ant.get(a["chave"])
        linhas.append([a["q_id"], a["ordem_pergunta"], a["safra"],
                       a["pergunta_pt"], a["pergunta_en"],
                       a["opcao_id"], a["opcao_pt"], a["opcao_en"],
                       a["chave"], a["n"], a["base"], a["pct"],
                       ant, (a["pct"] - ant) if ant is not None else None,
                       a.get("fonte", "calculado")])
    _aba(wb, "report",
         ["q_id", "ordem", "safra", "pergunta_pt", "pergunta_en", "opcao_id",
          "opcao_pt", "opcao_en", "chave", "n", "base", "pct", "pct_ant",
          "delta", "fonte"],
         linhas, [20, 7, 7, 50, 50, 18, 45, 45, 34, 7, 7, 9, 9, 9, 11])

    # ---------------- serie: janela movel p/ graficos de linha ----------------
    ser = [a for a in agg if a["onda"] in janela and a["familia"] != "mes"]
    _aba(wb, "serie",
         ["onda", "data", "chave", "q_id", "opcao_id", "opcao_pt", "opcao_en", "pct"],
         [[a["onda"], datetime(int(str(a["onda"])[:4]), int(str(a["onda"])[4:]), 1),
           a["chave"], a["q_id"], a["opcao_id"], a["opcao_pt"], a["opcao_en"], a["pct"]]
          for a in sorted(ser, key=lambda x: (x["chave"], x["onda"]))],
         [9, 12, 34, 20, 18, 45, 45, 9])

    # ---------------- medias ----------------
    _aba(wb, "medias", ["onda", "data", "q_id", "safra", "n", "media"],
         [[m["onda"], datetime(int(str(m["onda"])[:4]), int(str(m["onda"])[4:]), 1),
           m["q_id"], m["safra"], m["n"], m["media"]]
          for m in med if m["onda"] in janela],
         [9, 12, 20, 8, 7, 14])

    # =================================================================
    #  paineis: UM RETANGULO DE ENDERECO FIXO POR PERGUNTA
    # =================================================================
    #
    # E esta aba que sustenta a planilha nova. Cada pergunta ocupa um
    # bloco em endereco fixo; o grafico e montado UMA vez apontando
    # para ele e nunca mais precisa ser refeito -- nem quando a
    # pergunta ganha alternativa, nem quando ela falta num mes.
    #
    # A ordem dos blocos vem do REGISTRO (campo `ordem`), nao dos
    # dados. E isso que garante o endereco: num mes em que a pergunta
    # nao foi feita, o bloco continua no lugar, vazio, e nada abaixo
    # dele se desloca.
    #
    # Pergunta nova entra com `ordem` MAIOR que as existentes, senao
    # ela empurra os blocos de baixo e quebra os graficos ja montados.
    #
    DESLOC = 1      # o Power Query pousa a linha 1 da origem na linha 2
    layout = []

    def _end(aba, col, r1, r2=None):
        """Endereco como ele aparece NA PLANILHA DO REPORT (com o +1)."""
        a = f"{aba}!${col}${r1 + DESLOC}"
        return a if r2 is None else f"{a}:${col}${r2 + DESLOC}"

    lp = par.get("linhas_por_painel", 18)
    ws = wb.create_sheet("paineis")
    ws.append(["Painéis por pergunta — endereços fixos. "
               "Não ordene, não filtre, não insira nem remova linhas."])
    ws["A1"].font = openpyxl.styles.Font(bold=True)
    for c, w in zip("ABCDEFG", (60, 60, 10, 10, 10, 8, 8)):
        ws.column_dimensions[c].width = w

    por_q = defaultdict(list)
    for a in atual:
        if a["familia"] != "mes":
            por_q[a["q_id"]].append(a)

    do_painel = sorted(
        [p for p in registro.perguntas if p.familia in ("recorrente", "safra")],
        key=lambda p: p.ordem)

    for i, p in enumerate(do_painel):
        base = 2 + i * (lp + 3)
        # O corte aqui e so defensivo: `checar_paineis` ja barrou a
        # rodada antes de chegar neste ponto.
        dados = sorted(por_q.get(p.id, []),
                       key=lambda x: (x["ordem_opcao"], -x["pct"]))[:lp]
        titulo_pt = dados[0]["pergunta_pt"] if dados else p.pt
        titulo_en = dados[0]["pergunta_en"] if dados else p.en
        ws.cell(base, 1, titulo_pt).font = openpyxl.styles.Font(bold=True)
        ws.cell(base, 2, titulo_en)
        ws.cell(base, 3, p.id)
        ws.cell(base, 4, dados[0]["safra"] if dados else "")
        ws.cell(base, 5, dados[0]["base"] if dados else None)
        for j, t in enumerate(["opcao_pt", "opcao_en", "pct", "pct_ant",
                               "delta", "n", "base"]):
            ws.cell(base + 1, j + 1, t).font = openpyxl.styles.Font(
                italic=True, color="808080")
        for k in range(lp):
            if k >= len(dados):
                break
            d, r = dados[k], base + 2 + k
            ant = pct_ant.get(d["chave"])
            ws.cell(r, 1, d["opcao_pt"])
            ws.cell(r, 2, d["opcao_en"])
            ws.cell(r, 3, d["pct"]).number_format = "0%"
            ws.cell(r, 4, ant).number_format = "0%"
            ws.cell(r, 5, (d["pct"] - ant) if ant is not None else None)
            ws.cell(r, 6, d["n"])
            ws.cell(r, 7, d["base"])
        r1, r2 = base + 2, base + 1 + lp
        layout.append(["painel", p.id, titulo_pt[:70],
                       _end("paineis", "A", base),
                       _end("paineis", "A", r1, r2),
                       _end("paineis", "B", r1, r2),
                       _end("paineis", "C", r1, r2),
                       _end("paineis", "E", r1, r2)])

    # =================================================================
    #  tendencias: SERIES TEMPORAIS EM COLUNAS FIXAS
    # =================================================================
    #
    # Uma coluna por serie, na ordem declarada em config.yaml
    # (`series_do_report`). Acrescentar serie no FIM da lista nao
    # desloca as colunas ja existentes -- por isso os graficos de
    # linha (inclusive o da capa) tambem sao montados uma vez so.
    #
    # As linhas sao sempre `janela_serie`, alinhadas ao fim: o mes
    # corrente e sempre a ultima linha, independente de quantas ondas
    # existirem. Nos primeiros meses sobram linhas em branco no topo.
    #
    pedidos = cfg.get("series_do_report") or []
    pct_por, med_por = defaultdict(dict), defaultdict(dict)
    for a in agg:
        pct_por[a["chave"]][a["onda"]] = a["pct"]
    for m in med:
        med_por[m["q_id"]][m["onda"]] = m["media"]
    ibov = cfg.get("ibovespa_fechamento") or {}

    def _valor(spec, o):
        if spec == "ibovespa":
            return ibov.get(o)
        if spec.startswith("media:"):
            return med_por.get(spec[6:], {}).get(o)
        return pct_por.get(spec, {}).get(o)

    def _rot(spec):
        if spec == "ibovespa":
            return "Ibovespa"
        if spec.startswith("media:"):
            p = registro.por_id.get(spec[6:])
            return f"Média — {p.pt[:50] if p else spec[6:]}"
        a = next((x for x in agg if x["chave"] == spec), None)
        return a["opcao_pt"] if a else spec

    ws = wb.create_sheet("tendencias")
    ws.append(["Séries do report — colunas fixas. Não reordene."])
    ws["A1"].font = openpyxl.styles.Font(bold=True)
    ws.append(["data"] + [_rot(s) for s in pedidos])
    ws.append(["(chave)"] + list(pedidos))
    for c in ws[2] + ws[3]:
        c.font = openpyxl.styles.Font(italic=True, color="808080")
    ws["A2"].font = openpyxl.styles.Font(bold=True)

    n_lin = par["janela_serie"]
    vazias = n_lin - len(janela)
    for _ in range(vazias):
        ws.append([None] * (len(pedidos) + 1))
    for o in janela:
        ws.append([datetime(int(str(o)[:4]), int(str(o)[4:]), 1)]
                  + [_valor(s, o) for s in pedidos])
    ws.column_dimensions["A"].width = 12
    for j in range(len(pedidos)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j + 2)].width = 13

    t1, t2 = 4, 3 + n_lin
    for j, spec in enumerate(pedidos):
        col = openpyxl.utils.get_column_letter(j + 2)
        layout.append(["tendência", spec, _rot(spec)[:70],
                       _end("tendencias", col, 2),
                       _end("tendencias", "A", t1, t2),
                       "", _end("tendencias", col, t1, t2), ""])

    # ---------------- q_mes: SLOTS DE ENDERECO FIXO ----------------
    #
    # Layout deterministico. O grafico do slot 1 aponta sempre para
    #   rotulos  q_mes!$A$4:$A$15
    #   valores  q_mes!$C$4:$C$15
    # e nunca mais precisa ser refeito. Linhas nao usadas ficam vazias
    # (o grafico ignora). Ver LEIA-ME.md, secao "Pergunta do mes".
    #
    ws = wb.create_sheet("q_mes")
    ws.append(["Slots de pergunta do mês — endereços fixos, não insira/remova linhas"])
    ws["A1"].font = openpyxl.styles.Font(bold=True)
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 60
    for c in "CDE":
        ws.column_dimensions[c].width = 10

    por_slot = defaultdict(list)
    for a in atual:
        if a["familia"] == "mes":
            por_slot[a["q_id"]].append(a)

    alt = par["linhas_por_slot"]
    for s in range(1, par["slots_pergunta_mes"] + 1):
        base = 2 + (s - 1) * (alt + 3)
        dados = sorted(por_slot.get(f"q_mes_{s}", []), key=lambda x: -x["pct"])
        ws.cell(base, 1, dados[0]["pergunta_pt"] if dados else "")
        ws.cell(base, 2, dados[0]["pergunta_en"] if dados else "")
        ws.cell(base, 3, f"SLOT {s}")
        ws.cell(base, 1).font = openpyxl.styles.Font(bold=True)
        ws.cell(base, 3).font = openpyxl.styles.Font(bold=True, color="808080")
        for j, t in enumerate(["opcao_pt", "opcao_en", "pct", "n", "base"]):
            ws.cell(base + 1, j + 1, t).font = openpyxl.styles.Font(italic=True, color="808080")
        for k in range(alt):
            r = base + 2 + k
            if k < len(dados):
                d = dados[k]
                ws.cell(r, 1, d["opcao_pt"])
                ws.cell(r, 2, d["opcao_en"])
                ws.cell(r, 3, d["pct"]).number_format = "0%"
                ws.cell(r, 4, d["n"])
                ws.cell(r, 5, d["base"])
        r1, r2 = base + 2, base + 1 + alt
        layout.append(["pergunta do mês", f"slot {s}",
                       (dados[0]["pergunta_pt"][:70] if dados else "(vazio)"),
                       _end("q_mes", "A", base),
                       _end("q_mes", "A", r1, r2),
                       _end("q_mes", "B", r1, r2),
                       _end("q_mes", "C", r1, r2), ""])

    # ---------------- layout: a lista de enderecos, pronta ----------------
    #
    # Esta aba existe para voce nao ter que contar linha. Ela diz, para
    # cada grafico, exatamente qual intervalo apontar na planilha do
    # report. Os enderecos ja vem com o deslocamento do Power Query
    # aplicado -- e so copiar e colar em Selecionar Dados.
    #
    _aba(wb, "layout",
         ["tipo", "chave", "descrição", "título", "rótulos_pt",
          "rótulos_en", "valores", "delta"],
         layout, [16, 24, 70, 24, 26, 26, 26, 26])

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)


# =====================================================================
#  Relatorio da rodada
# =====================================================================

def escrever_log(cfg, onda, diag, n_total, respondentes, agg, origem):
    pasta = RAIZ / cfg["caminhos"]["logs"]
    pasta.mkdir(parents=True, exist_ok=True)
    p = pasta / f"rodada_{onda}.txt"
    L = []
    L.append("=" * 68)
    L.append(f" PESQUISA DE ASSESSORES — onda {onda}")
    L.append(f" gerado em {datetime.now():%d/%m/%Y %H:%M}")
    L.append(f" origem: {origem}")
    L.append("=" * 68)
    L.append(f"\nRespostas na onda: {n_total}")

    L.append("\n--- Perguntas reconhecidas ---")
    qs = sorted({(a["ordem_pergunta"], a["q_id"], a["familia"], a["safra"])
                 for a in agg if a["onda"] == onda})
    for ordem, qid, fam, safra in qs:
        n = len(respondentes.get(qid, ()))
        marca = f" [safra {safra}]" if safra else ""
        cob = f"{n}/{n_total}" if n else "-"
        L.append(f"  {qid:<22} {fam:<11}{marca:<14} respondentes {cob}")

    if diag.novas_perguntas:
        L.append("\n--- Perguntas do mês (viraram slots) ---")
        for i, (h, c) in enumerate(diag.novas_perguntas, 1):
            L.append(f"  SLOT {i}: {str(h).strip()[:88]}  ({c} respostas)")

    if diag.perguntas_ausentes:
        L.append("\n--- ATENÇÃO: recorrentes que não vieram nesta onda ---")
        for q in diag.perguntas_ausentes:
            L.append(f"  {q}")

    if diag.faixas_novas:
        L.append("\n--- Faixas novas (lidas pelo parser, nada a fazer) ---")
        for q, fs in diag.faixas_novas.items():
            for f in sorted(fs):
                L.append(f"  [{q}] {f}")

    if diag.livres:
        L.append("\n--- Texto livre agrupado ---")
        for q, c in diag.livres.items():
            for v, n in c.most_common():
                L.append(f"  [{q}] {v}  ({n})")

    if diag.avisos:
        L.append("\n--- Avisos (foram para 'Outros') ---")
        L += ["  " + a for a in diag.avisos]

    if diag.erros:
        L.append("\n" + "!" * 68)
        L.append("  ERRO — alternativa relevante não declarada em perguntas.yaml")
        L.append("!" * 68)
        L += ["  " + e for e in diag.erros]
        L.append("\n  Como resolver: abra config/perguntas.yaml, ache a pergunta")
        L.append("  e ou (a) adicione a alternativa em `opcoes:`, ou (b) se for")
        L.append("  o mesmo conceito com rótulo novo, mova o rótulo antigo para")
        L.append("  `aliases:` e ponha o novo em `pt:`. Depois rode de novo.")
    else:
        L.append("\nOK — nenhuma alternativa desconhecida relevante.")

    txt = "\n".join(L)
    p.write_text(txt, encoding="utf-8")
    return txt, p


# =====================================================================
#  Bootstrap: carga inicial a partir da Raw Data existente
# =====================================================================

def bootstrap(caminho_pa: Path, registro: Registro, cfg):
    """Importa as ondas historicas da aba 'Raw Data' da PA Principal."""
    cab, linhas = ler_planilha(caminho_pa, "Raw Data")
    col_onda = next((i for i, h in enumerate(cab)
                     if registro.eh_meta(h) == "onda"), 0)

    por_onda = defaultdict(list)
    for r in linhas:
        if r[col_onda] is None:
            continue
        por_onda[int(r[col_onda])].append(r)

    todos, diags = [], {}
    for onda in sorted(por_onda):
        regs, diag, n, resp = normalizar(cab, por_onda[onda], registro, onda,
                                         caminho_pa.name)
        todos += regs
        diags[onda] = (diag, n, resp)
        marca = "ERRO" if diag.erros else ("aviso" if diag.avisos else "ok")
        print(f"  {onda}  {n:>4} respostas  "
              f"{len({r['q_id'] for r in regs}):>2} perguntas  [{marca}]")
    return todos, diags


# =====================================================================
#  CLI
# =====================================================================

def achar_entrada(pasta: Path):
    if not pasta.exists():
        return None
    cand = [p for p in pasta.glob("*.xlsx") if not p.name.startswith("~$")]
    return max(cand, key=lambda p: p.stat().st_mtime) if cand else None


def main():
    ap = argparse.ArgumentParser(description="Pipeline da Pesquisa de Assessores XP")
    ap.add_argument("--arquivo", help="export do Forms (.xlsx)")
    ap.add_argument("--onda", type=int, help="AAAAMM (se não vier, é deduzido)")
    ap.add_argument("--conferir", action="store_true", help="só valida, não escreve")
    ap.add_argument("--bootstrap", help="carga inicial a partir de PA Principal.xlsx")
    ap.add_argument("--config", default=str(RAIZ / "config" / "config.yaml"))
    args = ap.parse_args()

    with open(args.config, encoding="utf-8") as fh:
        cfg = yaml.safe_load(fh)
    registro = Registro(RAIZ / "config" / "perguntas.yaml")
    par = cfg["parametros"]
    cam = cfg["caminhos"]

    store_path = RAIZ / cam["store"]
    saida = Path(cam["saida"])

    lp = par.get("linhas_por_painel", 18)

    def barrar_paineis(agg, onda):
        """-> True se a rodada tem que parar aqui."""
        probs = checar_paineis(agg, onda, registro, lp)
        if not probs:
            return False
        print("\n" + "!" * 68)
        print("  ERRO — painel pequeno demais; o gráfico sairia truncado")
        print("!" * 68)
        for x in probs:
            print("  " + x)
        print(f"\n  Aumente `linhas_por_painel` (hoje {lp}) em config/config.yaml.")
        print("  Atenção: isso desloca TODOS os painéis. Depois de aumentar,")
        print("  refaça os endereços dos gráficos usando a aba `layout`.")
        print("\n  Nada foi gravado.")
        return True

    # Numero publicado nao muda: ondas ate aqui saem congeladas.
    congelados, metas = carregar_congelados(
        RAIZ / "config" / "valores_publicados.csv")
    corte = par.get("ultima_onda_publicada")
    if congelados and corte:
        print(f"Histórico congelado até {corte} "
              f"({len(congelados)} valores publicados)")
    elif corte:
        print("AVISO: ultima_onda_publicada está definida, mas "
              "config/valores_publicados.csv não existe.\n"
              "       Rode src/congelar_historico.py antes, senão o "
              "histórico sai recalculado.")

    # ---------------- bootstrap ----------------
    if args.bootstrap:
        print(f"Carga inicial a partir de {args.bootstrap}\n")
        regs, diags = bootstrap(Path(args.bootstrap), registro, cfg)
        erros = [e for d, _, _ in diags.values() for e in d.erros]
        print(f"\n{len(regs)} registros, {len(diags)} ondas.")
        if erros:
            print(f"\n{len(erros)} alternativas não declaradas (as maiores):")
            for e in erros[:25]:
                print("  " + e)
        agg = agregar(regs, par["denominador"], congelados, corte,
                      metas, registro)
        ultima = max(int(r["onda"]) for r in regs)
        if barrar_paineis(agg, ultima):
            return 1
        if args.conferir:
            return 1 if erros else 0
        salvar_store(store_path, regs)
        med = medias(regs)
        escrever_historica(saida / cam["base_historica"], regs, agg, med, registro)
        print(f"\nGerado: {saida / cam['base_historica']}")
        n_ult = len({r["resp_id"] for r in regs if int(r["onda"]) == ultima})
        escrever_mes(saida / cam["base_mes"], agg, med, ultima, cfg, registro,
                     n_ult, args.bootstrap)
        print(f"Gerado: {saida / cam['base_mes']}  (onda {ultima})")
        return 1 if erros else 0

    # ---------------- rodada normal ----------------
    entrada = Path(args.arquivo) if args.arquivo else achar_entrada(Path(cam["entrada"]))
    if not entrada or not entrada.exists():
        print(f"Nenhum .xlsx encontrado em {cam['entrada']}")
        return 1
    print(f"Lendo {entrada.name}")

    cab, linhas = ler_planilha(entrada)
    onda = args.onda or deduzir_onda(cab, linhas, registro)
    if not onda:
        print("Não consegui deduzir a onda. Passe --onda AAAAMM.")
        return 1
    print(f"Onda {onda}, {len(linhas)} respostas\n")

    regs, diag, n_total, resp = normalizar(cab, linhas, registro, onda, entrada.name)

    store = carregar_store(store_path)
    completo = mesclar(store, regs, onda)
    agg = agregar(completo, par["denominador"], congelados, corte,
                  metas, registro)
    med = medias(completo)

    txt, log_path = escrever_log(cfg, onda, diag, n_total, resp, agg, entrada.name)
    print(txt)

    if diag.erros and par.get("falhar_em_desconhecida", True):
        print(f"\nNada foi gravado. Corrija perguntas.yaml e rode de novo.")
        print(f"Log: {log_path}")
        return 1
    if barrar_paineis(agg, onda):
        print(f"Log: {log_path}")
        return 1
    if args.conferir:
        print("\n--conferir: nada gravado.")
        return 0

    salvar_store(store_path, completo)
    escrever_historica(saida / cam["base_historica"], completo, agg, med, registro)
    escrever_mes(saida / cam["base_mes"], agg, med, onda, cfg, registro,
                 n_total, entrada.name)
    print(f"\nGerado: {saida / cam['base_historica']}")
    print(f"Gerado: {saida / cam['base_mes']}")
    print(f"Log:    {log_path}")
    print("\nAgora abra a PA Report.xlsx e clique em Dados > Atualizar Tudo.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
