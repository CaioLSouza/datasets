# -*- coding: utf-8 -*-
"""
=======================================================================
 GERADOR DA PA REPORT.XLSX   (roda UMA vez, na montagem)
=======================================================================

 Monta a planilha do report ja ligada as bases -- sem Power Query.

 -------------------------------------------------------------------
 POR QUE SEM POWER QUERY
 -------------------------------------------------------------------
 As consultas do Power Query nao ficam em XML dentro do .xlsx: elas
 vivem num blob chamado DataMashup, um ZIP em base64 de formato
 proprietario. Nenhuma biblioteca de Python escreve isso de forma
 confiavel -- forjar o blob na mao gera arquivo corrompido.

 Entao esta planilha usa REFERENCIA EXTERNA no lugar:

     ='\\\\xpdocs\\...\\bases\\[PA Base Mes Atual.xlsx]paineis'!A1

 Uma formula por celula, apontando para a base gerada. O efeito e o
 mesmo do Power Query -- o numero vem da base e se atualiza sozinho --
 so que a ligacao e nativa do Excel e nao precisa ser montada a mao.

 -------------------------------------------------------------------
 O DESLOCAMENTO DE UMA LINHA
 -------------------------------------------------------------------
 A linha 1 de cada aba fica com um cabecalho generico e a linha N da
 origem cai na linha N+1. Isso e de proposito: e exatamente o que o
 Power Query faria.

 Assim os enderecos que a aba `layout` publica valem para os dois
 caminhos. Se um dia voce trocar estas formulas por consultas de
 verdade, nenhum grafico precisa ser refeito.

 -------------------------------------------------------------------
 USO
 -------------------------------------------------------------------
   python montar_report.py
       gera na pasta `saida` do config.yaml, ao lado das bases

   python montar_report.py --destino "C:\\...\\PA Report.xlsx"
       gera em outro lugar

 O arquivo NAO e sobrescrito se ja existir -- os 22 links dos PPTs
 apontam para ele, e regerar mataria todos. Use --forcar so quando
 souber o que esta fazendo.
=======================================================================
"""

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
import yaml
from openpyxl.styles import Alignment, Font, PatternFill

from pipeline import RAIZ

# As abas que o report consome, na ordem em que entram na planilha.
# `layout` fica por ultimo: e referencia, nao alimenta grafico.
ABAS = ["paineis", "tendencias", "q_mes", "meta", "layout"]

# O mesmo deslocamento que o pipeline aplica ao publicar os enderecos:
# a linha 1 fica com cabecalho e a linha N da origem cai em N+1.
DESLOC = 1

# Abas de endereco fixo -- ordenar ou filtrar aqui quebra os graficos.
FIXAS = {"paineis", "tendencias", "q_mes"}


def ref_externa(pasta: Path, arquivo: str, aba: str, linha: int, col: str) -> str:
    """Formula de referencia externa a uma pasta de trabalho fechada.

    O caminho inteiro vai entre aspas simples porque tem espaco:
        ='\\\\servidor\\pasta com espaco\\[Arquivo.xlsx]Aba'!A1
    """
    return f"='{pasta}\\[{arquivo}]{aba}'!{col}{linha}"


def montar(caminho_base: Path, destino: Path, forcar: bool, modelo: Path = None):
    """caminho_base  o que vai DENTRO das formulas (a base na rede)
       modelo        de onde leio as dimensoes; por padrao, a propria base

    Os dois so diferem quando a base ainda nao existe na rede e voce quer
    gerar a planilha de antemao, a partir de uma copia local.
    """
    estrutura = modelo or caminho_base
    if destino.exists() and not forcar:
        sys.exit(
            f"Já existe:\n   {destino}\n\n"
            "NÃO vou sobrescrever. Os 22 links OLE dos dois PPTs apontam\n"
            "para este arquivo — regerar mataria todos eles, e a aba Charts\n"
            "que você montou seria perdida junto.\n\n"
            "Se é isso mesmo que você quer, passe --forcar. Mas antes:\n"
            "faça uma cópia do arquivo atual."
        )
    if not estrutura.exists():
        sys.exit(f"Não achei a base do mês:\n   {estrutura}\n\n"
                 "Rode o pipeline antes (montar.bat ou rodar.bat).")
    if modelo:
        print(f"   (estrutura lida de {modelo.name}; as fórmulas apontam\n"
              f"    para {caminho_base})\n")

    origem = openpyxl.load_workbook(estrutura, read_only=True)
    pasta, arquivo = caminho_base.parent, caminho_base.name

    # ------------------------------------------------------------------
    #  Ate onde cada aba precisa ir
    # ------------------------------------------------------------------
    #  Nao da para confiar no max_row da origem: ele para na ultima
    #  celula PREENCHIDA. Um mes em que `selic_alvo` nao foi perguntada
    #  deixa o bloco dela vazio, o max_row para em 234 -- e as formulas
    #  nunca chegariam na linha 253, que e onde o grafico dessa pergunta
    #  aponta. O grafico quebraria sem aviso.
    #
    #  A aba `layout` e a autoridade: ela publica o endereco de cada
    #  grafico. Extraio dela ate onde ir.
    # ------------------------------------------------------------------
    minimo = {}
    if "layout" in origem.sheetnames:
        import re
        padrao = re.compile(r"(\w+)!\$[A-Z]+\$(\d+)(?::\$[A-Z]+\$(\d+))?")
        for linha in origem["layout"].iter_rows(values_only=True):
            for cel in linha:
                for aba, r1, r2 in padrao.findall(str(cel or "")):
                    fim = int(r2 or r1) - DESLOC       # volta ao grão da origem
                    minimo[aba] = max(minimo.get(aba, 0), fim)
    if minimo:
        print("   (o layout exige, na origem: "
              + ", ".join(f"{a} até {r}" for a, r in sorted(minimo.items()))
              + ")\n")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    cinza = PatternFill("solid", fgColor="EEEEEE")
    resumo = []

    for nome in ABAS:
        if nome not in origem.sheetnames:
            print(f"   (a base não tem a aba `{nome}` — pulei)")
            continue
        src = origem[nome]
        n_lin = max(src.max_row, minimo.get(nome, 0))
        n_col = src.max_column
        ws = wb.create_sheet(nome)

        # Linha 1: cabecalho generico, igual ao que o Power Query poria.
        # E ele que cria o deslocamento de 1 que os enderecos assumem.
        for c in range(1, n_col + 1):
            cel = ws.cell(1, c, f"Column{c}")
            cel.font = Font(bold=True, color="808080", size=9)
            cel.fill = cinza
        ws.freeze_panes = "A2"

        for r in range(1, n_lin + 1):
            for c in range(1, n_col + 1):
                letra = openpyxl.utils.get_column_letter(c)
                ws.cell(r + 1, c,
                        ref_externa(pasta, arquivo, nome, r, letra))

        larg = {"paineis": [58, 58, 10, 10, 10, 8, 8],
                "tendencias": [12] + [14] * (n_col - 1),
                "q_mes": [58, 58, 10, 8, 8],
                "meta": [18, 46],
                "layout": [16, 24, 60, 24, 26, 26, 26, 26]}.get(nome, [])
        for i, w in enumerate(larg[:n_col], start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        resumo.append((nome, n_lin, n_col, nome in FIXAS))
        print(f"   {nome:<12} {n_lin:>4} x {n_col:<3} "
              f"-> linhas 2..{n_lin + 1}"
              + ("   [endereço fixo]" if nome in FIXAS else ""))

    # ---------------- aba de instrucoes, primeira da planilha ----------
    ws = wb.create_sheet("LEIA-ME", 0)
    ws.column_dimensions["A"].width = 100
    texto = [
        ("PA Report — a planilha do report", True),
        ("", False),
        ("Os números desta planilha vêm das bases geradas pelo pipeline,", False),
        ("por referência externa. Você não precisa montar consulta nenhuma.", False),
        ("", False),
        ("TODO MÊS", True),
        ("   1. rode o rodar.bat", False),
        ("   2. abra esta planilha", False),
        ("   3. se o Excel perguntar sobre links, clique em ATUALIZAR", False),
        ("   4. abra os dois PPTs e Atualizar Links", False),
        ("", False),
        ("SE O EXCEL NÃO PERGUNTAR", False),
        ("   Dados > Conexões > Editar Links > Atualizar Valores", False),
        ("   Para não perguntar mais: Editar Links > Prompt de Inicialização", False),
        ("   > 'Não exibir o alerta e atualizar os links'", False),
        ("", False),
        ("NÃO MEXA NAS ABAS paineis, tendencias E q_mes", True),
        ("   Os gráficos apontam para endereços absolutos delas.", False),
        ("   Ordenar, filtrar, inserir ou remover linha desloca tudo.", False),
        ("", False),
        ("ONDE ESTÃO OS ENDEREÇOS DOS GRÁFICOS", True),
        ("   Na aba `layout`. Já vêm com o deslocamento aplicado —", False),
        ("   é copiar e colar em Selecionar Dados.", False),
        ("", False),
        ("O QUE FALTA FAZER (uma vez só)", True),
        ("   Copiar as abas Charts e Gráfico capa da PA Principal antiga", False),
        ("   e repontar cada gráfico usando a aba `layout`.", False),
        ("   O passo a passo está em powerquery/INSTRUCOES.md", False),
        ("", False),
        ("A base de onde tudo vem:", False),
        (f"   {caminho_base}", False),
    ]
    for i, (linha, negrito) in enumerate(texto, start=1):
        cel = ws.cell(i, 1, linha)
        cel.font = Font(bold=negrito, size=12 if negrito and i == 1 else 11)
        cel.alignment = Alignment(vertical="top")

    origem.close()
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)

    print(f"\nGerado: {destino}")
    print(f"        {sum(l * c for _, l, c, _ in resumo):,} fórmulas de "
          f"referência externa".replace(",", "."))
    print("\nAo abrir pela primeira vez, o Excel vai perguntar sobre os")
    print("links. Clique em ATUALIZAR — sem isso as células ficam zeradas.")


def main(argv=None):
    ap = argparse.ArgumentParser(description="Gera a PA Report.xlsx")
    ap.add_argument("--destino", help="onde salvar (padrão: ao lado das bases)")
    ap.add_argument("--config", default=str(RAIZ / "config" / "config.yaml"))
    ap.add_argument("--forcar", action="store_true",
                    help="sobrescreve um arquivo existente (PERIGOSO: mata os "
                         "22 links dos PPTs)")
    ap.add_argument("--modelo",
                    help="lê as dimensões daqui em vez da base da rede — "
                         "serve para gerar a planilha antes de a base existir")
    args = ap.parse_args(argv)

    with open(args.config, encoding="utf-8") as fh:
        cfg = yaml.safe_load(fh)
    cam = cfg["caminhos"]
    saida = Path(cam["saida"])
    base_mes = saida / cam["base_mes"]

    destino = Path(args.destino) if args.destino else saida.parent / "PA Report.xlsx"

    print(f"lendo   {base_mes}")
    print(f"gerando {destino}\n")
    montar(base_mes, destino, args.forcar,
           Path(args.modelo) if args.modelo else None)
    return 0


if __name__ == "__main__":
    sys.exit(main())
