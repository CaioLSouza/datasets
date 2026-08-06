# -*- coding: utf-8 -*-
"""
=======================================================================
 PESQUISA DE ASSESSORES XP — versão célula a célula
=======================================================================

 Faz o mesmo que o rodar.bat, só que em pedaços — útil quando o
 terminal é chato de usar, ou quando você quer olhar o resultado de
 perto antes de seguir.

 COMO RODAR
   VS Code    abra este arquivo e clique em "Run Cell" acima de cada
              bloco `# %%`, ou Shift+Enter. Ele abre a Interactive
              Window sozinho.
   Jupyter    converta com jupytext, ou copie os blocos para células.
   Terminal   `py rodar_interativo.py` também funciona: roda tudo de
              cima a baixo. Mas para a rotina mensal prefira o
              rodar.bat, que devolve código de erro.

 A ORDEM
   Setup      sempre primeiro
   Parte 1    uma vez na vida, para montar
   Parte 2    todo mês
   Parte 3    quando quiser conferir alguma coisa

=======================================================================
"""

# %% [markdown]
# # Setup
#
# Rode esta célula primeiro, **sempre**. Ela acha a pasta do projeto,
# põe o `src` no path e importa os três módulos.

# %%
from pathlib import Path
import sys

# --- ONDE ESTÁ A PASTA DO PROJETO ------------------------------------
# Se este arquivo estiver dentro dela, a busca abaixo acha sozinha,
# subindo as pastas até encontrar src\pipeline.py.
try:
    PROJETO = Path(__file__).resolve().parent
except NameError:                     # rodando célula a célula
    PROJETO = Path.cwd()

while PROJETO != PROJETO.parent and not (PROJETO / "src" / "pipeline.py").exists():
    PROJETO = PROJETO.parent

# Se não achou (arquivo fora da pasta do projeto), ajuste na mão:
if not (PROJETO / "src" / "pipeline.py").exists():
    PROJETO = Path(r"\\xpdocs\Research\Equities\Estrategia\Reports"
                   r"\Pesquisa assessores\_motor")

# --- A PLANILHA ANTIGA, NA REDE --------------------------------------
# Só é usada na montagem (Parte 1). Depois disso, nunca mais.
PA_PRINCIPAL = (r"\\xpdocs\Research\Equities\Estrategia\Reports"
                r"\Pesquisa assessores\PA Principal.xlsx")

# ---------------------------------------------------------------------
sys.path.insert(0, str(PROJETO / "src"))

# autoreload só existe no IPython. Sem ele, editar um .py exigiria
# reiniciar o kernel para a mudança valer.
try:
    from IPython import get_ipython
    _ip = get_ipython()
    if _ip is not None:
        _ip.run_line_magic("reload_ext", "autoreload")
        _ip.run_line_magic("autoreload", "2")
except Exception:
    pass

import congelar_historico
import pipeline
import reconciliar

print("projeto ................", PROJETO)
print("src no path ............", (PROJETO / "src").exists())
print("PA Principal acessível .", Path(PA_PRINCIPAL).exists())
if not Path(PA_PRINCIPAL).exists():
    print("\n   -> rede fora do ar, ou ajuste PA_PRINCIPAL acima")


# %% [markdown]
# ---
# # Parte 1 — Montagem
#
# **Roda uma vez na vida.** Se a `PA Report.xlsx` já existe e está
# funcionando, pule direto para a Parte 2.

# %% [markdown]
# ## Fase 1 — Congelar o histórico publicado
#
# Lê a aba `Base` da PA Principal e grava, onda a onda, exatamente o que
# foi publicado. É isto que garante que **nenhum número que já foi ao ar
# mude de valor**.
#
# Espere **3.718 valores** em **76 ondas** (fev/2020 a jul/2026), mais
# três avisos — todos esperados, explicados depois da célula.

# %%
congelar_historico.main(PA_PRINCIPAL)


# %% [markdown]
# **Os três avisos:**
#
# 1. **Edição Coronavírus** — fev/2020 teve duas pesquisas: a regular
#    (dia 01) e uma extraordinária (dia 15). A chave de onda é `AAAAMM`
#    e não comporta as duas; fica a regular. Nada a fazer.
#
# 2. **Ranking em jun/2023** — naquele mês a Base guardou a *ordem de
#    preferência* dos setores (1 a 14), não percentual. Não existe
#    conversão honesta de posição para %, então jun/2023 fica sem essa
#    pergunta. Nada a fazer.
#
# 3. **Linha duplicada de `apetite_risco`** — **esta exige decisão
#    sua.** A Base publica 12,15% *e* 21,50% para *"Melhora na
#    recuperação econômica global"* em jul/2026. A de cima tem um `;`
#    colado no rótulo e subconta; a de baixo é a correção. O congelador
#    fica com a de baixo. Se você conferir e concluir o contrário, edite
#    o valor em `config/valores_publicados.csv`.

# %% [markdown]
# ## Fase 2 — Importar as respostas históricas
#
# Lê a aba `Raw Data` e normaliza as 37 ondas que têm resposta
# individual (jul/2023 em diante) contra o registro de perguntas.
#
# Espere **123.739 registros, 37 ondas**. Os `[aviso]` são normais — são
# write-ins de texto livre que foram para o balde *Outros*. Só `[ERRO]`
# trava a rodada.

# %%
codigo = pipeline.main(["--bootstrap", PA_PRINCIPAL])

print("\n" + ("OK — pode seguir para a Fase 3" if codigo == 0
              else "PAROU. Nada foi gravado — leia a mensagem acima."))


# %% [markdown]
# ## Fase 3 — Conferir que nada mudou
#
# O teste que autoriza tudo o que vem depois. Compara, célula a célula,
# o que o pipeline entrega contra o que a Base publica hoje.
#
# **O bloco 1 tem que dar ~100%** — hoje, 3.717 de 3.718. O único
# divergente é aquela linha duplicada da Fase 1.
#
# O bloco 2 é auditoria: mostra o que o recálculo do bruto *diria* se o
# histórico não estivesse congelado. É onde aparecem os erros do
# processo antigo. Nada disso vai para gráfico.

# %%
reconciliar.main(PA_PRINCIPAL)


# %% [markdown]
# > **Pare aqui se** o bloco 1 vier abaixo de ~99,9%, ou se aparecer
# > divergência que não seja a do `apetite_risco`.
# >
# > Confira também a linha `Ondas na Base`: tem que dizer **76**. Se
# > disser 37, o congelamento pegou só as colunas com código na linha 1
# > e perdeu fev/2020 a jun/2023.
#
# Passou? Falta só montar a `PA Report.xlsx` — essa parte é no Excel, e
# está em `powerquery/INSTRUCOES.md`.

# %% [markdown]
# ---
# # Parte 2 — Rotina mensal
#
# Depois que a `PA Report.xlsx` estiver montada, o mês inteiro é isto:
#
# 1. Exportar o Forms para `input_forms\`
# 2. **Rodar a célula abaixo**
# 3. Abrir a PA Report → *Dados › Atualizar Tudo*
# 4. Abrir os dois PPTs → *Atualizar Links*

# %%
codigo = pipeline.main([])

print("\n" + "=" * 62)
if codigo == 0:
    print("OK. Agora:")
    print("   1) abra a PA Report.xlsx")
    print("   2) Dados > Atualizar Tudo")
    print("   3) abra os dois PPTs e Atualizar Links")
else:
    print("PAROU — nada foi gravado.")
    print("Entrou alternativa nova na pesquisa, ou um painel estourou.")
    print("A mensagem acima diz qual é e o que fazer.")


# %% [markdown]
# **Se parar**, é sempre um destes dois, e ambos vêm explicados na
# mensagem:
#
# - **alternativa não declarada** → abra `config/perguntas.yaml`, ache a
#   pergunta e decida: é alternativa nova (adicione em `opcoes:`) ou é o
#   mesmo conceito com rótulo novo (mova o antigo para `aliases:`)?
# - **painel estourou** → aumente `linhas_por_painel` no `config.yaml`.
#   Atenção: isso desloca todos os painéis e obriga a refazer os
#   endereços dos gráficos.
#
# Depois de corrigir, rode a célula de novo. Reprocessar o mesmo mês é
# seguro: o pipeline substitui a onda inteira, não duplica.

# %% [markdown]
# ---
# # Parte 3 — Conferências
#
# Nada aqui altera arquivo nenhum. É só para olhar.

# %%
import collections

import openpyxl
import yaml

cfg = yaml.safe_load(open(PROJETO / "config" / "config.yaml", encoding="utf-8"))
BASES = Path(cfg["caminhos"]["saida"])
MES = BASES / cfg["caminhos"]["base_mes"]
HIST = BASES / cfg["caminhos"]["base_historica"]

print("base do mês ...", MES, "|", "ok" if MES.exists() else "NÃO EXISTE")
print("histórico .....", HIST, "|", "ok" if HIST.exists() else "NÃO EXISTE")


# %% [markdown]
# ## Os endereços dos gráficos
#
# É esta lista que você usa ao montar a `PA Report.xlsx` — os endereços
# já vêm com o deslocamento do Power Query aplicado.

# %%
wb = openpyxl.load_workbook(MES, read_only=True)
for linha in wb["layout"].iter_rows(values_only=True):
    print(" ".join(str(x or "")[:26].ljust(26) for x in linha[:5]))
wb.close()


# %% [markdown]
# ## Quanto histórico tem cada onda
#
# `publicado` = veio congelado da Base antiga, intocado.
# `calculado` = veio do dado bruto.
#
# Nas ondas antigas quase tudo é *publicado*. O que aparece como
# *calculado* nelas são os baldes `_outros`, os write-ins e a pergunta
# do mês — coisas que nunca foram linha fixa na Base.

# %%
wb = openpyxl.load_workbook(HIST, read_only=True)
it = wb["agregado"].iter_rows(values_only=True)
h = list(next(it))
i_onda, i_fonte = h.index("onda"), h.index("fonte")

por_onda = collections.defaultdict(collections.Counter)
for r in it:
    por_onda[r[i_onda]][r[i_fonte]] += 1
wb.close()

ondas = sorted(por_onda)
print(f"{len(ondas)} ondas, de {ondas[0]} a {ondas[-1]}\n")
print(f"{'onda':>8} {'publicado':>10} {'calculado':>10}")
for o in ondas:
    c = por_onda[o]
    print(f"{o:>8} {c['publicado']:>10} {c['calculado']:>10}")


# %% [markdown]
# ## Um mês específico, pergunta por pergunta

# %%
ONDA = 202607          # troque aqui

wb = openpyxl.load_workbook(HIST, read_only=True)
it = wb["agregado"].iter_rows(values_only=True)
h = list(next(it))
ix = {c: h.index(c) for c in
      ("onda", "q_id", "opcao_pt", "pct", "n", "base", "fonte")}

atual = None
for r in it:
    if r[ix["onda"]] != ONDA:
        continue
    if r[ix["q_id"]] != atual:
        atual = r[ix["q_id"]]
        print(f"\n--- {atual} ---")
    print(f"   {str(r[ix['opcao_pt']])[:46]:<46} "
          f"{r[ix['pct']]:>7.1%}  ({r[ix['fonte']]})")
wb.close()
