"""O motor. Roda todo mês.

    python atualizar.py
    python atualizar.py --bootstrap ["...\\PA Principal.xlsx"]

Lê os exports do Forms em input_forms\\, apura, junta com o histórico congelado
e grava dois arquivos em bases\\:

  PA Base.xlsx          a base geral -- todo o histórico em formato longo, mais
                        as respostas no grão respondente. Para análise.

  PA Charts Data.xlsx   uma aba por tabela PRONTA PARA GRÁFICO. É este arquivo
                        que a PA Charts.xlsx consulta via Power Query.

As abas de PA Charts Data.xlsx:

  d_<pergunta>   distribuição da onda corrente: rótulo, atual, anterior, delta.
                 Já ordenada. Serve barra, barra empilhada e pizza, em PT e EN
                 (são duas colunas de rótulo, você escolhe qual usar).
  s_<pergunta>   série temporal: uma linha por onda, uma coluna por
                 alternativa. Cabeçalho em português.
  medias         resposta média do sentimento e do Ibovespa esperado, por onda.
  capa           a série da capa mais o fechamento do Ibovespa.
  meta           uma linha por onda: data, respondentes, regime.
  corrente       uma linha só, a onda do report. Para título de slide.
  q_mes          a pergunta do mês. É a exceção: ela muda de forma todo mês,
                 então o gráfico dela é o único que você remonta.

Regras que o código garante:

* Múltipla escolha é comparada TOKEN A TOKEN, separando por ';'. Sobra ou falta
  de separador é indiferente -- é o que corrige o erro de contagem antigo.
* O denominador é quem respondeu AQUELA pergunta, não o total da onda.
* Ondas até ULTIMA_ONDA_PUBLICADA usam o valor publicado. Nada do que foi ao ar
  muda.
* Alternativa nova que não é texto livre PARA a rodada e não grava nada.
"""
from __future__ import annotations

import csv
import statistics
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

from comum import (BLOCO_POR_ID, BLOCOS, CAMINHOS, LIMITE_CATCHALL, LIXO,
                   ONDAS_NA_SERIE, ONDAS_VIVAS, ULTIMA_ONDA_PUBLICADA, Log,
                   catchall_por_bloco, data_da_onda, indice_de_rotulos,
                   ler_registro, normalizar, onda_anterior, onda_de, slug,
                   tokens)

META_COLS = ('Id', 'Hora de início', 'Hora de conclusão', 'Email', 'Nome',
             'Survey')
MESES_PT = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun',
            'jul', 'ago', 'set', 'out', 'nov', 'dez']
MESES_EN = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
            'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


# --------------------------------------------------------------------------
# leitura das fontes
# --------------------------------------------------------------------------
def ler_forms(caminho: Path, log) -> dict:
    ws = openpyxl.load_workbook(caminho, data_only=True,
                                read_only=True).worksheets[0]
    it = ws.iter_rows(values_only=True)
    cab = [str(c).strip() if c is not None else '' for c in next(it)]
    col_fim = next((i for i, c in enumerate(cab)
                    if normalizar(c).startswith('hora de conclusao')), None)
    respostas, datas = [], []
    for linha in it:
        if all(v is None for v in linha):
            continue
        respostas.append(dict(zip(cab, linha)))
        if col_fim is not None and hasattr(linha[col_fim], 'year'):
            datas.append(linha[col_fim])
    if not respostas:
        raise SystemExit(f'ERRO: {caminho.name} não tem resposta nenhuma.')
    if not datas:
        raise SystemExit(f'ERRO: {caminho.name} não tem "Hora de conclusão" '
                         f'-- não consigo deduzir a onda.')
    onda = Counter(onda_de(d) for d in datas).most_common(1)[0][0]
    log(f'  {caminho.name}: {len(respostas)} respostas -> onda {onda}')
    return dict(onda=onda, cabecalho=cab, respostas=respostas)


def ler_raw_data(caminho: Path, log) -> list[dict]:
    ws = openpyxl.load_workbook(caminho, data_only=True,
                                read_only=True)['Raw Data']
    it = ws.iter_rows(values_only=True)
    cab = [str(c).strip() if c is not None else '' for c in next(it)]
    por_onda: dict[int, list[dict]] = defaultdict(list)
    for linha in it:
        if all(v is None for v in linha[:6]):
            continue
        survey = str(linha[0] or '').strip()
        if not survey.isdigit() or len(survey) != 6:
            continue
        por_onda[int(survey)].append(dict(zip(cab, linha)))
    log(f'  Raw Data: {sum(len(v) for v in por_onda.values())} respostas em '
        f'{len(por_onda)} ondas ({min(por_onda)}..{max(por_onda)})')
    return [dict(onda=o, cabecalho=cab, respostas=r)
            for o, r in sorted(por_onda.items())]


# --------------------------------------------------------------------------
# apuração
# --------------------------------------------------------------------------
def casar_colunas(cabecalho, respostas, log) -> tuple[dict, list[str]]:
    """{pergunta_id: coluna} + colunas que sobraram (a pergunta do mês).

    Uma coluna sobrando só conta se alguém a respondeu -- a Raw Data antiga
    carrega 74 colunas, quase todas vazias em cada onda.
    """
    achados, usadas = {}, set()
    for bloco in BLOCOS:
        for col in cabecalho:
            n = normalizar(col)
            if not n or col in usadas:
                continue
            if any(p in n for p in bloco['match']):
                achados[bloco['id']] = col
                usadas.add(col)
                break
    meta = {normalizar(m) for m in META_COLS}
    sobras = [c for c in cabecalho
              if c and c not in usadas and normalizar(c) not in meta
              and any(r.get(c) is not None and str(r.get(c)).strip()
                      for r in respostas)]
    return achados, sobras


def maior_prefixo(pid: str, texto: str, idx: dict) -> dict | None:
    """O rótulo registrado mais longo que prefixa `texto`.

    A Raw Data antiga perdeu os ';': "Nordeste;Norte;" virou "NordesteNorte".
    """
    melhor = None
    for (bid, rotulo), reg in idx.items():
        if bid != pid or not rotulo:
            continue
        if texto.startswith(rotulo) and (melhor is None or len(rotulo) > melhor[0]):
            melhor = (len(rotulo), reg)
    return melhor[1] if melhor else None


def apurar(bloco, coluna, respostas, idx, catchall, log) -> dict:
    contagem: Counter = Counter()
    desconhecidas: Counter = Counter()
    absorvidas: Counter = Counter()
    fora_de_safra: Counter = Counter()
    lixo: Counter = Counter()
    multi_em_unica = respondeu = 0
    numeros: list[float] = []

    for r in respostas:
        cel = r.get(coluna)
        if cel is None or str(cel).strip() == '':
            continue
        if normalizar(cel) in LIXO:
            lixo[normalizar(cel)] += 1
            continue                     # fora do numerador e do denominador
        respondeu += 1

        if bloco['tipo'] == 'escala':
            try:
                v = float(str(cel).replace(',', '.'))
            except ValueError:
                desconhecidas[str(cel)[:60]] += 1
                continue
            numeros.append(v)
            contagem[normalizar(str(int(v)) if v == int(v) else str(v))] += 1
            continue

        brutos = tokens(cel)
        if bloco['tipo'] == 'unica' and len(brutos) > 1:
            multi_em_unica += 1
            brutos = brutos[:1]

        for t in brutos:
            reg = idx.get((bloco['id'], t))
            if reg is None and bloco['tipo'] == 'unica':
                reg = maior_prefixo(bloco['id'], t, idx)
                if reg is not None:
                    multi_em_unica += 1
            if reg is not None:
                contagem[reg['alternativa_id']] += 1
            elif catchall:
                contagem[catchall] += 1
                absorvidas[t[:60]] += 1
            elif bloco.get('safra_rolante'):
                fora_de_safra[t[:80]] += 1
            else:
                desconhecidas[t[:60]] += 1

    if multi_em_unica:
        log(f'      {bloco["id"]}: {multi_em_unica} respostas com mais de uma '
            f'alternativa numa pergunta de escolha única -- ficou a primeira')
    if lixo:
        log(f'      {bloco["id"]}: {sum(lixo.values())} respostas descartadas '
            f'como lixo ({", ".join(sorted(lixo))}) -- fora do denominador')

    pcts = {a: c / respondeu for a, c in contagem.items()} if respondeu else {}
    return dict(pcts=pcts, desconhecidas=desconhecidas, absorvidas=absorvidas,
                fora_de_safra=fora_de_safra, denominador=respondeu,
                pcts_safra=({slug(k): v / respondeu
                             for k, v in fora_de_safra.items()}
                            if respondeu else {}),
                media=statistics.fmean(numeros) if numeros else None)


def media_ponderada(pcts, regs) -> float | None:
    """Resposta média de ibovespa_alvo, pelo ponto médio de cada faixa."""
    num = den = 0.0
    for r in regs:
        if not r.get('valor_num'):
            continue
        p = pcts.get(r['alternativa_id'])
        if p:
            num += p * float(r['valor_num'])
            den += p
    return num / den if den else None


# --------------------------------------------------------------------------
def ler_congelado(log) -> dict:
    caminho = CAMINHOS['congelado']
    if not caminho.exists():
        log(f'AVISO: {caminho.name} não existe -- rode congelar.py primeiro.')
        return dict(valores={}, medias={}, respondentes={})
    valores, medias, resp = {}, {}, {}
    with open(caminho, encoding='utf-8-sig', newline='') as fh:
        for r in csv.DictReader(fh):
            if r['extraordinaria'] == '1':
                continue
            onda = int(r['onda'])
            if r['alternativa_id'] == '__media__':
                medias[(onda, r['pergunta_id'])] = float(r['valor'])
            else:
                valores[(onda, r['pergunta_id'], r['alternativa_id'])] = \
                    float(r['valor'])
            if r['respondentes']:
                resp[onda] = int(r['respondentes'])
    log(f'congelado: {len(valores)} percentuais, {len(medias)} médias, '
        f'{len(resp)} ondas com respondentes')
    return dict(valores=valores, medias=medias, respondentes=resp)


def ler_ibovespa(log) -> dict[int, float]:
    caminho = CAMINHOS['ibovespa']
    if not caminho.exists():
        log(f'AVISO: {caminho.name} não existe -- a capa sai sem Ibovespa.')
        return {}
    out = {}
    with open(caminho, encoding='utf-8-sig', newline='') as fh:
        for r in csv.DictReader(fh):
            if r.get('fechamento'):
                out[int(r['onda'])] = float(r['fechamento'])
    return out


# --------------------------------------------------------------------------
def gravar_csvs(pasta: Path, abas: dict[str, tuple[list, list[list]]], log) -> None:
    """Uma tabela por CSV. É desta pasta que o Power Query lê.

    Por que CSV e não as abas do xlsx: cada consulta lê só o seu arquivo. Lendo
    do xlsx, as 27 consultas reparseavam a pasta de trabalho inteira cada uma,
    e o Atualizar Tudo passava de dois minutos. Com CSV cai para segundos.

    Formato fixo e sem ambiguidade: UTF-8 com BOM, vírgula, ponto decimal,
    data ISO. A consulta declara Culture="en-US" para casar com isso -- assim
    não importa a configuração regional da máquina.
    """
    pasta.mkdir(parents=True, exist_ok=True)
    escritos = 0
    for nome, (cab, linhas) in abas.items():
        with open(pasta / f'{nome}.csv', 'w', encoding='utf-8-sig',
                  newline='') as fh:
            w = csv.writer(fh, delimiter=',', quoting=csv.QUOTE_MINIMAL)
            w.writerow(list(cab))
            for l in linhas:
                w.writerow(['' if v is None
                            else f'{v:%Y-%m-%d}' if isinstance(v, datetime)
                            else repr(v) if isinstance(v, float)
                            else v
                            for v in l])
        escritos += 1
    log(f'>> {pasta.name}\\: {escritos} CSVs (a fonte do Power Query)')


def gravar(caminho: Path, abas: dict[str, tuple[list, list[list]]], log) -> None:
    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for nome, (cab, linhas) in abas.items():
        ws = wb.create_sheet(nome[:31])
        ws.append(list(cab))
        for l in linhas:
            ws.append(l)
        ws.freeze_panes = 'A2'
    wb.save(caminho)
    log(f'>> {caminho.name}: {len(abas)} abas, '
        f'{sum(len(v[1]) for v in abas.values())} linhas')


# --------------------------------------------------------------------------
def main() -> int:
    bootstrap = '--bootstrap' in sys.argv
    log = Log(f'rodada_{datetime.now():%Y%m%d_%H%M}.txt')
    registro = ler_registro()
    idx = indice_de_rotulos(registro)
    catchall = catchall_por_bloco(registro)
    reg_por_bloco: dict[str, list[dict]] = defaultdict(list)
    for r in registro:
        reg_por_bloco[r['pergunta_id']].append(r)
    for v in reg_por_bloco.values():
        v.sort(key=lambda r: r['ordem'])
    log(f'registro: {len(registro)} alternativas, {len(idx)} rótulos reconhecíveis')
    sem = [b['id'] for b in BLOCOS if b['id'] not in catchall]
    log(f'blocos com "Outra" para absorver texto livre: {len(catchall)}/{len(BLOCOS)}'
        + (f' | sem "Outra": {", ".join(sem)}' if sem else ''))

    # ---- fontes
    fontes: list[dict] = []
    if bootstrap:
        resto = [a for a in sys.argv[1:] if not a.startswith('--')]
        pa = Path(resto[0]) if resto else CAMINHOS['pa_principal']
        log(f'\nbootstrap a partir de {pa}')
        if not pa.exists():
            log(f'ERRO: não achei {pa}')
            return 1
        fontes += ler_raw_data(pa, log)
    pasta = CAMINHOS['input_forms']
    if pasta.exists():
        log(f'\ninput_forms: {pasta}')
        for f in sorted(pasta.glob('*.xlsx')):
            if not f.name.startswith('~$'):
                fontes.append(ler_forms(f, log))
    elif not bootstrap:
        log(f'ERRO: pasta não encontrada: {pasta}')
        return 1
    if not fontes:
        log('ERRO: nenhuma fonte para processar.')
        return 1

    # export do Forms ganha da Raw Data para a mesma onda
    por_onda: dict[int, dict] = {}
    for f in fontes:
        if f['onda'] in por_onda:
            log(f'  onda {f["onda"]}: vinha da Raw Data '
                f'({len(por_onda[f["onda"]]["respostas"])} respostas) -- fica o '
                f'export do Forms ({len(f["respostas"])})')
        por_onda[f['onda']] = f
    fontes = [por_onda[o] for o in sorted(por_onda)]

    # ---- apuração
    log('\napuração:')
    calc: dict[tuple, float] = {}
    calc_safra: dict[tuple, float] = {}
    rot_safra: dict[tuple, str] = {}
    medias_calc: dict[tuple, float] = {}
    respondentes: dict[int, int] = {}
    denominadores: dict[tuple, int] = {}
    q_mes: dict[int, list[dict]] = {}
    respostas_longas: list[list] = []
    problemas: list[str] = []

    for fonte in fontes:
        onda = fonte['onda']
        achados, sobras = casar_colunas(fonte['cabecalho'], fonte['respostas'], log)
        faltando = [b['id'] for b in BLOCOS if b['id'] not in achados]
        respondentes[onda] = len(fonte['respostas'])
        log(f'  onda {onda}: {len(achados)}/{len(BLOCOS)} perguntas casadas'
            + (f' | faltando: {", ".join(faltando)}' if faltando else '')
            + (f' | pergunta do mês: {len(sobras)}' if sobras else ''))

        for pid, coluna in achados.items():
            bloco = BLOCO_POR_ID[pid]
            res = apurar(bloco, coluna, fonte['respostas'], idx,
                         catchall.get(pid), log)
            absorv, den = sum(res['absorvidas'].values()), res['denominador']

            if den and absorv / den > LIMITE_CATCHALL:
                amostra = ', '.join(repr(k) for k, _ in
                                    res['absorvidas'].most_common(3))
                if onda <= ULTIMA_ONDA_PUBLICADA:
                    log(f'      {pid}: {absorv} marcações de {den} respondentes '
                        f'cairiam na "Outra" -- conjunto de alternativas diferente '
                        f'do registro nesta onda. Não recalculado; fica o '
                        f'publicado. ({amostra})')
                    continue
                problemas.append(
                    f'onda {onda} / {pid}: {absorv} marcações de {den} '
                    f'respondentes não casam com o registro.\n'
                    f'      Isso não é texto livre -- é conjunto de alternativas '
                    f'diferente.\n      {amostra}\n'
                    f'      Mapeie os rótulos em registro.csv (coluna aliases).')
                continue

            if res['desconhecidas']:
                problemas.append(
                    f'onda {onda} / {pid}: alternativa não registrada, e o bloco '
                    f'não tem "Outra" para absorver.\n      '
                    + '; '.join(f'{k!r} ({v}x)'
                                for k, v in res['desconhecidas'].most_common(6))
                    + '\n      Acrescente em registro.csv, ou ponha como alias '
                      'de uma alternativa existente.')
            if res['absorvidas']:
                log(f'      {pid}: {absorv} respostas de texto livre em "Outra" '
                    f'({len(res["absorvidas"])} distintas, ex.: '
                    f'{", ".join(repr(k) for k, _ in res["absorvidas"].most_common(2))})')
            if res['fora_de_safra']:
                log(f'      {pid}: {len(res["fora_de_safra"])} faixas de safra '
                    f'antiga -- vão para a base geral, não para os gráficos')
                for rot in res['fora_de_safra']:
                    rot_safra[(onda, pid, slug(rot))] = rot
                for aid, p in res['pcts_safra'].items():
                    calc_safra[(onda, pid, aid)] = p

            for aid, p in res['pcts'].items():
                calc[(onda, pid, aid)] = p
            if den:
                denominadores[(onda, pid)] = den
            if bloco['tipo'] == 'escala' and res['media'] is not None:
                medias_calc[(onda, pid)] = res['media']
            if pid == 'ibovespa_alvo':
                m = media_ponderada(res['pcts'], reg_por_bloco[pid])
                if m:
                    medias_calc[(onda, pid)] = m
            if den and den != len(fonte['respostas']):
                log(f'      {pid}: {den} de {len(fonte["respostas"])} responderam')

        if sobras:
            itens = []
            for col in sobras:
                c: Counter = Counter()
                original: dict[str, str] = {}
                n = 0
                for r in fonte['respostas']:
                    cel = r.get(col)
                    if cel is None or str(cel).strip() == '':
                        continue
                    n += 1
                    for bruto in str(cel).split(';'):
                        t = normalizar(bruto)
                        if not t:
                            continue
                        c[t] += 1
                        original.setdefault(t, str(bruto).strip().strip(';').strip())
                for rot, qtd in c.most_common():
                    itens.append(dict(pergunta=col, rotulo=original.get(rot, rot),
                                      qtd=qtd, respondentes=n,
                                      pct=qtd / n if n else 0))
            q_mes[onda] = itens

        for i, r in enumerate(fonte['respostas'], 1):
            for pid, coluna in achados.items():
                cel = r.get(coluna)
                if cel is not None and str(cel).strip():
                    respostas_longas.append([onda, i, pid, str(cel)[:250]])

    if problemas:
        log('\n' + '=' * 70)
        log('PAREI. Nada foi gravado -- a base não fica meio atualizada.')
        log('')
        for p in problemas[:15]:
            log(f'  {p}')
        log('=' * 70)
        log.gravar()
        return 2

    # ---- histórico: congelado manda nas ondas publicadas
    cong = ler_congelado(log)
    ibov = ler_ibovespa(log)
    v_cong, m_cong, r_cong = cong['valores'], cong['medias'], cong['respondentes']

    historico: list[list] = []
    valor_de: dict[tuple, float] = {}          # (onda,pid,aid) -> valor final
    reg_por_alt = {(r['pergunta_id'], r['alternativa_id']): r for r in registro}

    for chave in sorted(set(v_cong) | set(calc)):
        onda, pid, aid = chave
        congelada = onda <= ULTIMA_ONDA_PUBLICADA
        pub, c = v_cong.get(chave), calc.get(chave)
        valor = pub if (congelada and pub is not None) else c
        if valor is None:
            continue
        fonte_v = 'publicado' if (congelada and pub is not None) else 'calculado'
        r0 = reg_por_alt.get((pid, aid), {})
        valor_de[chave] = valor
        historico.append([
            onda, f'{onda // 100}-{onda % 100:02d}-01', pid,
            r0.get('serie_id', aid), aid, r0.get('rotulo_pt', ''),
            r0.get('rotulo_en', ''), valor,
            r_cong.get(onda) or respondentes.get(onda) or '',
            denominadores.get((onda, pid), ''), fonte_v,
            c if c is not None else '', pub if pub is not None else ''])

    for (onda, pid, aid), p in sorted(calc_safra.items()):
        historico.append([onda, f'{onda // 100}-{onda % 100:02d}-01', pid, aid,
                          aid, rot_safra.get((onda, pid, aid), ''), '', p,
                          respondentes.get(onda, ''), '', 'fora_de_safra', p, ''])

    medias_finais = {}
    for chave in set(m_cong) | set(medias_calc):
        onda, pid = chave
        congelada = onda <= ULTIMA_ONDA_PUBLICADA
        pub, c = m_cong.get(chave), medias_calc.get(chave)
        medias_finais[chave] = pub if (congelada and pub is not None) else c

    ondas = sorted({h[0] for h in historico})
    corrente = ondas[-1]
    n_pub = sum(1 for h in historico if h[10] == 'publicado')
    log(f'\nondas: {len(ondas)} ({ondas[0]}..{corrente})')
    log(f'valores: {len(historico)} ({n_pub} publicados, '
        f'{len(historico) - n_pub} calculados)')

    # ================================================================== #
    # as tabelas prontas para gráfico
    # ================================================================== #
    def resp_de(o):
        return r_cong.get(o) or respondentes.get(o) or ''

    tabelas: dict[str, tuple[list, list[list]]] = {}
    anterior = onda_anterior(corrente, ondas)
    log(f'\ntabelas de gráfico (onda {corrente}, comparando com {anterior}):')

    suspeitos = []
    for bloco in BLOCOS:
        pid = bloco['id']
        regs = reg_por_bloco[pid]

        # Quais alternativas entram no gráfico. A regra depende do tipo de
        # pergunta, porque "zero neste mês" significa coisas diferentes:
        #
        # ordenar='natural' (regiões, faixas, escala 0-10) -- o conjunto é uma
        #   enumeração fixa. Região sem resposta em julho tem que continuar na
        #   pizza. Então entram todas que tiveram resposta na janela LONGA;
        #   isso só descarta o que foi aposentado de fato (a faixa legada
        #   "0% a 25%", por exemplo).
        #
        # ordenar='valor' (rankings) -- o conjunto muda com o tempo e barra de
        #   0% não interessa. Então entram só as que a onda corrente usou de
        #   fato. É isto que tira as sete alternativas que a planilha antiga
        #   manteve zeradas desde o cutover de abr/2026.
        do_bruto = {aid for (o, p, aid) in calc if o == corrente and p == pid}
        if bloco['ordenar'] == 'valor':
            vivas = do_bruto or {
                r['alternativa_id'] for r in regs
                if (valor_de.get((corrente, pid, r['alternativa_id'])) or 0) > 0}
        else:
            recentes = ondas[-ONDAS_VIVAS:]
            vivas = {r['alternativa_id'] for r in regs
                     if any((valor_de.get((o, pid, r['alternativa_id'])) or 0) > 0
                            for o in recentes)} | do_bruto
        if not vivas:
            vivas = {r['alternativa_id'] for r in regs}

        # Um conceito renomeado tem duas alternativas com o mesmo serie_id e o
        # mesmo valor. No gráfico isso vira barra repetida -- fica só a que a
        # onda corrente usa.
        por_serie: dict[str, list[dict]] = defaultdict(list)
        for r in regs:
            por_serie[r['serie_id'] or r['alternativa_id']].append(r)
        escolhida = {}
        for sid, grupo in por_serie.items():
            atual_do_grupo = [r for r in grupo if r['alternativa_id'] in vivas]
            escolhida[sid] = (atual_do_grupo or grupo)[-1]

        # ---- d_<pergunta>: a distribuição da onda corrente
        linhas = []
        for sid, r in escolhida.items():
            aid = r['alternativa_id']
            if aid not in vivas:
                continue
            a = valor_de.get((corrente, pid, aid))
            b = valor_de.get((anterior, pid, aid)) if anterior else None
            if a is None and b is None:
                continue
            # O valor congelado e o recalculado do bruto discordando muito é o
            # sinal de que o publicado daquela célula está errado. Não mexo
            # nele -- publicado é publicado -- mas aviso, porque é com esse
            # número que você vai montar o gráfico agora.
            c_atual = calc.get((corrente, pid, aid))
            if (isinstance(a, float) and c_atual is not None
                    and abs(a - c_atual) > 0.20):
                suspeitos.append(
                    f'{pid}/{aid}: publicado {100 * a:.1f}% vs recalculado '
                    f'{100 * c_atual:.1f}% ({100 * (c_atual - a):+.1f} pp)')
            linhas.append([aid, r['rotulo_pt'], r['rotulo_en'], a, b,
                           (a - b) if (a is not None and b is not None) else None])
        if bloco['ordenar'] == 'valor':
            linhas.sort(key=lambda l: -(l[3] if isinstance(l[3], float) else -1))
        else:
            ordem_reg = {r['alternativa_id']: r['ordem'] for r in regs}
            linhas.sort(key=lambda l: ordem_reg.get(l[0], 999))
        for i, l in enumerate(linhas, 1):
            l.insert(0, i)
        tabelas[f'd_{pid}'] = (
            ['ordem', 'alternativa_id', 'rotulo_pt', 'rotulo_en',
             'atual', 'anterior', 'delta'], linhas)

        # ---- s_<pergunta>: série temporal, uma coluna por SÉRIE (não por
        # alternativa), para a renomeação não partir a linha do gráfico em duas
        janela = ondas[-ONDAS_NA_SERIE:]
        def valor_serie(o, sid):
            for r in por_serie[sid]:
                v = valor_de.get((o, pid, r['alternativa_id']))
                if v is not None:
                    return v
            return None

        series = [sid for sid in escolhida
                  if escolhida[sid]['alternativa_id'] in vivas
                  and any(valor_serie(o, sid) is not None for o in janela)]
        series.sort(key=lambda sid: escolhida[sid]['ordem'])
        cab = ['onda', 'data'] + [escolhida[sid]['rotulo_pt'] for sid in series]
        serie = []
        for o in janela:
            vals = [valor_serie(o, sid) for sid in series]
            if all(v is None for v in vals):
                continue
            # Numa pergunta de safra rolante, as ondas em que a pesquisa
            # perguntava sobre OUTRO ano deixam as faixas atuais quase vazias
            # (somam 8% em vez de 100%). Linha assim parece dado e não é.
            # O corte em 80% também tira os meses de faixas de 10 mil, antes
            # de abr/2026, que só preenchem parte das faixas de hoje.
            if (bloco.get('safra_rolante') and bloco['tipo'] == 'unica'
                    and sum(v for v in vals if v is not None) < 0.8):
                continue
            serie.append([o, data_da_onda(o)] + vals)
        tabelas[f's_{pid}'] = (cab, serie)
        log(f'  d_{pid}: {len(linhas)} alternativas | s_{pid}: {len(serie)} ondas '
            f'× {len(series)} séries')

    # ---- medias
    linhas = []
    for o in ondas[-ONDAS_NA_SERIE:]:
        s = medias_finais.get((o, 'sentimento'))
        i = medias_finais.get((o, 'ibovespa_alvo'))
        if s is None and i is None:
            continue
        linhas.append([o, data_da_onda(o), s, i, resp_de(o) or None])
    tabelas['medias'] = (['onda', 'data', 'sentimento_media',
                          'ibovespa_media', 'respondentes'], linhas)

    # ---- capa: proximos_meses + fechamento do Ibovespa
    regs_pm = reg_por_bloco['proximos_meses']
    cab = ['onda', 'data'] + [r['rotulo_pt'] for r in regs_pm] + ['ibovespa']
    linhas, sem_ibov = [], []
    for o in ondas[-ONDAS_NA_SERIE:]:
        if not any((o, 'proximos_meses', r['alternativa_id']) in valor_de
                   for r in regs_pm):
            continue
        linha = [o, data_da_onda(o)]
        for r in regs_pm:
            linha.append(valor_de.get((o, 'proximos_meses', r['alternativa_id'])))
        linha.append(ibov.get(o))
        if o not in ibov:
            sem_ibov.append(str(o))
        linhas.append(linha)
    tabelas['capa'] = (cab, linhas)
    if sem_ibov:
        log(f'  capa: sem fechamento do Ibovespa em {", ".join(sem_ibov[-6:])} '
            f'-- acrescente em ibovespa.csv (segue manual)')

    # ---- meta e corrente
    tabelas['meta'] = (
        ['onda', 'data', 'respondentes', 'regime'],
        [[o, data_da_onda(o), resp_de(o) or None,
          'publicado' if o <= ULTIMA_ONDA_PUBLICADA else 'calculado']
         for o in ondas])
    d = data_da_onda(corrente)
    tabelas['corrente'] = (
        ['onda', 'data', 'respondentes', 'mes_pt', 'mes_en'],
        [[corrente, d, resp_de(corrente) or None,
          f'{MESES_PT[d.month - 1]}/{d.year}', f'{MESES_EN[d.month - 1]} {d.year}']])

    # ---- q_mes: a exceção
    itens = q_mes.get(corrente, [])
    reais = [i for i in itens if i['qtd'] > 1]
    livres = [i for i in itens if i['qtd'] <= 1]
    linhas = [[i + 1, it['pergunta'][:200], it['rotulo'][:200], None, it['pct'],
               it['qtd']] for i, it in enumerate(sorted(reais,
                                                        key=lambda r: r['pct']))]
    if livres:
        soma = sum(i['pct'] for i in livres)
        linhas.append([len(linhas) + 1,
                       livres[0]['pergunta'][:200] if livres else '',
                       'Outra', 'Other', soma, sum(i['qtd'] for i in livres)])
        linhas.sort(key=lambda l: l[4])
        for i, l in enumerate(linhas, 1):
            l[0] = i
    tabelas['q_mes'] = (['ordem', 'pergunta', 'rotulo_pt', 'rotulo_en', 'pct',
                         'qtd'], linhas)
    if itens:
        log(f'  q_mes: {len(reais)} alternativas'
            + (f' + {len(livres)} de texto livre somadas em "Outra"' if livres else '')
            + ' -- a tradução para inglês (coluna rotulo_en) segue manual')

    if suspeitos:
        log('')
        log(f'  ATENÇÃO -- na onda {corrente} o valor publicado discorda do '
            f'recalculado em {len(suspeitos)} alternativa(s):')
        for s in suspeitos:
            log(f'    {s}')
        log('    As tabelas trazem o PUBLICADO (a onda está congelada). Se você')
        log('    quiser o número corrigido nestas células, é uma decisão de')
        log('    republicação -- baixe ULTIMA_ONDA_PUBLICADA no comum.py.')

    # ---- gravação
    log('')
    gravar(CAMINHOS['base_geral'], {
        'historico': (['onda', 'data', 'pergunta_id', 'serie_id',
                       'alternativa_id', 'rotulo_pt', 'rotulo_en', 'valor',
                       'respondentes', 'responderam_a_pergunta', 'fonte',
                       'pct_calculado', 'pct_publicado'], historico),
        'respostas': (['onda', 'respondente', 'pergunta_id', 'resposta'],
                      respostas_longas),
        'meta': tabelas['meta'],
    }, log)
    gravar(CAMINHOS['charts_data'], tabelas, log)
    gravar_csvs(CAMINHOS['charts_csv'], tabelas, log)
    log.gravar()
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
