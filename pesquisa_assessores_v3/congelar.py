"""Roda UMA vez, na instalação.

Extrai da PA Principal.xlsx o que só existe lá dentro e grava em texto:

  registro.csv                       -- perguntas e alternativas
  historico_congelado.csv            -- os percentuais publicados
  _saida/base_completa_publicada.csv -- todas as perguntas que já existiram

Depois disto a PA Principal não é mais lida por nada. Ela continua sendo
necessária como fonte, porém: a aba `Base` cobre 76 ondas e a `Raw Data` só 37,
então 39 ondas -- 3 anos e 5 meses -- não têm dado bruto para recalcular. Sem
ela não dá para regerar este CSV. **Não apague.**

Este é o único arquivo do pipeline que conhece endereço de célula da planilha
antiga. Está tudo em FAIXAS_BASE, abaixo.

    python congelar.py ["caminho\\da\\PA Principal.xlsx"]
"""
from __future__ import annotations

import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

from comum import (BASE_ANTIGA, CAMINHOS, Log, escrever_registro, normalizar,
                   onda_de, slug)

B = BASE_ANTIGA


def limpar_rotulo(texto: str) -> str:
    """Tira a sujeira que a Base antiga acumulou no texto do rótulo.

    Três rótulos de apetite_risco tinham um ";" colado no fim
    ("Cortes de juros no Brasil;"), e alguns trazem espaço-não-quebrável
    (U+00A0) no lugar do espaço. Isso apareceria no eixo do gráfico.

    Não afeta o casamento nem o alternativa_id: os dois passam por
    normalizar(), que já ignora ";" e U+00A0. É só o texto de exibição.
    """
    t = str(texto).replace(" ", " ").replace("​", "")
    t = re.sub(r"\s+", " ", t).strip()
    return t.rstrip(";").strip()


# --------------------------------------------------------------------------
# Onde cada pergunta mora na aba Base da PA Principal.
#
# Levantado lendo a planilha: o cabeçalho de cada bloco está na coluna C com a
# coluna A vazia, e as alternativas vêm abaixo. As faixas foram conferidas
# contra os intervalos que os 14 gráficos do deck liam.
#
#   linhas   : as alternativas
#   media    : linha da resposta média, quando existe
#   contagem : linhas de contagem absoluta (a escala guarda as duas coisas)
FAIXAS_BASE = {
    'regiao':                  dict(linhas=(8, 12)),
    'alocacao_rv':             dict(linhas=(16, 21)),
    'proximos_meses':          dict(linhas=(26, 28)),
    'classes_ativos':          dict(linhas=(35, 44)),
    'pct_internacional':       dict(linhas=(54, 58)),
    'interesse_internacional': dict(linhas=(62, 68)),
    'riscos_bolsa':            dict(linhas=(169, 188)),
    'setores':                 dict(linhas=(204, 217)),
    'sentimento':              dict(linhas=(234, 244), media=232,
                                    contagem=(221, 231)),
    'ibovespa_alvo':           dict(linhas=(456, 461), media=455),
    'apetite_risco':           dict(linhas=(320, 326)),
}

# --------------------------------------------------------------------------
# Renomeações confirmadas nos dados.
#
# Em abr/2026 o Forms reescreveu alternativas de riscos_bolsa. Cada par foi
# conferido: a linha antiga zera (ou espelha) a partir de 202604 e a nova começa
# exatamente ali.
#
# Elas NÃO viram o mesmo alternativa_id -- viram o mesmo serie_id. A série fica
# contínua para análise sem mexer em número publicado, e cada rótulo continua
# apurando na sua própria alternativa.
RENOMEACOES = [
    ('Política monetária mais dura nos mercados desenvolvidos',
     'Juros mais altos nos mercados desenvolvidos'),
    ('Juros mais altos no Brasil', 'Juros mais altos que o esperado o Brasil'),
    ('Instabilidade política', 'Instabilidade política/ Eleições'),
    ('Riscos geopolíticos', 'Riscos geopolíticos/ Guerra'),
]
# 'Crescimento econômico fraco na China' + 'Recessão nos EUA' parecem ter sido
# absorvidas por 'Desaceleração econômica global', mas é 2->1: não há conversão
# honesta, então ficam como séries separadas.

# Rótulos que o Forms usa hoje e que a Base nunca registrou -- o mesmo conceito
# escrito de outro jeito. Vão para a coluna aliases da alternativa existente.
ALIASES_EXTRA = {
    ('classes_ativos', 'Ouro e commodities'): ['Ouro'],
    ('interesse_internacional', 'Bonds'): ['Bonds e crédito'],
    ('interesse_internacional',
     'Não estão interessados por investimentos internacionais'):
        ['Não estão interessados'],
}

# Alternativas que existem no Forms e não têm linha na Base.
#   (pergunta_id, rotulo_pt, rotulo_en)
NOVAS = [
    ('setores', 'Não estão interessados', 'Not interested'),
]

# --------------------------------------------------------------------------
# Rótulos da coluna C que estão ERRADOS na Base. Chaveado por linha.
#
# A linha 323 de apetite_risco tem na coluna C "Melhora na recuperação
# econômica global;" -- o mesmo texto da linha 325. Mas a coluna A, que é por
# onde as fórmulas do deck casam, diz "Melhora na recuperação econômica DA
# CHINA". São duas alternativas diferentes, e o dado bruto confirma: "Melhora
# na recuperação econômica da China" aparece como resposta nas ondas de 2024.
#
# Sem esta correção as duas colapsam numa alternativa só e o valor publicado
# sai errado -- 12,15% (a linha de cima) em vez dos 21,5% que o deck mostra.
CORRIGIR_ROTULO = {
    ('apetite_risco', 323): 'Melhora na recuperação econômica da China',
}

# --------------------------------------------------------------------------
# Legendas em inglês, copiadas das células auxiliares da aba Charts -- que é
# de onde o deck em inglês as lê. São elas que foram publicadas.
#
# A coluna A da Base não serve para isto: em ibovespa_alvo ela guarda o ponto
# médio numérico (150000), em interesse_internacional está vazia, e em
# apetite_risco tem o texto em PORTUGUÊS. Nos três casos o inglês só existe
# na Charts.
ROTULOS_EN = {
    'ibovespa_alvo': {
        'Abaixo de 150 mil pontos': 'Below 150k',
        'Entre 150 mil e 170 mil pontos': 'Between 150k - 170k',
        'Entre 170 mil e 190 mil pontos': 'Between 170k - 190k',
        'Entre 190 e 200 mil pontos': 'Between 190k - 200k',
        'Entre 200 e 210 mil pontos': 'Between 200k - 210k',
        'Acima de 210 mil pontos': 'Above 210k points',
    },
    'interesse_internacional': {
        'Fundos Internacionais': 'International Funds',
        'Ações internacionais': 'BDRs',
        'ETFs': 'ETFs',
        'Dólar': 'Dollar',
        'Bonds': 'Bonds',
        'Não estão interessados por investimentos internacionais': 'Not interested',
        'Outros': 'Other',
    },
    'apetite_risco': {
        'Cortes de juros no Brasil': 'Rate cuts in Brazil',
        'Queda de juros nos EUA e no mundo': 'Rate cuts in the US and globally',
        'Mercado voltando a ter performance sólida':
            'Market returning to show solid performance',
        'Melhora na recuperação econômica global':
            'Improvement in global economy recovery',
        'Mudança de rumo na política econômica':
            'Change of direction in political economy',
        'Outra': 'Other',
    },
    # riscos_bolsa e classes_ativos TÊM inglês na coluna A da Base, mas com
    # outra redação. Aqui fica a do deck, que é a publicada.
    'riscos_bolsa': {
        'Desaceleração econômica global': 'Global economic slowdown',
        'Juros mais altos nos mercados desenvolvidos':
            'Higher rates in developed markets',
        'Riscos fiscais no Brasil': 'Fiscal risks in Brazil',
        'Inflação em alta no Brasil': 'Rising inflation in Brazil',
        'Juros mais altos que o esperado o Brasil':
            'Higher-than-expected local rates',
        'Dólar mais alto': 'Stronger US dollar',
        'Riscos geopolíticos/ Guerra': 'Geopolitical risks / war',
        'Instabilidade política/ Eleições': 'Political instability / elections',
        'Choque do petróleo': 'Oil shock',
    },
    'classes_ativos': {
        'Ouro e commodities': 'Gold and commodities',
        'Investimentos Internacionais (dólar, fundos, ETFs, etc)':
            'International Investments (USD, funds, ETFs, etc)',
    },
}

# Alternativas que o deck NÃO mostra no gráfico, embora sejam apuradas.
# Conferido nas células auxiliares da Charts: riscos mostra 9 alternativas,
# classes 9 e interesse 6 -- em nenhuma delas o "Outra" aparece. Já em
# apetite_risco ele aparece, como "Other". Por isso é por alternativa e não
# uma regra geral.
FORA_DO_GRAFICO = [
    ('classes_ativos', 'Outra'),
    ('riscos_bolsa', 'Outra'),
    ('interesse_internacional', 'Outros'),
]


# --------------------------------------------------------------------------
def ler_ondas(ws, log) -> list[dict]:
    """Eixo de ondas, lido pela LINHA DA DATA.

    A linha 1 (código da onda) só ganhou valor a partir da 38ª coluna -- ler por
    ela perde 39 ondas em silêncio.
    """
    import re
    ondas = []
    vistos: dict[int, int] = defaultdict(int)
    for col in range(B['primeira_col'], ws.max_column + 1):
        data = ws.cell(B['linha_data'], col).value
        if not hasattr(data, 'year'):
            continue
        onda = onda_de(data)
        vistos[onda] += 1
        resp = ws.cell(B['linha_resp'], col).value
        if isinstance(resp, str):                       # "590 respostas"
            m = re.search(r'\d+', resp)
            resp = int(m.group()) if m else None
        ondas.append(dict(col=col, data=data, onda=onda,
                          extraordinaria=vistos[onda] > 1, respondentes=resp))
    log(f'ondas: {len(ondas)}  ({ondas[0]["data"]:%Y-%m} a {ondas[-1]["data"]:%Y-%m})')
    for d in [o for o in ondas if o['extraordinaria']]:
        log(f'  aviso: onda {d["onda"]} tem 2ª coluna em {d["data"]:%Y-%m-%d} '
            f'-> extraordinaria=1 (edição extra, não substitui a regular)')
    sem = [o['onda'] for o in ondas if not isinstance(o['respondentes'], int)]
    if sem:
        log(f'  aviso: {len(sem)} ondas sem nº de respondentes na linha 2: '
            f'{sem[:8]}{"..." if len(sem) > 8 else ""}')
    return ondas


def montar_registro(ws, log) -> tuple[list[dict], dict]:
    """Registro das perguntas recorrentes. -> (registro, linhas_por_alternativa)

    linhas_por_alternativa guarda de que linha(s) da Base cada alternativa saiu.
    Serve só para o congelamento; não vai para o registro.csv.
    """
    registro: list[dict] = []
    de_onde: dict[tuple, list[int]] = defaultdict(list)

    for pid, faixa in FAIXAS_BASE.items():
        ini, fim = faixa['linhas']
        vistos: dict[str, str] = {}
        ordem = 0
        for linha in range(ini, fim + 1):
            pt = ws.cell(linha, B['col_pt']).value
            en = ws.cell(linha, B['col_en']).value
            pt_txt = limpar_rotulo(pt) if pt is not None else ''
            en_txt = limpar_rotulo(en) if en is not None else ''
            corrigido = CORRIGIR_ROTULO.get((pid, linha))
            if corrigido:
                log(f'  rótulo corrigido L{linha}: {pt_txt[:40]!r} -> '
                    f'{corrigido[:40]!r} (a coluna A da Base, que o deck casa)')
                pt_txt = corrigido
            if not pt_txt and not en_txt:
                continue
            chave = normalizar(pt_txt) or normalizar(en_txt)
            alt_id = vistos.get(chave)
            if alt_id is None:
                alt_id = slug(pt_txt or en_txt)
                vistos[chave] = alt_id
                # valor_num: o ponto médio numérico que a coluna A guarda em
                # ibovespa_alvo, usado para a resposta média
                valor_num = ''
                import re as _re
                if en_txt and _re.fullmatch(r'\d{4,7}', en_txt):
                    valor_num, en_txt = en_txt, ''
                ordem += 1
                registro.append(dict(
                    pergunta_id=pid, alternativa_id=alt_id, serie_id=alt_id,
                    ordem=ordem, rotulo_pt=pt_txt, rotulo_en=en_txt,
                    aliases='', valor_num=valor_num, ativa=True,
                    no_grafico=True))
            de_onde[(pid, alt_id)].append(linha)

    espelhadas = {k: v for k, v in de_onde.items() if len(v) > 1}
    for (pid, aid), linhas in espelhadas.items():
        log(f'  {pid}/{aid}: mesmo rótulo em {len(linhas)} linhas '
            f'({", ".join(map(str, linhas))})')

    # renomeações: serie_id compartilhado, alias só na linha NOVA
    por_rotulo = {(r['pergunta_id'], normalizar(r['rotulo_pt'])): r
                  for r in registro}
    for velho, novo in RENOMEACOES:
        rv = rn = None
        for (pid, rot), r in por_rotulo.items():
            if rot == normalizar(velho):
                rv = r
            elif rot == normalizar(novo):
                rn = r
        if not rv or not rn:
            log(f'  aviso: renomeação não encontrada: {velho!r} -> {novo!r}')
            continue
        rv['serie_id'] = rn['serie_id'] = rn['alternativa_id']
        # O alias vai só na alternativa NOVA. Se fosse também na antiga, o
        # rótulo novo do Forms resolveria para a alternativa antiga (que vem
        # primeiro no registro) e a série se partiria ao contrário.
        rn['aliases'] = velho
        log(f'  série emendada: {velho[:40]!r} -> {novo[:40]!r}')

    for (pid, rot), extras in ALIASES_EXTRA.items():
        alvo = por_rotulo.get((pid, normalizar(rot)))
        if alvo is None:
            log(f'  aviso: alias extra sem alternativa: {pid}/{rot!r}')
            continue
        atuais = [a for a in (alvo['aliases'] or '').split('|') if a]
        alvo['aliases'] = '|'.join(atuais + extras)
        log(f'  alias: {pid}/{rot[:34]!r} <- {extras}')

    for pid, pt, en in NOVAS:
        if (pid, normalizar(pt)) in por_rotulo:
            continue
        ordem = max((r['ordem'] for r in registro if r['pergunta_id'] == pid),
                    default=0) + 1
        registro.append(dict(
            pergunta_id=pid, alternativa_id=slug(pt), serie_id=slug(pt),
            ordem=ordem, rotulo_pt=pt, rotulo_en=en, aliases='',
            valor_num='', ativa=True, no_grafico=True))
        log(f'  nova: {pid}/{pt!r} (não existia na Base)')

    # legendas em inglês do deck
    n_en = 0
    for pid, mapa in ROTULOS_EN.items():
        for pt, en in mapa.items():
            alvo = por_rotulo.get((pid, normalizar(pt)))
            if alvo is None:
                log(f'  aviso: legenda EN sem alternativa: {pid}/{pt[:40]!r}')
                continue
            if alvo['rotulo_en'] != en:
                alvo['rotulo_en'] = en
                n_en += 1
    log(f'  legendas em inglês do deck aplicadas: {n_en}')

    # alternativas que o deck não mostra no gráfico
    for pid, pt in FORA_DO_GRAFICO:
        alvo = por_rotulo.get((pid, normalizar(pt)))
        if alvo is None:
            log(f'  aviso: exclusão sem alternativa: {pid}/{pt!r}')
            continue
        alvo['no_grafico'] = False
        log(f'  fora do gráfico: {pid}/{pt!r} (o deck não mostra)')

    # rótulo puramente numérico (a escala 0-10) é igual nos dois idiomas
    sem_en = [(r['pergunta_id'], r['rotulo_pt'][:38]) for r in registro
              if not r['rotulo_en'] and r['no_grafico']
              and not r['rotulo_pt'].strip().isdigit()]
    if sem_en:
        log(f'  AVISO: {len(sem_en)} alternativas de gráfico sem legenda em '
            f'inglês -- o deck EN vai sair sem rótulo nelas:')
        for pid, pt in sem_en[:12]:
            log(f'    {pid}/{pt!r}')

    log(f'registro: {len(registro)} alternativas em {len(FAIXAS_BASE)} perguntas')
    return registro, de_onde


def congelar_valores(ws, ondas, registro, de_onde, log) -> list[dict]:
    """Os percentuais publicados, um por (onda, pergunta, alternativa)."""
    saida: dict[tuple, dict] = {}
    conflitos = 0
    for r in registro:
        chave_de_onde = (r['pergunta_id'], r['alternativa_id'])
        for linha in de_onde.get(chave_de_onde, []):
            for o in ondas:
                v = ws.cell(linha, o['col']).value
                if v is None or isinstance(v, str):
                    continue
                k = (o['onda'], int(o['extraordinaria']),
                     r['pergunta_id'], r['alternativa_id'])
                novo = dict(onda=o['onda'],
                            extraordinaria=int(o['extraordinaria']),
                            data=f'{o["data"]:%Y-%m-%d}',
                            pergunta_id=r['pergunta_id'],
                            alternativa_id=r['alternativa_id'],
                            linha_origem=linha, valor=v,
                            respondentes=o['respondentes'] or '')
                if k not in saida:
                    saida[k] = novo
                elif abs(saida[k]['valor'] - v) > 1e-9:
                    conflitos += 1
                    if conflitos <= 6:
                        log(f'  conflito {k[2]}/{k[3]} onda {k[0]}: '
                            f'L{saida[k]["linha_origem"]}={saida[k]["valor"]:.4f} '
                            f'vs L{linha}={v:.4f} -> fica a de cima')
    if conflitos:
        log(f'  {conflitos} conflitos entre linhas espelhadas. O MATCH das '
            f'fórmulas do deck usava a de cima, então é ela que foi publicada.')

    # médias e contagens: linhas soltas, sem alternativa
    extras = []
    for pid, faixa in FAIXAS_BASE.items():
        if 'media' not in faixa:
            continue
        for o in ondas:
            v = ws.cell(faixa['media'], o['col']).value
            if v is None or isinstance(v, str):
                continue
            extras.append(dict(onda=o['onda'],
                               extraordinaria=int(o['extraordinaria']),
                               data=f'{o["data"]:%Y-%m-%d}', pergunta_id=pid,
                               alternativa_id='__media__',
                               linha_origem=faixa['media'], valor=v,
                               respondentes=o['respondentes'] or ''))
    log(f'congelado: {len(saida)} percentuais + {len(extras)} médias')
    return sorted(list(saida.values()) + extras,
                  key=lambda s: (s['pergunta_id'], s['alternativa_id'], s['onda']))


def base_completa(ws, ondas, log) -> list[dict]:
    """Toda pergunta que já existiu na Base -- vai para a base geral."""
    def tem_dado(linha):
        return any(ws.cell(linha, o['col']).value is not None for o in ondas)

    saida, pergunta, n = [], None, 0
    for linha in range(5, ws.max_row + 1):
        pt = ws.cell(linha, B['col_pt']).value
        en = ws.cell(linha, B['col_en']).value
        pt_txt = str(pt).strip() if pt is not None else ''
        if pt_txt and en is None and not tem_dado(linha) and len(pt_txt) > 25:
            pergunta, n = pt_txt, n + 1
            continue
        if not pergunta or not tem_dado(linha):
            continue
        if normalizar(pt_txt) in ('total', ''):
            continue
        for o in ondas:
            v = ws.cell(linha, o['col']).value
            if v is None or isinstance(v, str):
                continue
            saida.append(dict(onda=o['onda'],
                              extraordinaria=int(o['extraordinaria']),
                              data=f'{o["data"]:%Y-%m-%d}',
                              pergunta=pergunta[:180], rotulo_pt=pt_txt,
                              rotulo_en=str(en).strip() if en is not None else '',
                              valor=v, respondentes=o['respondentes'] or ''))
    log(f'base completa: {len(saida)} valores, {n} perguntas já feitas')
    return saida


def gravar_csv(caminho: Path, linhas: list[dict]) -> None:
    caminho.parent.mkdir(parents=True, exist_ok=True)
    with open(caminho, 'w', encoding='utf-8-sig', newline='') as fh:
        w = csv.DictWriter(fh, fieldnames=list(linhas[0].keys()))
        w.writeheader()
        w.writerows(linhas)
    print(f'>> {caminho}  ({len(linhas)} linhas)')


def main() -> int:
    origem = Path(sys.argv[1]) if len(sys.argv) > 1 else CAMINHOS['pa_principal']
    log = Log('congelamento.txt')
    log(f'origem: {origem}')
    if not origem.exists():
        log(f'ERRO: não achei {origem}')
        log('  se a rede estiver fora, passe o caminho de uma cópia local.')
        return 1

    ws = openpyxl.load_workbook(origem, data_only=True)['Base']
    ondas = ler_ondas(ws, log)
    log('')
    registro, de_onde = montar_registro(ws, log)
    log('')
    congelado = congelar_valores(ws, ondas, registro, de_onde, log)
    log('')
    completa = base_completa(ws, ondas, log)

    escrever_registro(registro)
    print(f'>> {CAMINHOS["registro"]}  ({len(registro)} linhas)')
    gravar_csv(CAMINHOS['congelado'], congelado)
    gravar_csv(CAMINHOS['saida'] / 'base_completa_publicada.csv', completa)

    # semente do fechamento do Ibovespa: estava digitado à mão na coluna E da
    # aba 'Gráfico capa'. Vira arquivo de texto, e é nele que você acrescenta
    # uma linha por mês -- o único dado do report que não sai da pesquisa.
    ibov = ibovespa_da_capa(openpyxl.load_workbook(origem, data_only=True), log)
    if ibov and not CAMINHOS['ibovespa'].exists():
        gravar_csv(CAMINHOS['ibovespa'], ibov)
    elif ibov:
        log(f'  {CAMINHOS["ibovespa"].name} já existe -- não sobrescrevi. '
            f'({len(ibov)} fechamentos disponíveis no template, se precisar)')

    log.gravar()
    return 0


def ibovespa_da_capa(wb, log) -> list[dict]:
    """O fechamento do Ibovespa que estava digitado na aba da capa."""
    if 'Gráfico capa' not in wb.sheetnames:
        log('  aviso: aba "Gráfico capa" não encontrada -- sem semente do Ibovespa')
        return []
    ws = wb['Gráfico capa']
    out = {}
    for linha in range(2, ws.max_row + 1):
        d = ws.cell(linha, 1).value
        v = ws.cell(linha, 5).value
        if hasattr(d, 'year') and isinstance(v, (int, float)):
            out[onda_de(d)] = dict(onda=onda_de(d), data=f'{d:%Y-%m-%d}',
                                   fechamento=v)
    log(f'ibovespa: {len(out)} fechamentos extraídos da aba da capa')
    return [out[k] for k in sorted(out)]


if __name__ == '__main__':
    raise SystemExit(main())
