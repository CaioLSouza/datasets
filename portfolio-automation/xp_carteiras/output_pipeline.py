"""Gera somente os arquivos Excel operacionais da pasta de output."""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook

from .components import gerar_decomposicoes, gerar_tabelas_componentes
from .constants import (
    MES_LBL_PT,
    arq_base100,
    arq_comp,
    arq_decomp,
    arq_fig2,
    arq_performance,
    cols_pct,
    cols_pct_decomp,
    mapa_arquivo,
    nomes_exib,
    portfolio_names,
)
from .excel_reports import _preencher_sheet, _preencher_sheet_flat, exportar_tabela_retornos
from .performance import indicadores_12m, tabela_retornos
from .pipeline_data import PipelineContext


def _export_base100(context: PipelineContext) -> None:
    output_dir = context.settings.output_dir
    for portfolio in portfolio_names:
        df_b100 = context.resultado_dfs.get(portfolio)
        if df_b100 is None:
            print(f"[AVISO] base 100 não exportada (carteira não processada): {portfolio}")
            continue
        caminho = output_dir / f"{arq_base100[portfolio]}.xlsx"
        with pd.ExcelWriter(caminho, engine="xlsxwriter") as writer:
            df_b100.to_excel(writer, sheet_name="base100")
            workbook = writer.book
            worksheet = writer.sheets["base100"]
            fmt_data = workbook.add_format({"num_format": "dd/mm/yyyy", "align": "center"})
            fmt_num = workbook.add_format({"num_format": "0.00", "align": "center"})
            worksheet.set_column(0, 0, 12, fmt_data)
            worksheet.set_column(1, len(df_b100.columns), 16, fmt_num)
        print(f"Base 100 exportada: {caminho}")


def _export_metrics_and_performance(context: PipelineContext) -> None:
    output_dir = context.settings.output_dir
    for portfolio in portfolio_names:
        metrics = indicadores_12m(context.resultado_dfs[portfolio], context.cdi)
        caminho = output_dir / f"{arq_fig2[portfolio]}.xlsx"
        metrics.to_excel(caminho)
        print(f"Exportado: {caminho}")

        table_raw, _ = tabela_retornos(context.resultado_dfs[portfolio])
        table_raw = table_raw.rename(index={portfolio: nomes_exib[portfolio]})
        exportar_tabela_retornos(table_raw, output_dir / arq_performance[portfolio])


def _export_components(context: PipelineContext) -> None:
    output_dir = context.settings.output_dir
    for portfolio in portfolio_names:
        atual, ultimo, anterior_mtd = gerar_tabelas_componentes(
            context.composition_dict[portfolio],
            context.market_data,
            context.mapa_nome,
            context.mapa_setor,
        )
        variantes = (
            ("atual", atual),
            ("ultimo_rebal", ultimo),
            ("comp_mes_passado_mtd_atual", anterior_mtd),
        )
        for sufixo, table in variantes:
            caminho = output_dir / f"componentes_{mapa_arquivo[portfolio]}_{sufixo}.xlsx"
            with pd.ExcelWriter(caminho, engine="xlsxwriter") as writer:
                table.to_excel(writer, index=False, sheet_name="componentes")
                workbook = writer.book
                worksheet = writer.sheets["componentes"]
                fmt_pct = workbook.add_format({"num_format": "0.0%"})
                for column in cols_pct:
                    index = table.columns.get_loc(column)
                    worksheet.set_column(index, index, 14, fmt_pct)


def _export_compositions(context: PipelineContext) -> None:
    output_dir = context.settings.output_dir
    with pd.ExcelWriter(output_dir / "sector_weights_ibovespa.xlsx") as writer:
        context.sector_weight_ibovespa.to_excel(
            writer, sheet_name="Individual Sectors", index=False
        )
        context.segment_weight_ibovespa.to_excel(
            writer, sheet_name="Sector Groups", index=False
        )

    for portfolio in portfolio_names:
        nome = nomes_exib[portfolio]
        df_en = context.dfs_composicao[portfolio]["EN"]
        df_pt = context.dfs_composicao[portfolio]["PT"]
        workbook = Workbook()
        workbook.remove(workbook.active)
        _preencher_sheet(
            workbook.create_sheet(title="ENG"), df_en, f"{nome} Portfolio", idioma="EN"
        )
        _preencher_sheet(
            workbook.create_sheet(title="PT"), df_pt, f"Carteira {nome}", idioma="PT"
        )
        _preencher_sheet_flat(workbook.create_sheet(title="ENG_data"), df_en, idioma="EN")
        _preencher_sheet_flat(workbook.create_sheet(title="PT_data"), df_pt, idioma="PT")
        workbook.save(output_dir / f"{arq_comp[portfolio]}.xlsx")
        print(f"Composição gerada (visual + data): {nome}")


def _export_return_attribution(context: PipelineContext) -> None:
    output_dir = context.settings.output_dir
    for portfolio in portfolio_names:
        atual, anterior = gerar_decomposicoes(
            context.composition_dict[portfolio],
            context.market_data,
            context.mapa_nome,
            context.mapa_setor,
        )
        data_max = context.market_data["data"].max()
        lbl_atual = f"{MES_LBL_PT[data_max.month]}-{str(data_max.year)[2:]}"
        ref_ant = pd.Timestamp(data_max.year, data_max.month, 1) - pd.Timedelta(days=1)
        lbl_ant = f"{MES_LBL_PT[ref_ant.month]}-{str(ref_ant.year)[2:]}"
        caminho = output_dir / arq_decomp[portfolio]
        with pd.ExcelWriter(caminho, engine="xlsxwriter") as writer:
            workbook = writer.book
            fmt_pct = workbook.add_format({"num_format": "0.0%"})
            for sheet, table in (
                (f"mes_atual_{lbl_atual}", atual),
                (f"mes_anterior_{lbl_ant}", anterior),
            ):
                table.to_excel(writer, index=False, sheet_name=sheet)
                worksheet = writer.sheets[sheet]
                for column in cols_pct_decomp:
                    index = table.columns.get_loc(column)
                    worksheet.set_column(index, index, 14, fmt_pct)
                worksheet.set_column(0, 0, 24)
                worksheet.set_column(2, 2, 22)
        print(f"Decomposição de retorno exportada: {caminho}")


def generate_output_files(context: PipelineContext) -> None:
    """Atualiza todos os Excel de output sem abrir ou gerar PowerPoint."""
    context.settings.output_dir.mkdir(parents=True, exist_ok=True)
    _export_base100(context)
    _export_metrics_and_performance(context)
    _export_components(context)
    _export_compositions(context)
    _export_return_attribution(context)
