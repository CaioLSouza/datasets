"""Pipeline do pipeline de carteiras XP."""

from __future__ import annotations

import os
import pandas as pd
from openpyxl import Workbook

from .components import calcular_pesos_ibovespa, gerar_decomposicoes, gerar_tabelas_componentes, montar_df_composicao
from .excel_reports import _preencher_sheet, _preencher_sheet_flat, exportar_tabela_retornos
from .performance import _df_para_lamina, calcular_performance, indicadores_12m, indice_base100, tabela_retornos
from .powerpoint_reports import atualizar_ppt
from .constants import (
    MES_LBL_PT,
    arq_base100,
    arq_comp,
    arq_decomp,
    arq_fig2,
    arq_performance,
    benchmarks_por_carteira,
    cols_pct,
    cols_pct_decomp,
    mapa_arquivo,
    nomes_exib,
    portfolio_names,
    sector_to_segment,
)
from .monthly_config import commercial_ppt_config
from .settings import Settings, load_settings

def main(settings: Settings | None = None) -> None:
    """Executa o pipeline completo e grava os artefatos configurados."""
    settings = settings or load_settings()
    settings.output_dir.mkdir(parents=True, exist_ok=True)

    # --- Carga, preparação e performance ---

    xpqs = pd.read_excel(settings.sector_classification_path)[['cod_ativo', 'name', 'adjusted_GICS_sector', 'sector_xp']]

    market_data = pd.read_parquet(
        settings.market_data_path
    )

    bdr_market_data = pd.read_csv(
        settings.bdr_market_data_path
    )

    bdr_market_data.rename(columns={'Ativo': 'cod_ativo', 'Data': 'data'}, inplace=True)

    bdr_market_data['cod_ativo'] = bdr_market_data['cod_ativo'].str.replace("<XBSP>", "")

    bdr_market_data['data'] = pd.to_datetime(bdr_market_data['data'])

    valid_dates = market_data['data'].unique()

    bdr_market_data = bdr_market_data[bdr_market_data['data'].isin(valid_dates)].copy()

    bdr_market_data['adj_close_price'] = pd.to_numeric(bdr_market_data['adj_close_price'], errors='coerce')

    market_data = pd.concat([market_data, bdr_market_data], ignore_index=True)

    indices = pd.read_parquet(settings.indices_path)

    indices = indices.loc[indices['cod_ativo'].isin(['Ibovespa', 'SMLL', 'ISEE'])].copy()

    indices['data'] = pd.to_datetime(indices['data'])

    file_path = settings.performance_workbook_path

    composition_dict = {}

    for portfolio in portfolio_names:
        df = pd.read_excel(file_path, sheet_name=portfolio, skiprows=5)

        if portfolio == 'Carteira - TOP Ações XP':
            df = df.iloc[:, 89:]
            df.rename(columns={'Ticker.1': 'cod_ativo'}, inplace=True)
        else:
            df = df.iloc[:, 3:]
            df.rename(columns={'Ticker': 'cod_ativo'}, inplace=True)

        composition_dict[portfolio] = df

    md = market_data.copy()

    md['data'] = pd.to_datetime(md['data'])

    md = md[['cod_ativo', 'data', 'adj_close_price']].dropna()

    md = md.sort_values(['cod_ativo', 'data'])

    md['ret'] = md.groupby('cod_ativo')['adj_close_price'].pct_change()

    resultado_dfs = {}

    for portfolio in portfolio_names:
        curva_carteira = calcular_performance(composition_dict[portfolio], md)

        df_port = pd.DataFrame({portfolio: curva_carteira})

        for bench in benchmarks_por_carteira[portfolio]:
            df_port[bench] = indice_base100(bench, curva_carteira.index, indices)

        df_port.index.name = 'data'
        resultado_dfs[portfolio] = df_port

    output_dir = str(settings.output_dir)

    for portfolio in portfolio_names:
        df_b100 = resultado_dfs.get(portfolio)
        if df_b100 is None:
            print(f"[AVISO] base 100 não exportada (carteira não processada): {portfolio}")
            continue
     
        caminho = f"{output_dir}\\{arq_base100[portfolio]}.xlsx"
     
        with pd.ExcelWriter(caminho, engine='xlsxwriter') as writer:
            # exporta a curva base 100 (carteira + benchmarks) com a data no índice
            df_b100.to_excel(writer, sheet_name='base100')
     
            wb = writer.book
            ws = writer.sheets['base100']
     
            fmt_data  = wb.add_format({'num_format': 'dd/mm/yyyy', 'align': 'center'})
            fmt_num   = wb.add_format({'num_format': '0.00', 'align': 'center'})
     
            # coluna A = datas (índice)
            ws.set_column(0, 0, 12, fmt_data)
            # demais colunas = valores base 100
            ws.set_column(1, len(df_b100.columns), 16, fmt_num)
     
        print(f"Base 100 exportada: {caminho}")

    indices = pd.read_parquet(settings.indices_path)

    cdi = (indices.loc[indices['cod_ativo'] == 'CDI Acumulado', ['data', 'close_price']]
                  .drop_duplicates('data')
                  .set_index('data')['close_price']
                  .sort_index())

    for portfolio in portfolio_names:
        tab = indicadores_12m(resultado_dfs[portfolio], cdi)
        caminho = f"{output_dir}\\{arq_fig2[portfolio]}.xlsx"
        tab.to_excel(caminho)
        print(f"Exportado: {caminho}")

    # --- Tabelas de retorno e componentes ---

    for portfolio in portfolio_names:
        table_raw, _ = tabela_retornos(resultado_dfs[portfolio])
        table_raw = table_raw.rename(index={portfolio: nomes_exib[portfolio]})
        exportar_tabela_retornos(
            table_raw,
            f"{output_dir}\\{arq_performance[portfolio]}",
        )

    mapa_info = xpqs.drop_duplicates('cod_ativo').set_index('cod_ativo')

    mapa_nome   = mapa_info['name'].to_dict()

    mapa_setor  = mapa_info['sector_xp'].to_dict()

    output_dir = str(settings.output_dir)

    tabelas_componentes = {}

    for portfolio in portfolio_names:
        tab_atual, tab_ultimo, tab_comp_anterior_mtd_atual = gerar_tabelas_componentes(
            composition_dict[portfolio], md, mapa_nome, mapa_setor
        )

        tabelas_componentes[portfolio] = {
            'atual': tab_atual,
            'ultimo': tab_ultimo,
            'comp_anterior_mtd_atual': tab_comp_anterior_mtd_atual,
        }

        nome = mapa_arquivo[portfolio]

        variantes = [
            ('atual', tab_atual),
            ('ultimo_rebal', tab_ultimo),
            ('comp_mes_passado_mtd_atual', tab_comp_anterior_mtd_atual),
        ]
        for sufixo, tab in variantes:
            caminho = f"{output_dir}\\componentes_{nome}_{sufixo}.xlsx"

            with pd.ExcelWriter(caminho, engine='xlsxwriter') as writer:
                tab.to_excel(writer, index=False, sheet_name='componentes')
                wb = writer.book
                ws = writer.sheets['componentes']

                fmt_pct = wb.add_format({'num_format': '0.0%'})

                for c in cols_pct:
                    col_idx = tab.columns.get_loc(c)
                    ws.set_column(col_idx, col_idx, 14, fmt_pct)

    # --- Composição e exportações Excel ---

    output_dir = str(settings.output_dir)

    comp_sheet = pd.read_excel(
        settings.comp_sheet_path, sheet_name='Sheet1'
    )[['TICKER', 'NAME', 'TARGET', 'RECOMMENDATION']].rename(columns={
        'TICKER': 'Ticker', 'NAME': 'Company', 'TARGET': 'Target Price', 'RECOMMENDATION': 'Rating'
    })

    sector_classification_raw = xpqs[['cod_ativo', 'adjusted_GICS_sector']].rename(
        columns={'cod_ativo': 'Ticker', 'adjusted_GICS_sector': 'Sector'}
    ).drop_duplicates('Ticker')

    sector_weight_ibovespa, segment_weight_ibovespa = calcular_pesos_ibovespa(
        settings.index_composition_path, xpqs
    )

    segment_sector = pd.DataFrame(
        [{'Sector': s, 'Segment': seg} for s, seg in sector_to_segment.items()]
    )

    with pd.ExcelWriter(f"{output_dir}\\sector_weights_ibovespa.xlsx") as writer:
        sector_weight_ibovespa.to_excel(writer, sheet_name='Individual Sectors', index=False)
        segment_weight_ibovespa.to_excel(writer, sheet_name='Sector Groups', index=False)

    dfs_composicao = {}

    for portfolio in portfolio_names:
        nome = nomes_exib[portfolio]
        composition_args = (
            composition_dict[portfolio], comp_sheet, sector_classification_raw,
            sector_weight_ibovespa, segment_sector,
        )
        df_en = montar_df_composicao(*composition_args, idioma='EN')
        df_pt = montar_df_composicao(*composition_args, idioma='PT')
        dfs_composicao[portfolio] = {'EN': df_en, 'PT': df_pt}

        wb = Workbook()
        wb.remove(wb.active)

        # abas VISUAIS (com merge) — para relatório
        ws_en = wb.create_sheet(title='ENG')
        _preencher_sheet(ws_en, df_en, f"{nome} Portfolio", idioma='EN')

        ws_pt = wb.create_sheet(title='PT')
        _preencher_sheet(ws_pt, df_pt, f"Carteira {nome}", idioma='PT')

        # abas FLAT (sem merge) — para query / Power Query
        ws_en_q = wb.create_sheet(title='ENG_data')
        _preencher_sheet_flat(ws_en_q, df_en, idioma='EN')

        ws_pt_q = wb.create_sheet(title='PT_data')
        _preencher_sheet_flat(ws_pt_q, df_pt, idioma='PT')

        wb.save(f"{output_dir}\\{arq_comp[portfolio]}.xlsx")
        print(f"Composição gerada (visual + data): {nome}")

    ppt_config = commercial_ppt_config(settings)

    # --- Atualização da lâmina em PowerPoint ---

    for portfolio, cfg in ppt_config.items():
        template = cfg.get('template', '')
        if not template or not os.path.exists(template):
            print(f"[PULADO] template não encontrado para {portfolio}: {template}")
            continue
        atualizar_ppt(
            caminho_template=template,
            caminho_saida=cfg['saida'],
            df_port=_df_para_lamina(portfolio, resultado_dfs),
            composition=composition_dict[portfolio],
            serie_cdi=cdi,
            portfolio=portfolio,
        )

    # --- Atribuição de retorno ---

    for portfolio in portfolio_names:
        dec_atual, dec_ant = gerar_decomposicoes(
            composition_dict[portfolio], md, mapa_nome, mapa_setor
        )
        caminho = f"{output_dir}\\{arq_decomp[portfolio]}.xlsx"

        data_max = md['data'].max()
        lbl_atual = f"{MES_LBL_PT[data_max.month]}-{str(data_max.year)[2:]}"
        ref_ant = pd.Timestamp(data_max.year, data_max.month, 1) - pd.Timedelta(days=1)
        lbl_ant = f"{MES_LBL_PT[ref_ant.month]}-{str(ref_ant.year)[2:]}"

        with pd.ExcelWriter(caminho, engine='xlsxwriter') as writer:
            wb = writer.book
            fmt_pct = wb.add_format({'num_format': '0.0%'})

            for sheet, tab in [(f"mes_atual_{lbl_atual}", dec_atual),
                               (f"mes_anterior_{lbl_ant}", dec_ant)]:
                tab.to_excel(writer, index=False, sheet_name=sheet)
                ws = writer.sheets[sheet]
                for c in cols_pct_decomp:
                    col_idx = tab.columns.get_loc(c)
                    ws.set_column(col_idx, col_idx, 14, fmt_pct)
                ws.set_column(0, 0, 24)  # Companhia
                ws.set_column(2, 2, 22)  # Setor

        print(f"Decomposição de retorno exportada: {caminho}")

