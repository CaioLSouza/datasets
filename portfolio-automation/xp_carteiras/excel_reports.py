"""Excel Reports do pipeline de carteiras XP."""

from __future__ import annotations

import datetime
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

def _col_para_data(col):
    """
    Se a coluna for um rótulo de mês no formato '%b-%y' (ex: 'Jun-25'),
    retorna um datetime (1º dia do mês: 01/Jun/25). Senão, retorna None.
    """
    if not isinstance(col, str) or col == '':
        return None
    try:
        return datetime.datetime.strptime(col, '%b-%y')
    except ValueError:
        return None


def exportar_tabela_retornos(tab_raw, caminho, sheet='performance'):
    """
    Exporta a tabela de retornos:
      - Header dos meses gravado como DATA real (1º dia do mês), formato 'mmm/yy'.
      - Colunas de texto (Since inception, YTD, LTM, anos, separador '') ficam texto.
      - Valores em % (número de verdade, formato 0.0%).
    """
    tab = tab_raw.copy()

    with pd.ExcelWriter(caminho, engine='xlsxwriter') as writer:
        tab.to_excel(writer, sheet_name=sheet, startrow=1, header=False)
        wb = writer.book
        ws = writer.sheets[sheet]

        fmt_pct  = wb.add_format({'num_format': '0.0%', 'align': 'center'})
        fmt_date = wb.add_format({'num_format': 'mmm/yy', 'align': 'center', 'bold': True})
        fmt_txt  = wb.add_format({'align': 'center', 'bold': True})

        ws.write(0, 0, '')

        for j, col in enumerate(tab.columns, start=1):
            dt = _col_para_data(col)
            if dt is not None:
                ws.write_datetime(0, j, dt, fmt_date)
            else:
                ws.write(0, j, str(col), fmt_txt)

        for j, col in enumerate(tab.columns, start=1):
            if col == '':
                continue
            ws.set_column(j, j, 9, fmt_pct)

        ws.set_column(0, 0, 16)

    print(f"Performance exportada: {caminho}")


def _merge_groups(ws, col, first, last, body_font):
    start = first
    cur = ws.cell(row=start, column=col).value
    for row in range(first + 1, last + 2):
        val = ws.cell(row=row, column=col).value if row <= last else None
        if val != cur:
            if row - 1 > start:
                ws.merge_cells(start_row=start, start_column=col, end_row=row - 1, end_column=col)
                c = ws.cell(row=start, column=col)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.font = body_font
            start = row
            if row <= last:
                cur = val


def _preencher_sheet(ws, df, titulo, idioma='EN'):
    header_font = Font(name='Roboto Light', size=7, bold=True, color='000000')
    body_font   = Font(name='Roboto Light', size=7.5, bold=False, color='000000')
    title_font  = Font(name='Roboto Light', size=7.5, bold=True, color='000000')

    if idioma == 'EN':
        header_color, border_color = '8EB3DF', '8EB3DF'
    else:
        header_color, border_color = 'FFBC00', 'D9D9D9'

    ws['A1'] = titulo
    ws['A1'].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    start_row = 3
    header_fill = PatternFill('solid', fgColor=header_color)
    for ci, col in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=ci, value=col)
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = header_fill

    for ri, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for cj, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=cj, value=val)
            c.font = body_font
            c.alignment = Alignment(horizontal='center', vertical='center')

    last_row = start_row + len(df)
    for r in range(start_row + 1, last_row + 1):
        ws.cell(row=r, column=3).number_format = '0.0%'
        ws.cell(row=r, column=4).number_format = '0.0%'
        ws.cell(row=r, column=7).number_format = '0.0%'
        ws.cell(row=r, column=9).number_format = '"R$" #,##0.00'

    thin = Side(border_style='thin', color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(start_row, last_row + 1):
        for c in range(1, len(df.columns) + 1):
            ws.cell(row=r, column=c).border = border

    for col in [1, 2, 3, 4]:
        _merge_groups(ws, col, start_row + 1, last_row, body_font)

    widths = {1: 15, 2: 18, 3: 22, 4: 24, 5: 25, 6: 10, 7: 10, 8: 10, 9: 15}
    for ci, w in widths.items():
        ws.column_dimensions[chr(ord('A') + ci - 1)].width = w


def _preencher_sheet_flat(ws, df, idioma='EN'):
    """
    Versão SEM merge: repete Segment/Sector/Pesos em todas as linhas.
    Ideal para query (Power Query, Ctrl+T, fórmulas). Header na linha 1.
    """
    header_font = Font(name='Roboto Light', size=8, bold=True, color='000000')
    body_font   = Font(name='Roboto Light', size=8, bold=False, color='000000')

    header_color = '8EB3DF' if idioma == 'EN' else 'FFBC00'
    header_fill = PatternFill('solid', fgColor=header_color)

    for ci, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = header_fill

    for ri, row in enumerate(df.itertuples(index=False), start=2):
        for cj, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=cj, value=val)
            c.font = body_font
            c.alignment = Alignment(horizontal='center', vertical='center')

    last_row = 1 + len(df)
    for r in range(2, last_row + 1):
        ws.cell(row=r, column=3).number_format = '0.0%'
        ws.cell(row=r, column=4).number_format = '0.0%'
        ws.cell(row=r, column=7).number_format = '0.0%'
        ws.cell(row=r, column=9).number_format = '"R$" #,##0.00'

    widths = {1: 15, 2: 18, 3: 22, 4: 24, 5: 25, 6: 10, 7: 10, 8: 10, 9: 15}
    for ci, w in widths.items():
        ws.column_dimensions[chr(ord('A') + ci - 1)].width = w


def _eh_nan(x):
    return x is None or (isinstance(x, float) and np.isnan(x))


def _fmt_pct_br(x, dec=1, dash='-'):
    """0.123 -> '12,3%' ; negativo -> '-6,7%' ; NaN -> '-'."""
    if _eh_nan(x):
        return dash
    return f"{x * 100:.{dec}f}%".replace('.', ',')


def _fmt_num_br(x, dec=2, dash='-'):
    """0.46 -> '0,46' ; NaN -> '-'."""
    if _eh_nan(x):
        return dash
    return f"{x:.{dec}f}".replace('.', ',')


def _fmt_int_br(x, dash='-'):
    if _eh_nan(x):
        return dash
    return str(int(round(x)))


def _fmt_dur(n, dash='-'):
    if _eh_nan(n):
        return dash
    n = int(round(n))
    return f"{n} mês" if n == 1 else f"{n} meses"


