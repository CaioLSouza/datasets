"""Stub out ONLY the Bloomberg calls, keep every Excel formula, then check the maths in real Excel.
Verifies the aggregation, sector discovery, momentum-from-TR-index and missing-data guards.
It cannot verify BDP/BDH/BDS themselves.
"""
import openpyxl, os, pythoncom, win32com.client as w32, pandas as pd, numpy as np

SRC = 'Ibovespa Factor Screen (BBG).xlsx'
TST = 'zz_test_stubbed.xlsx'
DR0 = 5
# columns that hold a Bloomberg call and therefore get stubbed
BBG_COLS = (4, 5, 6, 9, 10, 11, 14, 15, 16)

#            ticker   sector      shout   price  eps_now eps_A eps_B  trNow   trA     trB
T = [('PETR4 BZ','Energy',      13044.0, 38.50,  10.20,  9.80,  9.00, 112.50, 100.00,  92.21),
     ('PETR3 BZ','Energy',       7442.0, 41.20,  10.20,  9.80,  9.00, 111.80, 100.00,  92.40),
     ('PRIO3 BZ','Energy',        865.0, 44.10,   6.10,  6.40,  7.00,  91.80, 100.00, 108.64),
     ('ITUB4 BZ','Financials',   5286.0, 36.90,   4.10,  3.95,  3.80, 106.40, 100.00,  93.17),
     ('BBDC4 BZ','Financials',   5238.0, 17.30,   2.05,  2.10,  2.20,  97.90, 100.00,  94.77),
     ('BBAS3 BZ','Financials',   5730.0, 28.40,   7.80,  7.20,  6.90, 115.00, 100.00,  90.27),
     ('MGLU3 BZ','Retail',      66000.0,  9.15,   0.12,  0.09,  0.05, 133.00, 100.00,  88.08),
     ('LREN3 BZ','Retail',       1010.0, 18.60,   1.40,  1.55,  1.62,  90.50, 100.00, 102.84),
     ('VALE3 BZ','Materials',    4300.0, 61.70,   8.90,  8.40,  8.10, 107.70, 100.00,  98.00),
     ('ABEV3 BZ','Consumer',    15700.0, 13.05,   0.95,  0.95,  0.92, 101.20, 100.00,  96.84),
     # edge cases ---------------------------------------------------------------------------
     ('WEGE3 BZ','Materials',    4200.0, 52.30,   1.80,  0.00,  1.55, 104.00, 100.00,  None),
     ('XPTO3 BZ', None,          1000.0, 10.00,   1.00,  0.90,  0.80, 105.00, 100.00,  95.45)]
#  WEGE3: eps_A = 0 -> rev A blank, rev B still computes; trB missing -> out of momentum B only
#  XPTO3: no sector -> excluded from every sector row AND from the index total

wb = openpyxl.load_workbook(SRC)
mem, mp, d = wb['Index Members'], wb['Sector Map'], wb['Data']
mem['A4'] = None
for i, row in enumerate(T):
    mem.cell(4 + i, 1, row[0])
    if row[1] is not None:
        mp.cell(4 + i, 2, row[1])
    r = DR0 + i
    d.cell(r, 4, row[2] * row[3])     # D company mkt cap
    d.cell(r, 5, row[2])              # E shares out
    d.cell(r, 6, row[3])              # F price
    d.cell(r, 9, row[4])              # I fwd EPS now
    d.cell(r,10, row[5])              # J fwd EPS at A
    d.cell(r,11, row[6])              # K fwd EPS at B
    d.cell(r,14, row[7])              # N TR index now
    d.cell(r,15, row[8])              # O TR index at A
    d.cell(r,16, row[9])              # P TR index at B
for i in range(len(T), 120):
    for c in BBG_COLS: d.cell(DR0 + i, c, None)
wb.save(TST)

# ---------------- expected values, computed independently -------------------------
df = pd.DataFrame(T, columns=['tk','sec','sh','px','e0','eA','eB','trN','trA','trB'])
df['w'] = df.sh * df.px
df['revA'] = np.where(df.eA > 0, df.e0/df.eA - 1, np.nan)
df['revB'] = np.where(df.eB > 0, df.e0/df.eB - 1, np.nan)
df['momA'] = np.where(df.trA > 0, df.trN/df.trA - 1, np.nan)
trB = df.trB.astype(float)
df['momB'] = np.where(trB > 0, df.trN/trB - 1, np.nan)
ok = df.sec.notna()
def wavg(sub, col):
    m = sub[col].notna()
    return np.nan if sub.loc[m,'w'].sum() == 0 else (sub.loc[m,col]*sub.loc[m,'w']).sum()/sub.loc[m,'w'].sum()
exp_sec = {s: {c: wavg(g, c) for c in ('revA','revB','momA','momB')} | {'n': len(g), 'w': g.w.sum()}
           for s, g in df[ok].groupby('sec')}
exp_tot = {c: wavg(df[ok], c) for c in ('revA','revB','momA','momB')}
exp_tot['w'] = df[ok].w.sum(); exp_tot['n'] = int(ok.sum())
exp_n = {c: int(df[ok][c].notna().sum()) for c in ('revA','revB','momA','momB')}

# ---------------- read Excel back ------------------------------------------------
pythoncom.CoInitialize()
xl = w32.gencache.EnsureDispatch('Excel.Application'); xl.Visible=False; xl.DisplayAlerts=False
try:
    bk = xl.Workbooks.Open(os.path.abspath(TST)); xl.Application.CalculateFullRebuild()
    ag, da, dg = (bk.Worksheets('Sector Aggregation'), bk.Worksheets('Data'),
                  bk.Worksheets('Diagnostics'))
    nerr = 0
    for sh in bk.Worksheets:
        if sh.Name == 'Diagnostics':          # raw BBG calls by design; errors expected here
            continue
        try: rg = sh.UsedRange.SpecialCells(-4123, 16)
        except Exception: continue
        nerr += rg.Count
        try:
            locs = [c.Address(False,False)+'='+str(c.Text) for ar in rg.Areas for c in ar][:8]
        except Exception as e: locs = [f'({e})']
        print(f'  ERRORS on "{sh.Name}": {rg.Count} -> {locs}')
    print(f'error cells outside Diagnostics: {nerr}')

    print('\n--- per-ticker guards ---')
    for i in (10, 11):
        r = DR0 + i
        print(f'  {da.Cells(r,1).Value:9s} sector={str(da.Cells(r,2).Value):14s} '
              f'revA={da.Cells(r,12).Text:>8s} revB={da.Cells(r,13).Text:>8s} '
              f'momA={da.Cells(r,17).Text:>8s} momB={da.Cells(r,18).Text:>8s} '
              f'flag={da.Cells(r,19).Value}')

    print('\n--- sector rows (discovered, order of first appearance) ---')
    worst = 0.0
    for r in range(4, 34):
        s = ag.Cells(r,1).Value
        if not s: continue
        e = exp_sec[s]
        line = f'  {s:12s} n={ag.Cells(r,2).Value:2.0f} w={ag.Cells(r,3).Value:12,.0f}  '
        for k, c in enumerate(('revA','revB','momA','momB')):
            gv, ev = ag.Cells(r, 5+k).Value, e[c]
            if gv in ('', None) or (isinstance(ev, float) and np.isnan(ev)):
                line += f'{"blank":>10s} '
            else:
                worst = max(worst, abs(gv-ev)); line += f'{gv*100:9.4f}% '
        assert abs(ag.Cells(r,3).Value - e['w']) < 1e-6, f'weight mismatch {s}'
        assert ag.Cells(r,2).Value == e['n'], f'count mismatch {s}'
        print(line)

    rt = 35
    print(f'\n--- index total ---')
    print(f'  {ag.Cells(rt,1).Value}: n={ag.Cells(rt,2).Value:.0f} (exp {exp_tot["n"]})  '
          f'w={ag.Cells(rt,3).Value:,.0f} (exp {exp_tot["w"]:,.0f})')
    for k, c in enumerate(('revA','revB','momA','momB')):
        gv = ag.Cells(rt, 5+k).Value
        worst = max(worst, abs(gv-exp_tot[c]))
        gn, en = ag.Cells(rt, 9+k).Value, exp_n[c]
        flag = 'ok' if gn == en else f'MISMATCH exp {en}'
        print(f'  {c}: excel={gv*100:9.5f}%  python={exp_tot[c]*100:9.5f}%   n={gn:.0f} {flag}')
    print(f'  coverage: {ag.Cells(37,2).Value}')

    print('\n--- diagnostics sheet cross-references ---')
    for r in range(4, 22):
        lab, txt = dg.Cells(r,2).Value, str(dg.Cells(r,3).Text)
        if lab and str(lab)[0].isdigit():
            print(f'  {str(lab):30s} -> {txt[:34]}')
    print(f'  test 7 formula: {dg.Cells(14,3).Formula[:76]}')
    print(f'  verdict cell  : {dg.Cells(21,3).Formula[:60]}')

    print(f'\nMAX |excel - python| across every weighted average: {worst:.3e}')
    bk.Close(SaveChanges=False)
finally:
    xl.Quit(); pythoncom.CoUninitialize()
