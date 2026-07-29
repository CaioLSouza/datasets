"""Ibovespa factor screen driven by Bloomberg formulas, aggregated by a user-supplied sector."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

NROW = 120                      # ticker capacity (Ibovespa carries ~87)
R0, R1 = 4, 3 + NROW            # first / last data row
NSEC = 30                       # sector slots
S0, S1 = 4, 3 + NSEC

SH_READ, SH_IN, SH_MAP, SH_MEM = 'Read me', 'Inputs', 'Sector Map', 'Index Members'
SH_DATA, SH_AGG = 'Data', 'Sector Aggregation'

A = 'Arial'
F_T  = Font(name=A, size=13, bold=True, color='1A2226')
F_H  = Font(name=A, size=9,  bold=True, color='FFFFFF')
F_SB = Font(name=A, size=9,  bold=True, color='1A2226')
F_B  = Font(name=A, size=10)
F_IN = Font(name=A, size=11, bold=True, color='0000FF')
F_C  = Font(name=A, size=10, color='000000')
F_K  = Font(name=A, size=10, bold=True, color='000000')
F_N  = Font(name=A, size=9,  color='595959')
F_BBG= Font(name=A, size=10, color='7030A0')      # cells that hit Bloomberg
FH   = PatternFill('solid', fgColor='1F3864')
FSB  = PatternFill('solid', fgColor='D9E2F3')
FIN  = PatternFill('solid', fgColor='FFFF00')
FKEY = PatternFill('solid', fgColor='EAF2F2')
FHLP = PatternFill('solid', fgColor='F2F2F2')
BOX  = Border(*[Side(style='thin', color='BFBFBF')]*4)
PCT, PCT1, NUM0, NUM2, DT = '0.0%', '0.00%', '#,##0', '#,##0.00', 'yyyy-mm-dd'
WRAP = Alignment(wrap_text=True, vertical='top')

wb = Workbook()

def head(ws, row, labels, widths, groups=None):
    if groups:
        for lab, c0, c1 in groups:
            cell = ws.cell(row-1, c0, lab); cell.font, cell.fill = F_SB, FSB
            cell.alignment = Alignment(horizontal='center')
            if c1 > c0: ws.merge_cells(start_row=row-1, start_column=c0, end_row=row-1, end_column=c1)
    for j, s in enumerate(labels, 1):
        c = ws.cell(row, j, s); c.font, c.fill = F_H, FH
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 40
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[L(j)].width = w

def put(ws, r, c, v, nf=None, font=F_C, fill=None):
    cell = ws.cell(r, c, v)
    if nf: cell.number_format = nf
    cell.font = font
    if fill: cell.fill = fill
    return cell

# ==================================================================== Inputs
ws = wb.create_sheet(SH_IN)
ws.cell(1,1,'Inputs — the whole workbook is driven from here').font = F_T
ws.cell(2,1,'Edit only the YELLOW cells. Field mnemonics live here on purpose: if your terminal '
            'names a field differently, change it once here and every ticker follows.').font = F_N
for col, w in [('A',3),('B',30),('C',30),('D',3),('E',74)]:
    ws.column_dimensions[col].width = w

PARAMS = [
    ('SCOPE', None, None, None, None),
    ('Index ticker', 'IBOV Index', None, 'IdxTicker',
     'Feeds BDS(...,"INDX_MEMBERS") on the Index Members sheet. Swap for "IBX Index", "SMLL Index", etc.'),
    ('Ticker suffix', ' Equity', None, 'TkSuffix',
     'INDX_MEMBERS returns "PETR4 BZ", so " Equity" is appended to make a full BDP ticker. '
     'Keep the leading space.'),
    ('As-of date', '=TODAY()', DT, 'AsOf',
     'Anchor for every lookback. Leave as TODAY() for a live screen, or hardcode a date to freeze a snapshot.'),
    ('WINDOWS', None, None, None, None),
    ('Momentum window A (months)', 3, NUM0, 'MomA', 'Short momentum horizon.'),
    ('Momentum window B (months)', 6, NUM0, 'MomB', 'Long momentum horizon.'),
    ('Revision window A (months)', 3, NUM0, 'RevA', 'Short earnings-revision horizon.'),
    ('Revision window B (months)', 6, NUM0, 'RevB', 'Long earnings-revision horizon.'),
    ('FIELDS', None, None, None, None),
    ('Market cap', 'CUR_MKT_CAP', None, 'FldMcap',
     'Bloomberg reports this at COMPANY level. For a Brazilian name with ON and PN lines both in the '
     'index it is the same number on both, which double-counts if you weight by it. See Weighting basis.'),
    ('Shares outstanding', 'EQY_SH_OUT', None, 'FldShOut',
     'Shares of THIS listed line, in millions. Used to build a security-level cap.'),
    ('Last price', 'PX_LAST', None, 'FldPx', 'Used with shares outstanding for the security-level cap.'),
    ('Forward EPS estimate', 'BEST_EPS', None, 'FldEps',
     'Consensus EPS. The revision is the change in this number, so the fiscal period must be pinned — '
     'see the override below.'),
    ('Fiscal period override', 'BF', None, 'FpOvr',
     'BEST_FPERIOD_OVERRIDE. "BF" is blended forward 12m, which rolls continuously and is the right '
     'choice for a revision. A fixed year ("2027") also works. Leaving it unpinned would compare '
     'different fiscal years and produce a meaningless revision at every annual roll.'),
    ('Total return field', 'CUST_TRR_RETURN_HOLDING_PER', None, 'FldTRR',
     'Total return over a custom window, driven by the two date overrides. Includes dividends, which is '
     'the correct momentum measure. If your terminal rejects it, try CHG_PCT_3M / CHG_PCT_6M instead '
     '(price only, fixed windows).'),
    ('WEIGHTING', None, None, None, None),
    ('Weighting basis', 'SEC_CAP', None, 'WgtBasis',
     'SEC_CAP = shares outstanding x price, i.e. the market cap of THIS listed line. '
     'COMPANY_CAP = the CUR_MKT_CAP field. Use SEC_CAP for Brazil: it is what you asked for, and it '
     'stops PETR3+PETR4 or ITUB3+ITUB4 from counting the same company twice.'),
]
r = 4
NAMED = {}
for lab, val, nf, nm, note in PARAMS:
    if val is None:
        ws.cell(r, 2, lab).font = F_SB
        ws.cell(r, 2).fill = FSB; ws.cell(r, 3).fill = FSB
        r += 1; continue
    put(ws, r, 2, lab, None, F_B)
    c = put(ws, r, 3, val, nf, F_IN, FIN); c.border = BOX
    n = put(ws, r, 5, note, None, F_N); n.alignment = WRAP
    ws.row_dimensions[r].height = 30 if len(str(note)) < 110 else 44
    wb.defined_names.add(DefinedName(nm, attr_text=f"'{SH_IN}'!$C${r}"))
    NAMED[nm] = r
    r += 1

dv = DataValidation(type='list', formula1='"SEC_CAP,COMPANY_CAP"', allow_blank=False)
ws.add_data_validation(dv); dv.add(ws.cell(NAMED['WgtBasis'], 3))

# derived lookback dates
r += 1
ws.cell(r, 2, 'DERIVED DATES (calculated)').font = F_SB
ws.cell(r, 2).fill = FSB; ws.cell(r, 3).fill = FSB
r += 1
for lab, nm, src in [('Momentum A start','DtMomA','MomA'), ('Momentum B start','DtMomB','MomB'),
                     ('Revision A base date','DtRevA','RevA'), ('Revision B base date','DtRevB','RevB')]:
    put(ws, r, 2, lab, None, F_B)
    put(ws, r, 3, f'=EDATE(AsOf,-{src})', DT, F_C, FKEY)
    put(ws, r, 5, 'EDATE back from the as-of date. Bloomberg overrides receive it as YYYYMMDD.', None, F_N)
    wb.defined_names.add(DefinedName(nm, attr_text=f"'{SH_IN}'!$C${r}"))
    r += 1
ws.freeze_panes = 'A4'

# ==================================================================== Index Members
ws = wb.create_sheet(SH_MEM)
ws.cell(1,1,'Index members — live from Bloomberg').font = F_T
ws.cell(2,1,'One BDS call. It spills down column A when the add-in resolves it. Everything else in the '
            'workbook reads this list, so nothing needs to be pasted by hand.').font = F_N
ws.column_dimensions['A'].width = 18
ws.column_dimensions['C'].width = 90
put(ws, 3, 1, 'Member', None, F_H, FH)
put(ws, 4, 1, '=BDS(IdxTicker,"INDX_MEMBERS")', None, F_BBG)
put(ws, 4, 3, '↑ This single cell spills the full member list. If you see #NAME? the Bloomberg add-in '
              'is not loaded; if you see one ticker only, your Excel is not spilling — copy the cell '
              'down instead.', None, F_N).alignment = WRAP
put(ws, 6, 3, '=IF(COUNTA($A$4:$A$400)=0,"No members returned — check the add-in and the index ticker.",'
              '"Members returned: "&COUNTA($A$4:$A$400))', None, F_K)

MEMRNG = f"'{SH_MEM}'!$A$4:$A$403"

def member(row_offset):
    """Nth member, or "" past the end of the list.

    INDEX over an empty cell returns 0, not blank, so a bare IFERROR(INDEX(...),"") leaves a
    numeric 0 in every unused row. That 0 is not equal to "", so downstream IF($A="") guards
    fall through and the row gets treated as a real ticker.
    """
    ix = f'INDEX({MEMRNG},ROW()-{row_offset})'
    return f'=IFERROR(IF({ix}=0,"",{ix}),"")'

# ==================================================================== Sector Map
ws = wb.create_sheet(SH_MAP)
ws.cell(1,1,'Sector Map — the only sheet you fill in').font = F_T
ws.cell(2,1,'Column A arrives from the index automatically. Type your sector in column B. '
            'Nothing else in the workbook needs touching.').font = F_N
head(ws, 3, ['Ticker','Sector (fill this in)','Status'], [16, 30, 46])
ws.freeze_panes = 'A4'
for i in range(NROW):
    r = R0 + i
    put(ws, r, 1, member(R0 - 1), None, F_C)
    c = put(ws, r, 2, None, None, F_IN, FIN); c.border = BOX
    put(ws, r, 3, f'=IF($A{r}="","",IF($B{r}="","EMPTY","ok"))', None, F_N)
MAPRNG = f"'{SH_MAP}'!$A${R0}:$B${R1}"
r = R1 + 2
put(ws, r, 1, 'Filled:', None, F_K)
put(ws, r, 2, f'=COUNTIF($C${R0}:$C${R1},"ok")&" of "&COUNTA($A${R0}:$A${R1})&" tickers — "'
              f'&COUNTIF($C${R0}:$C${R1},"EMPTY")&" to go"', None, F_K, FKEY)

# ==================================================================== Data
ws = wb.create_sheet(SH_DATA)
ws.cell(1,1,'Security-level factor screen — every value pulled from Bloomberg').font = F_T
ws.cell(2,1,'Purple cells call Bloomberg. Grey columns from Q onward are aggregation helpers; they are '
            'grouped and hidden, and can be ignored.').font = F_N
COLS = ['Ticker','Sector','Bloomberg ticker','Company mkt cap','Shares out (mm)','Price',
        'Security mkt cap','Weight used','Fwd EPS now','Fwd EPS at A','Fwd EPS at B',
        'Earnings rev. A','Earnings rev. B','Momentum A','Momentum B','Data flag']
HELP = ['mapped','w·revA','w revA','w·revB','w revB','w·momA','w momA','w·momB','w momB','sector #']
head(ws, 4, COLS + HELP, [13,24,19,15,13,10,15,15,12,12,12,13,13,12,12,30] + [11]*10,
     groups=[('IDENTITY',1,3), ('SIZE',4,8), ('EARNINGS REVISIONS',9,13),
             ('MOMENTUM',14,15), ('QUALITY',16,16), ('AGGREGATION HELPERS',17,26)])
ws.freeze_panes = 'C5'
DR0, DR1 = 5, 4 + NROW
for i in range(NROW):
    r = DR0 + i
    blank = f'$C{r}=""'
    put(ws, r, 1, member(DR0 - 1), None, F_C)
    # sentinels are plain ASCII and matched exactly everywhere. A non-ASCII marker survives a
    # LEFT() comparison but silently degrades to a bare wildcard inside COUNTIF criteria.
    put(ws, r, 2, f'=IF($A{r}="","",IFERROR(IF(VLOOKUP($A{r},{MAPRNG},2,FALSE)=0,"NO SECTOR",'
                  f'VLOOKUP($A{r},{MAPRNG},2,FALSE)),"NOT IN MAP"))', None, F_C)
    put(ws, r, 3, f'=IF($A{r}="","",$A{r}&TkSuffix)', None, F_C)
    put(ws, r, 4, f'=IF({blank},"",BDP($C{r},FldMcap))', NUM0, F_BBG)
    put(ws, r, 5, f'=IF({blank},"",BDP($C{r},FldShOut))', NUM0, F_BBG)
    put(ws, r, 6, f'=IF({blank},"",BDP($C{r},FldPx))', NUM2, F_BBG)
    put(ws, r, 7, f'=IFERROR($E{r}*$F{r},"")', NUM0, F_C)
    put(ws, r, 8, f'=IFERROR(IF(WgtBasis="SEC_CAP",$G{r},$D{r}),"")', NUM0, F_K)
    put(ws, r, 9, f'=IF({blank},"",BDP($C{r},FldEps,"BEST_FPERIOD_OVERRIDE="&FpOvr))', NUM2, F_BBG)
    for col, dt in [(10,'DtRevA'), (11,'DtRevB')]:
        put(ws, r, col,
            f'=IF({blank},"",IFERROR(INDEX(BDH($C{r},FldEps,{dt},{dt},'
            f'"BEST_FPERIOD_OVERRIDE="&FpOvr,"Days=A","Fill=P"),1,1),""))', NUM2, F_BBG)
    put(ws, r, 12, f'=IFERROR(IF($J{r}<=0,"",$I{r}/$J{r}-1),"")', PCT1, F_K)
    put(ws, r, 13, f'=IFERROR(IF($K{r}<=0,"",$I{r}/$K{r}-1),"")', PCT1, F_K)
    for col, dt in [(14,'DtMomA'), (15,'DtMomB')]:
        put(ws, r, col,
            f'=IF({blank},"",IFERROR(BDP($C{r},FldTRR,'
            f'"CUST_TRR_START_DT="&TEXT({dt},"yyyymmdd"),'
            f'"CUST_TRR_END_DT="&TEXT(AsOf,"yyyymmdd"))/100,""))', PCT1, F_BBG)
    put(ws, r, 16,
        f'=IF($A{r}="","",IF($Q{r}=0,"sector missing",'
        f'IF(NOT(ISNUMBER($H{r})),"no market cap",'
        f'IF(COUNT($L{r}:$O{r})=4,"complete",'
        f'"missing "&(4-COUNT($L{r}:$O{r}))&" of 4 factors"))))', None, F_N)
    # one central "is this row usable" flag, so every downstream guard shares the same definition
    put(ws, r, 17, f'=IF(OR($B{r}="",$B{r}="NO SECTOR",$B{r}="NOT IN MAP"),0,1)', NUM0, F_N, FHLP)
    # numerator / denominator per metric, zero unless the metric, the weight and the sector are all good
    for k, mcol in enumerate(['L','M','N','O']):
        ok = f'AND($Q{r}=1,ISNUMBER(${mcol}{r}),ISNUMBER($H{r}))'
        put(ws, r, 18+2*k, f'=IF({ok},${mcol}{r}*$H{r},0)', NUM2, F_N, FHLP)
        put(ws, r, 19+2*k, f'=IF({ok},$H{r},0)', NUM0, F_N, FHLP)
    put(ws, r, 26,
        f'=IF($Q{r}=0,"",IF(COUNTIF($B${DR0}:$B{r},$B{r})>1,"",'
        + (f'MAX(0,MAX($Z${DR0}:$Z{r-1}))+1))' if i else '1))'), NUM0, F_N, FHLP)
for c in range(17, 27):
    ws.column_dimensions[L(c)].outline_level = 1
    ws.column_dimensions[L(c)].hidden = True
ws.conditional_formatting.add(f'$P${DR0}:$P${DR1}',
    CellIsRule(operator='equal', formula=['"complete"'], font=Font(name=A, size=9, color='548235')))

# ==================================================================== Sector Aggregation
ws = wb.create_sheet(SH_AGG)
ws.cell(1,1,'Sector aggregation — market-cap weighted').font = F_T
ws.cell(2,1,'Sectors are discovered from what you typed on Sector Map; no list to maintain here. '
            'A security is only weighted into a factor where BOTH that factor and its market cap came '
            'back as numbers, so a single missing datapoint cannot poison a sector.').font = F_N
head(ws, 3, ['Sector','Names','Total mkt cap','Share of index','Earnings rev. A','Earnings rev. B',
             'Momentum A','Momentum B','n (rev A)','n (rev B)','n (mom A)','n (mom B)'],
     [26,8,17,13,14,14,13,13,10,10,10,10])
ws.freeze_panes = 'B4'
SB, SW = f"'{SH_DATA}'!$B${DR0}:$B${DR1}", f"'{SH_DATA}'!$H${DR0}:$H${DR1}"
HN = {k: f"'{SH_DATA}'!${L(18+2*k)}${DR0}:${L(18+2*k)}${DR1}" for k in range(4)}
HD = {k: f"'{SH_DATA}'!${L(19+2*k)}${DR0}:${L(19+2*k)}${DR1}" for k in range(4)}
for i in range(NSEC):
    r = S0 + i
    put(ws, r, 1, f"=IFERROR(INDEX('{SH_DATA}'!$B${DR0}:$B${DR1},"
                  f"MATCH({i+1},'{SH_DATA}'!$Z${DR0}:$Z${DR1},0)),\"\")", None, F_K)
    g = f'$A{r}=""'
    put(ws, r, 2, f'=IF({g},"",COUNTIF({SB},$A{r}))', NUM0)
    put(ws, r, 3, f'=IF({g},"",SUMIF({SB},$A{r},{SW}))', NUM0)
    put(ws, r, 4, f'=IF(OR({g},$C${S1+2}=0),"",$C{r}/$C${S1+2})', PCT)
    for k in range(4):
        put(ws, r, 5+k,
            f'=IF({g},"",IF(SUMIF({SB},$A{r},{HD[k]})=0,"",'
            f'SUMIF({SB},$A{r},{HN[k]})/SUMIF({SB},$A{r},{HD[k]})))', PCT1, F_K)
        put(ws, r, 9+k, f'=IF({g},"",COUNTIF({SB},$A{r})-COUNTIFS({SB},$A{r},{HD[k]},0))', NUM0, F_N)

rt = S1 + 2
put(ws, rt, 1, 'INDEX TOTAL', None, F_K, FKEY)
put(ws, rt, 2, f'=SUM($B${S0}:$B${S1})', NUM0, F_K, FKEY)
put(ws, rt, 3, f'=SUM($C${S0}:$C${S1})', NUM0, F_K, FKEY)
put(ws, rt, 4, f'=IF($C{rt}=0,"",1)', PCT, F_K, FKEY)
for k in range(4):
    put(ws, rt, 5+k, f'=IF(SUM({HD[k]})=0,"",SUM({HN[k]})/SUM({HD[k]}))', PCT1, F_K, FKEY)
    put(ws, rt, 9+k, f'=SUM({L(9+k)}{S0}:{L(9+k)}{S1})', NUM0, F_N, FKEY)

rc = rt + 2
put(ws, rc, 1, 'Coverage check', None, F_K)
UNM = (f'(COUNTIF(\'{SH_DATA}\'!$B${DR0}:$B${DR1},"NO SECTOR")'
       f'+COUNTIF(\'{SH_DATA}\'!$B${DR0}:$B${DR1},"NOT IN MAP"))')
put(ws, rc, 2, f'=IF({UNM}>0,{UNM}&" ticker(s) still have no sector — they are excluded from every '
               f'sector row AND from the index total above.",'
               f'"All "&COUNTIF(\'{SH_DATA}\'!$Q${DR0}:$Q${DR1},1)&" tickers are mapped to a sector.")',
    None, F_K, FKEY)
ws.merge_cells(start_row=rc, start_column=2, end_row=rc, end_column=8)

# ==================================================================== Read me
ws = wb.create_sheet(SH_READ, 0)
for col, w in [('A',3),('B',26),('C',104)]:
    ws.column_dimensions[col].width = w
ws.cell(1,1,'Ibovespa factor screen — revisions, momentum and size, aggregated by sector').font = F_T
ws.cell(2,1,'Open with the Bloomberg add-in running. Fill one column on Sector Map; '
            'everything else calculates.').font = F_N
def sect(r, h, items, ht=44):
    ws.cell(r,2,h).font = F_K; r += 1
    for a,b in items:
        ws.cell(r,2,a).font = F_B
        c = ws.cell(r,3,b); c.font = F_B; c.alignment = WRAP
        ws.row_dimensions[r].height = ht; r += 1
    return r+1

r = sect(4, 'WHAT YOU DO', [
    ('1. Open in Bloomberg', 'The add-in resolves one BDS call for the member list and roughly 700 BDP/BDH '
                             'calls for the data. Give it a moment and press F9 if anything still reads #N/A Requesting.'),
    ('2. Fill column B on Sector Map', 'Ticker arrives on its own. Type your sector next to it. '
                                       'The Status column tracks what is still empty.'),
    ('3. Read Sector Aggregation', 'It discovers your sector names automatically — there is no list to '
                                   'maintain. Add a name, and a new row appears.'),
], ht=44)
r = sect(r, 'ONE THING WORTH DOING ONCE', [
    ('Freeze the ticker column',
     'Column A of Sector Map is a formula reading the live index. That is convenient now, but the '
     'Ibovespa rebalances three times a year, and when a name enters or leaves, the formula shifts every '
     'ticker below it while your typed sectors stay put — silently misaligning them. After you have '
     'filled the sectors, select column A, copy, and Paste Special > Values. The list becomes static, '
     'your mapping is anchored to it permanently, and any future entrant simply shows "◄ not in map" '
     'on Data so you can add one row. Five seconds, and the file stops being able to lie to you.'),
], ht=100)
r = sect(r, 'THE FIELDS, AND WHY', [
    ('Earnings revisions', 'Forward EPS today divided by forward EPS as of the lookback date, minus one. '
                           'The fiscal period is pinned with BEST_FPERIOD_OVERRIDE (default "BF", blended '
                           'forward 12m). Without pinning it, every annual roll would compare two '
                           'different fiscal years and the "revision" would be an artefact.'),
    ('Momentum', 'Total return over the window via CUST_TRR_RETURN_HOLDING_PER with explicit start and '
                 'end date overrides. Total return rather than price return, so a large dividend is not '
                 'read as a fall. Divided by 100 because the field returns percentage points.'),
    ('Market cap', 'Two are pulled. CUR_MKT_CAP is company-level, so for a name with both an ON and a PN '
                   'line in the index it reports the same number twice. Shares outstanding x price gives '
                   'the cap of the individual listed line. Weighting basis defaults to the latter — this '
                   'is what you asked for, and it keeps PETR3+PETR4 from counting Petrobras twice.'),
    ('Missing data', 'A security enters a weighted average only where both that factor and its market cap '
                     'returned numbers, and only once it has a sector. The n columns on Sector '
                     'Aggregation show how many names actually stood behind each figure, and the Data '
                     'flag column says what is missing per ticker.'),
])
r = sect(r, 'WHAT I COULD NOT VERIFY', [
    ('No terminal here',
     'The IRR workbook I built earlier was checked cell by cell against Excel. This one could not be: '
     'there is no Bloomberg here, so no formula in it has ever returned a value. The structure, the '
     'Excel-side logic and the aggregation are sound, but treat the first open as a test run.'),
    ('Field names are the likely failure',
     'BDP, BDH and BDS syntax is stable, and CUR_MKT_CAP, EQY_SH_OUT, PX_LAST and BEST_EPS are '
     'long-standing mnemonics. CUST_TRR_RETURN_HOLDING_PER is the one I would check first — it is '
     'entitlement-sensitive and the override names vary. If it fails, put CHG_PCT_3M in the Total return '
     'field cell, clear the date overrides, and accept price-only momentum over a fixed window.'),
    ('Every mnemonic is a cell, not a formula',
     'This is why the Inputs sheet holds the field names. A wrong mnemonic is a one-cell fix that '
     'propagates to all 87 tickers, not a find-and-replace across the file.'),
    ('Scale conventions', 'Check two things on the first pull: that EQY_SH_OUT is in millions on your '
                          'terminal, and that the total return field returns 60 rather than 0.60 for 60%. '
                          'Both are divided or multiplied accordingly and both differ between setups.'),
])
r = sect(r, 'COLOUR KEY', [
    ('Yellow', 'You edit these.'),
    ('Purple', 'Calls Bloomberg.'),
    ('Black', 'Calculated in Excel.'),
    ('Grey, hidden', 'Aggregation helpers on Data, columns Q to Y. Ungroup to inspect.'),
], ht=18)

del wb['Sheet']
wb._sheets = [wb[n] for n in [SH_READ, SH_IN, SH_MAP, SH_DATA, SH_AGG, SH_MEM]]
OUT = 'Ibovespa Factor Screen (BBG).xlsx'
wb.save(OUT)
print('saved', OUT)
print('sheets:', wb.sheetnames)
print(f'ticker rows {DR0}-{DR1} ({NROW}) | sector rows {S0}-{S1} ({NSEC})')
