"""Formula-driven version: an Inputs sheet drives every downstream calculation."""
import pandas as pd, numpy as np, pickle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import LineChart, Reference

with open('raw_panels.pkl', 'rb') as f:
    P = pickle.load(f)

NAMES = {
 'RAIL3':'Rumo','HBSA3':'Hidrovias do Brasil','STBP3':'Santos Brasil','CCRO3':'CCR','ECOR3':'EcoRodovias',
 'EQTL3':'Equatorial Energia','ENGI11':'Energisa','CMIG4':'Cemig','CPLE3':'Copel','AXIA3':'Eletrobras (Axia)',
 'ALUP11':'Alupar','ISAE4':'ISA Energia (CTEEP)','TAEE11':'Taesa','AURE3':'Auren Energia','EGIE3':'Engie Brasil',
 'CPFE3':'CPFL Energia','LIGT3':'Light','ENEV3':'Eneva','CSMG3':'Copasa','SBSP3':'Sabesp','SAPR11':'Sanepar',
 'ORVR3':'Orizon','MULT3':'Multiplan','IGTI11':'Iguatemi','ALOS3':'Allos'}
SECTOR = {**{t:'Infra & logistics' for t in ['RAIL3','HBSA3','STBP3','CCRO3','ECOR3']},
          **{t:'Electric utilities' for t in ['EQTL3','ENGI11','CMIG4','CPLE3','AXIA3','ALUP11','ISAE4',
                                             'TAEE11','AURE3','EGIE3','CPFE3','LIGT3','ENEV3']},
          **{t:'Water & sanitation' for t in ['CSMG3','SBSP3','SAPR11','ORVR3']},
          **{t:'Malls' for t in ['MULT3','IGTI11','ALOS3']}}

INFRA_T = ['RAIL3','HBSA3','STBP3','CCRO3','ECOR3']
UTIL_T  = ['EQTL3','ENGI11','CMIG4','CPLE3','AXIA3','ALUP11','ISAE4','TAEE11','AURE3','EGIE3',
           'CSMG3','SBSP3','CPFE3','SAPR11','LIGT3','ENEV3','ORVR3']
MALL_T  = ['MULT3','IGTI11','ALOS3']
ORDER = INFRA_T + UTIL_T + MALL_T
SRC = {**{t:'infra' for t in INFRA_T}, **{t:'util' for t in UTIL_T}, **{t:'malls' for t in MALL_T}}

SH = dict(inp='Inputs', read='Read me', idx='IRR Index (Monthly)', sm='Company Summary',
          pan='Panel (Monthly)', dl='Chain-link Deltas',
          infra='Infra (Daily)', util='Utilities (Daily)', mall='Malls (Weekly)')
RAWSH = {'infra': SH['infra'], 'util': SH['util'], 'malls': SH['mall']}
RAW_TICKERS = {'infra': INFRA_T, 'util': UTIL_T, 'malls': MALL_T}
RAW_LAST = {k: 3 + len(P[k]) for k in P}                       # last data row per raw sheet
def raw_col(key, ticker):                                       # data col letter on a raw sheet
    return L(2 + RAW_TICKERS[key].index(ticker))
def raw_nt(key, tenor):
    return L(2 + len(RAW_TICKERS[key]) + (0 if tenor == 10 else 1))
def rng(key, col):
    return f"'{RAWSH[key]}'!${col}$4:${col}${RAW_LAST[key]}"
def drng(key):
    return f"'{RAWSH[key]}'!$A$4:$A${RAW_LAST[key]}"

A = 'Arial'
F_T   = Font(name=A, size=13, bold=True, color='1A2226')
F_H   = Font(name=A, size=9,  bold=True, color='FFFFFF')
F_B   = Font(name=A, size=10)
F_IN  = Font(name=A, size=11, bold=True, color='0000FF')   # user-editable input
F_SRC = Font(name=A, size=10, color='0000FF')              # source data
F_C   = Font(name=A, size=10, color='000000')              # formula
F_K   = Font(name=A, size=10, bold=True, color='000000')
F_N   = Font(name=A, size=9,  color='595959')
FH    = PatternFill('solid', fgColor='2F6F74')
FIN   = PatternFill('solid', fgColor='FFFF00')             # cells the user edits
FKEY  = PatternFill('solid', fgColor='EAF2F2')
FWARN = PatternFill('solid', fgColor='FFF4E0')
BOX   = Border(*[Side(style='thin', color='BFBFBF')]*4)
PCT, PCT3, DT, INT = '0.00%', '0.000%', 'yyyy-mm-dd', '0'
WRAP = Alignment(wrap_text=True, vertical='top')

wb = Workbook()

def head(ws, row, labels, widths):
    for j, s in enumerate(labels, 1):
        c = ws.cell(row, j, s); c.font, c.fill = F_H, FH
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 32
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[L(j)].width = w

def put(ws, r, c, v, nf=None, font=F_C, fill=None):
    cell = ws.cell(r, c)
    if v is not None and not (isinstance(v, float) and pd.isna(v)):
        cell.value = float(v) if isinstance(v, (np.floating, np.integer)) else v
    if nf: cell.number_format = nf
    cell.font = font
    if fill: cell.fill = fill
    return cell

# ============================================================ raw sheets (only hardcoded data)
for key, title_txt in [('infra','Infrastructure & logistics — real equity IRR (daily)'),
                       ('util','Electric utilities, water & sanitation — real equity IRR (daily)'),
                       ('malls','Malls / income properties — real equity IRR (weekly)')]:
    df, tk = P[key], RAW_TICKERS[key]
    ws = wb.create_sheet(RAWSH[key])
    ws.cell(1,1,title_txt).font = F_T
    ws.cell(2,1,'Source data exactly as delivered by XP. This is the ONLY hardcoded data in the workbook — '
                'everything else is formulas. Paste a refreshed extract here and the whole file updates.').font = F_N
    head(ws, 3, ['Date'] + tk + ['NTN-B 10y','NTN-B 15y'], [12] + [11]*len(tk) + [11,11])
    ws.freeze_panes = 'B4'
    for i, (dt, row) in enumerate(df.iterrows(), 4):
        put(ws, i, 1, dt.to_pydatetime(), DT, F_SRC)
        for j, col in enumerate(tk + ['NTNB10','NTNB15'], 2):
            put(ws, i, j, row.get(col), PCT3, F_SRC)

# ============================================================ Inputs
ws = wb.create_sheet(SH['inp'])
ws.cell(1,1,'Inputs — every number in this workbook is driven from this sheet').font = F_T
ws.cell(2,1,'Edit only the YELLOW cells. Every other sheet recalculates automatically.').font = F_N
for col, w in [('A',3),('B',34),('C',16),('D',12),('E',62)]:
    ws.column_dimensions[col].width = w

R_PARAM = 5
ws.cell(4,2,'ANALYSIS PARAMETERS').font = F_K
PARAMS = [
    ('Analysis start date', pd.Timestamp('2022-12-01').to_pydatetime(), DT, 'WinStart',
     'Everything before this date is ignored: panel, index and percentiles. Default 2022-12-01 cuts the '
     'stretch where the chain-link rests on only 2 mall names. Set 2019-02-01 to see the full history.'),
    ('NTN-B tenor (10 or 15)', 10, INT, 'Tenor',
     'Which sovereign real yield to benchmark against. Drives every spread in the file.'),
    ('Expected inflation', 0.04, PCT, 'Infl',
     'Used only to convert the real IRRs into nominal terms on Company Summary. Does not affect spreads.'),
    ('Real hurdle rate (IPCA + x)', 0.08, PCT, 'Hurdle',
     'The bar a stock must clear. 8% reproduces the "Can equity IRRs compete with IPCA+8%?" question.'),
    ('Min names for a reliable link', 8, INT, 'MinNames',
     'If fewer companies than this carry a month of the chain-link, that month is flagged "Thin".'),
]
for i, (lab, val, nf, nm, note) in enumerate(PARAMS):
    r = R_PARAM + i
    put(ws, r, 2, lab, None, F_B)
    c = put(ws, r, 3, val, nf, F_IN, FIN); c.border = BOX
    n = put(ws, r, 5, note, None, F_N); n.alignment = WRAP
    ws.row_dimensions[r].height = 30
    wb.defined_names.add(DefinedName(nm, attr_text=f"'{SH['inp']}'!$C${r}"))

R_CO = 13
ws.cell(R_CO-2, 2, 'COMPANY COVERAGE').font = F_K
ws.cell(R_CO-2, 5, 'Set Include to No to drop a name from the index, the medians and every total. '
                   'Santos Brasil ships as No: its series ends Nov-2025 (take-private).').font = F_N
for j, s in enumerate(['Ticker','Company','Sector','Include','Source sheet'], 1):
    c = ws.cell(R_CO-1, j+1 if j > 1 else 1, s)
for j, (s, col) in enumerate([('Ticker',1),('Company',2),('Sector',3),('Include',4),('Source sheet',5)]):
    c = ws.cell(R_CO-1, col, s); c.font, c.fill = F_H, FH
    c.alignment = Alignment(horizontal='center', vertical='center')
ws.column_dimensions['A'].width = 10
dv_yn = DataValidation(type='list', formula1='"Yes,No"', allow_blank=False)
ws.add_data_validation(dv_yn)
dv_ten = DataValidation(type='list', formula1='"10,15"', allow_blank=False)
ws.add_data_validation(dv_ten); dv_ten.add(ws.cell(R_PARAM+1, 3))
for i, t in enumerate(ORDER):
    r = R_CO + i
    put(ws, r, 1, t, None, F_K)
    put(ws, r, 2, NAMES[t], None, F_B)
    put(ws, r, 3, SECTOR[t], None, F_B)
    c = put(ws, r, 4, 'No' if t == 'STBP3' else 'Yes', None, F_IN, FIN)
    c.border = BOX; c.alignment = Alignment(horizontal='center'); dv_yn.add(c)
    put(ws, r, 5, RAWSH[SRC[t]], None, F_N)
INC = {t: f"'{SH['inp']}'!$D${R_CO + i}" for i, t in enumerate(ORDER)}
ws.freeze_panes = 'A4'

# ============================================================ Panel (monthly)
months = pd.date_range('2019-02-28', '2026-07-31', freq='ME')
NM = len(months)
P_FIRST, P_LAST = 4, 3 + NM
ws = wb.create_sheet(SH['pan'])
ws.cell(1,1,'Monthly panel — average real IRR per company per calendar month').font = F_T
ws.cell(2,1,'AVERAGEIFS over the raw sheets. A cell is blank when the name is outside the analysis window, '
            'excluded on Inputs, or had no data that month.').font = F_N
head(ws, 3, ['Month end'] + ORDER + ['NTN-B 10y','NTN-B 15y'], [12] + [10]*len(ORDER) + [11,11])
ws.freeze_panes = 'B4'
for i, dt in enumerate(months):
    r = P_FIRST + i
    put(ws, r, 1, dt.to_pydatetime(), DT, F_SRC)
    win = f'EOMONTH($A{r},-1)+1'
    for j, t in enumerate(ORDER, 2):
        k = SRC[t]
        avg = (f'AVERAGEIFS({rng(k, raw_col(k,t))},{drng(k)},">="&{win},{drng(k)},"<="&$A{r})')
        put(ws, r, j, f'=IF(OR($A{r}<WinStart,{INC[t]}<>"Yes"),"",IFERROR({avg},""))', PCT3)
    for j, ten in [(2+len(ORDER), 10), (3+len(ORDER), 15)]:
        parts = [f'AVERAGEIFS({rng(k, raw_nt(k,ten))},{drng(k)},">="&{win},{drng(k)},"<="&$A{r})'
                 for k in ('infra','util','malls')]
        nested = f'IFERROR({parts[0]},IFERROR({parts[1]},IFERROR({parts[2]},"")))'
        put(ws, r, j, f'=IF($A{r}<WinStart,"",{nested})', PCT3)
CO_FIRST, CO_LAST = L(2), L(1 + len(ORDER))
NT10, NT15 = L(2 + len(ORDER)), L(3 + len(ORDER))

# ============================================================ Chain-link deltas
ws = wb.create_sheet(SH['dl'])
ws.cell(1,1,'Month-over-month change in IRR, per company').font = F_T
ws.cell(2,1,'Populated only where the company has a value in BOTH months, so a name entering or leaving '
            'coverage never moves the index. The row average drives the chain-link.').font = F_N
head(ws, 3, ['Month end'] + ORDER + ['Avg. change','Paired names'], [12] + [10]*len(ORDER) + [12,12])
ws.freeze_panes = 'B4'
for i in range(NM):
    r = P_FIRST + i
    put(ws, r, 1, f"='{SH['pan']}'!A{r}", DT, F_SRC)
    for j in range(2, 2 + len(ORDER)):
        c = L(j)
        if i == 0:
            put(ws, r, j, '=""', PCT3)
        else:
            put(ws, r, j, f'=IF(OR(\'{SH["pan"]}\'!{c}{r}="",\'{SH["pan"]}\'!{c}{r-1}=""),"",'
                          f'\'{SH["pan"]}\'!{c}{r}-\'{SH["pan"]}\'!{c}{r-1})', PCT3)
    rowr = f'{CO_FIRST}{r}:{CO_LAST}{r}'
    put(ws, r, 2+len(ORDER), f'=IF(COUNT({rowr})=0,"",AVERAGE({rowr}))', PCT3, F_K, FKEY)
    put(ws, r, 3+len(ORDER), f'=COUNT({rowr})', INT)
D_AVG, D_N = L(2+len(ORDER)), L(3+len(ORDER))

# ============================================================ IRR Index
ws = wb.create_sheet(SH['idx'])
ws.cell(1,1,'Consolidated equity IRR index — all included companies').font = F_T
ws.cell(2,1,'Column E is the headline series: chain-linked and composition-neutral. Column J flags months '
            'carried by too few names to trust.').font = F_N
head(ws, 3, ['Month end','Companies','Simple average','Simple median','Chain-linked IRR index',
             'Avg. MoM change','Paired names','NTN-B (real)','Index spread vs NTN-B','Reliability'],
     [12,11,13,13,16,13,11,13,16,11])
ws.freeze_panes = 'B4'
for i in range(NM):
    r = P_FIRST + i
    rowr = f"'{SH['pan']}'!{CO_FIRST}{r}:{CO_LAST}{r}"
    put(ws, r, 1, f"='{SH['pan']}'!A{r}", DT, F_SRC)
    put(ws, r, 2, f'=COUNT({rowr})', INT)
    put(ws, r, 3, f'=IF($B{r}=0,"",AVERAGE({rowr}))', PCT)
    put(ws, r, 4, f'=IF($B{r}=0,"",MEDIAN({rowr}))', PCT)
    if i == NM - 1:
        put(ws, r, 5, f'=IF($C{r}="","",$C{r})', PCT, F_K, FKEY)
    else:
        put(ws, r, 5, f'=IF(OR($C{r}="",$E{r+1}="",$F{r+1}=""),"",$E{r+1}-$F{r+1})', PCT, F_K, FKEY)
    put(ws, r, 6, f"='{SH['dl']}'!{D_AVG}{r}", PCT3)
    put(ws, r, 7, f"='{SH['dl']}'!{D_N}{r}", INT)
    put(ws, r, 8, f'=IFERROR(IF(Tenor=10,\'{SH["pan"]}\'!{NT10}{r},\'{SH["pan"]}\'!{NT15}{r})+0,"")', PCT)
    put(ws, r, 9, f'=IF(OR($E{r}="",$H{r}=""),"",$E{r}-$H{r})', PCT)
    put(ws, r,10, f'=IF($G{r}="","",IF($G{r}=0,"",IF($G{r}<MinNames,"Thin","OK")))', None, F_N)

ch = LineChart()
ch.title = 'Consolidated equity IRR index vs NTN-B (real)'
ch.y_axis.numFmt = '0%'; ch.height, ch.width = 10, 26
ch.add_data(Reference(ws, min_col=5, max_col=5, min_row=3, max_row=P_LAST), titles_from_data=True)
ch.add_data(Reference(ws, min_col=8, max_col=8, min_row=3, max_row=P_LAST), titles_from_data=True)
ch.set_categories(Reference(ws, min_col=1, min_row=4, max_row=P_LAST))
ch.series[0].graphicalProperties.line.width = 24000
ch.series[1].graphicalProperties.line.width = 14000
ch.series[1].graphicalProperties.line.dashStyle = 'dash'
for s in ch.series: s.smooth = False; s.marker.symbol = 'none'
ws.add_chart(ch, 'L4')

# ============================================================ Company Summary
ws = wb.create_sheet(SH['sm'])
ws.cell(1,1,"Company summary — current IRR against each name's own history").font = F_T
ws.cell(2,1,'Spot IRR and dates come from the raw sheets; the percentile distribution comes from the monthly '
            'panel, so it respects the analysis window on Inputs.').font = F_N
COLS = ['Ticker','Company','Sector','Included','Obs. (raw)','First date','Last date','Current IRR (spot)',
        'P10','P25','Median','P75','P90','Current percentile','Months in window','NTN-B (paired)',
        'Spread vs NTN-B','Nominal IRR','Spread vs hurdle','Beats hurdle?','_inc IRR','_inc spread','_top q']
head(ws, 3, COLS, [9,21,17,9,9,11,11,12,9,9,9,9,9,12,10,12,13,11,12,10,10,10,8])
ws.freeze_panes = 'C4'
S_FIRST = 4
for i, t in enumerate(ORDER):
    r = S_FIRST + i
    k = SRC[t]
    rc, dc = rng(k, raw_col(k, t)), drng(k)
    pc = f"'{SH['pan']}'!{L(2+i)}${P_FIRST}:{L(2+i)}${P_LAST}"
    nt10, nt15 = rng(k, raw_nt(k,10)), rng(k, raw_nt(k,15))
    warn = (t == 'STBP3')
    fl = FWARN if warn else None
    put(ws, r, 1, t, None, F_K, fl)
    put(ws, r, 2, NAMES[t], None, F_B, fl)
    put(ws, r, 3, SECTOR[t], None, F_B, fl)
    put(ws, r, 4, f'={INC[t]}', None, F_C, fl)
    put(ws, r, 5, f'=COUNT({rc})', INT, F_C, fl)
    put(ws, r, 6, f'=INDEX({dc},MATCH(TRUE,INDEX({rc}<>"",0),0))', DT, F_C, fl)
    put(ws, r, 7, f'=LOOKUP(2,1/({rc}<>""),{dc})', DT, F_C, fl)
    put(ws, r, 8, f'=LOOKUP(2,1/({rc}<>""),{rc})', PCT, F_K, fl)
    for j, q in enumerate([0.10, 0.25, 0.50, 0.75, 0.90]):
        put(ws, r, 9+j, f'=IF(COUNT({pc})=0,"",PERCENTILE({pc},{q}))', PCT, F_C, fl)
    put(ws, r, 14, f'=IF(COUNT({pc})=0,"",COUNTIF({pc},"<"&$H{r})/COUNT({pc}))', '0%', F_C, fl)
    put(ws, r, 15, f'=COUNT({pc})', INT, F_C, fl)
    # last NON-BLANK NTN-B on or before this name's own last date. INDEX/MATCH alone returns 0
    # when that exact row happens to have an empty yield cell (IGTI11 does).
    put(ws, r, 16,
        f'=IF(Tenor=10,LOOKUP(2,1/(({dc}<=$G{r})*({nt10}<>"")),{nt10}),'
        f'LOOKUP(2,1/(({dc}<=$G{r})*({nt15}<>"")),{nt15}))', PCT, F_C, fl)
    put(ws, r, 17, f'=$H{r}-$P{r}', PCT, F_K, fl)
    put(ws, r, 18, f'=(1+$H{r})*(1+Infl)-1', PCT, F_C, fl)
    put(ws, r, 19, f'=$H{r}-Hurdle', PCT, F_C, fl)
    put(ws, r, 20, f'=IF($H{r}>Hurdle,"Yes","No")', None, F_C, fl)
    put(ws, r, 21, f'=IF($D{r}="Yes",$H{r},"")', PCT, F_N, fl)
    put(ws, r, 22, f'=IF($D{r}="Yes",$Q{r},"")', PCT, F_N, fl)
    put(ws, r, 23, f'=IF(AND($D{r}="Yes",$N{r}<>"",$N{r}>=0.75),1,0)', INT, F_N, fl)
S_LAST = S_FIRST + len(ORDER) - 1
for c in ('U','V','W'):
    ws.column_dimensions[c].outline_level = 1
    ws.column_dimensions[c].hidden = True

r0 = S_LAST + 2
TOT = [('Included companies',      f'=COUNTIF($D${S_FIRST}:$D${S_LAST},"Yes")', INT),
       ('Median IRR (included)',   f'=MEDIAN($U${S_FIRST}:$U${S_LAST})', PCT),
       ('Median spread vs NTN-B',  f'=MEDIAN($V${S_FIRST}:$V${S_LAST})', PCT),
       ('Names with positive spread', f'=COUNTIF($V${S_FIRST}:$V${S_LAST},">0")', INT),
       ('Names beating the hurdle',
        f'=SUMPRODUCT(($D${S_FIRST}:$D${S_LAST}="Yes")*($T${S_FIRST}:$T${S_LAST}="Yes"))', INT),
       ('Names in their own top quartile', f'=SUM($W${S_FIRST}:$W${S_LAST})', INT)]
for i, (lab, f, nf) in enumerate(TOT):
    put(ws, r0+i, 2, lab, None, F_K)
    put(ws, r0+i, 8, f, nf, F_K, FKEY)

# ============================================================ Read me
ws = wb.create_sheet(SH['read'], 0)
for col, w in [('A',3),('B',26),('C',104)]:
    ws.column_dimensions[col].width = w
ws.cell(1,1,'Equity IRR — consolidated coverage (formula-driven)').font = F_T
ws.cell(2,1,'All IRRs are REAL (inflation-adjusted) and compare directly with the NTN-B real yield.').font = F_N
def sect(r, h, items, ht=42):
    ws.cell(r,2,h).font = F_K; r += 1
    for a,b in items:
        ws.cell(r,2,a).font = F_B
        c = ws.cell(r,3,b); c.font = F_B; c.alignment = WRAP
        ws.row_dimensions[r].height = ht; r += 1
    return r+1
r = sect(4, 'HOW TO USE IT', [
    ('Inputs', 'The only sheet you edit. Yellow cells only. Change the window, the NTN-B tenor, the hurdle, '
               'expected inflation, or switch any company on and off — everything else recalculates.'),
    ('Refreshing the data', 'Paste a newer extract over the three raw sheets, keeping the column layout. '
                            'Add rows at the bottom and widen the ranges if the history grows.'),
    ('Colour key', 'YELLOW = you edit this. BLUE = source data from XP. BLACK = calculated, do not overwrite.'),
], ht=44)
r = sect(r, 'SHEETS', [
    (SH['idx'], 'The consolidated index, one row per month. Column E is the headline, column J flags thin months.'),
    (SH['sm'],  'One row per company: spot IRR, its own percentile distribution, spread over the NTN-B and over the hurdle.'),
    (SH['pan'], 'Monthly average IRR per company — the grid behind everything.'),
    (SH['dl'],  'Month-over-month change per company, only where the name exists in both months.'),
    ('Raw sheets', 'The three XP extracts, unmodified. The only hardcoded numbers in the file.'),
], ht=30)
r = sect(r, 'WHY THE INDEX IS CHAIN-LINKED', [
    ('The problem', 'Coverage starts at different dates: malls from Feb-2019, infra from Dec-2022, utilities only '
                    'from Jul-2025. A plain average jumps whenever the roster changes rather than when anything reprices.'),
    ('The fix', 'The index moves by the average month-over-month change of only those names present in BOTH months, '
                'cumulated. A name entering or leaving never moves the level.'),
    ('The anchor', 'The chain is anchored so the LAST month equals the actual average of the names trading then, and '
                   'extended backwards. The current level is a real number; the history is the composition-neutral path to it.'),
    ('Index vs simple average', 'With a fixed roster the two move in perfect lockstep and the gap between them is '
                                'constant. The gap only steps at composition changes, and it measures how much the '
                                'coverage of that date differs from today\'s.'),
])
r = sect(r, 'LIMITS WORTH KNOWING', [
    ('Why the window defaults to Dec-2022',
     'Before Dec-2022 the chain rests on just 2 mall names for 46 months. Propagating the standard error of each '
     'monthly step gives a band of roughly +/-7.7 p.p. on the Feb-2019 level — the early history is extrapolation, '
     'not measurement. Set the window to 2019-02-01 if you want to see it, but do not quote its level.'),
    ('Percentile windows differ by name',
     'Utilities only have data from Jul-2025, so their percentiles rest on far fewer months than infra or malls. '
     'Check "Months in window" on Company Summary before quoting one.'),
    ('Santos Brasil', 'Series ends Nov-2025 (take-private). Ships as Include = No.'),
    ('NTN-B differs slightly between files',
     'Each XP file carries its own NTN-B column and they disagree by about 0.15 p.p. at the current snapshot. '
     'Company Summary pairs each name with its own file; the index prefers infra, then utilities, then malls. '
     'Normalising to one source moves no name more than 1 rank and flips no sign.'),
    ('Malls NTN-B in 2022', 'That column dips to 3-4% for four months in Q2-2022, roughly 2.3 p.p. below the other '
                            'files. Only affects the index spread if you widen the window back that far.'),
    ('Spreads are percentage points', 'Formatted as % but they are differences between two rates.'),
    ('Not in the source data', 'No dividend / growth / re-rating decomposition and no exit-multiple sensitivity exist '
                               'in these files, so neither can be built here.'),
])
r = sect(r, 'SOURCES', [
    ('Infra', 'https://researchxp1.s3.sa-east-1.amazonaws.com/Infra+Historical+IRRs.xlsx'),
    ('Income Properties', 'https://researchxp1.s3.sa-east-1.amazonaws.com/Income+Properties+-+Historical+IRR+(Weekly).xlsb'),
    ('Utilities', 'https://researchxp1.s3.sa-east-1.amazonaws.com/Utilities+Historical+IRR.xlsx'),
], ht=18)

del wb['Sheet']
wb._sheets = [wb[SH[k]] for k in ['read','inp','idx','sm','pan','dl','infra','util','mall']]
OUT = 'Equity IRR - Consolidated Coverage (Auto).xlsx'
wb.save(OUT)
print('saved', OUT)
print('sheets:', wb.sheetnames)
print('months:', NM, '| panel rows', P_FIRST, '-', P_LAST, '| summary rows', S_FIRST, '-', S_LAST)
